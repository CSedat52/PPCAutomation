"""
Amazon PPC Analyst Agent (Agent 2) — v5 Multi-Account
=======================================================
DEGISIKLIKLER (v4 -> v5):
  1. Multi-account: hesap_key + marketplace parametresi alir.
  2. Tum path'ler init_paths() ile dinamik olusturulur.
  3. Veri okuma: data/{hesap_key}_{marketplace}/ altindan.
  4. Config okuma: config/{hesap_key}_{marketplace}/ altindan.
  5. Cikti yazma: data/{hesap_key}_{marketplace}/analysis/ ve decisions/ altina.
  6. CLI: python agent2/analyst.py <hesap_key> <marketplace>

DEGISIKLIKLER (v3 -> v4):
  1. save_error_log(): Yeni fonksiyon — tum beklenmeyen hatalar ve preflight
     hatalari data/logs/agent2_errors.json dosyasina eklenir.
  2. run_analysis() artik _run_analysis_impl()'i try/except ile sarmaliyor.
     Beklenmeyen exception'lar yakalanip loglanir, agent graceful bicimde
     BASARISIZ donus yapar.
  3. Agent 4 (Learning Agent) agent2_errors.json'u okuyarak hata kaliplarini
     analiz eder ve tekrar eden hatalara cozum onerisi uretir.
  4. Son 200 hata kaydi tutulur (eski kayitlar otomatik temizlenir).

DEGISIKLIKLER (v2 -> v3):
  1. ASIN bazli tanh parametreleri: bid_functions.json > asin_parametreleri
  2. get_tanh_params_for_asin(): oncelik: ASIN ozel (aktif=True) → global
  3. calculate_new_bid(): param_kaynagi sebep string'ine eklendi

DUZELTMELER (v1 -> v2):
  1. SP kolon isimleri: sales7d → sales14d, purchases7d → purchases14d
  2. SB bid eslestirme: keywordBid alani
  3. SD bid eslestirme: targetId ile eslestirme
  4. Portfolio bilgisi: campaign entity listesinden portfolioId
  5. Para birimi: marketplace'e gore
  6. Excel formatlama
  7. Sutun ve satir sirasi duzenlemeleri
  8. ASIN ve Orders sutunlari kaldirildi
  9. Harvesting'de purchases14d

Desteklenen Reklam Tipleri: SP, SB, SD
"""

import os
import sys
import json
import math
import logging
import traceback
from datetime import datetime
from pathlib import Path
from copy import deepcopy

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("agent2_analyst")
sys.path.insert(0, str(Path(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))))
from log_utils import save_error_log as _central_save_error_log, save_log as _save_log


def _dashboard_status(agent_name, status, health_detail=None):
    """Dashboard agent_status tablosunu gunceller."""
    try:
        from supabase.db_client import SupabaseClient
        db = SupabaseClient()
        db.update_agent_status_detail(agent_name, status, health_detail)
    except Exception as e:
        logger.warning("Supabase yazim hatasi: %s", e)


def _dashboard_pipeline(session_id, hesap_key, marketplace, step, status, error_msg=None):
    """Dashboard pipeline_runs tablosunu gunceller."""
    if not session_id:
        return
    try:
        from supabase.db_client import SupabaseClient
        db = SupabaseClient()
        db.upsert_pipeline_run(session_id, hesap_key, marketplace, step, status, error_msg)
    except Exception as e:
        logger.warning("Supabase yazim hatasi: %s", e)

# ============================================================================
# DOSYA YOLLARI — hesap_key + marketplace'den dinamik olusturulur
# ============================================================================

BASE_DIR = Path(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Bu degiskenler init_paths() ile set edilir (run_analysis basinda cagirilir)
CONFIG_DIR = None
DATA_DIR = None
ANALYSIS_DIR = None
DECISIONS_DIR = None
SETTINGS_FILE = None
BID_FUNCTIONS_FILE = None

# Maestro pipeline session ID'si (env var ile iletilir, korelasyon icin)
MAESTRO_SESSION_ID = os.environ.get("MAESTRO_SESSION_ID")


def init_paths(hesap_key, marketplace):
    """Hesap+marketplace icin tum path'leri set eder."""
    global CONFIG_DIR, DATA_DIR, ANALYSIS_DIR, DECISIONS_DIR
    global SETTINGS_FILE, BID_FUNCTIONS_FILE

    dir_name = f"{hesap_key}_{marketplace}"
    CONFIG_DIR = BASE_DIR / "config" / dir_name
    DATA_DIR = BASE_DIR / "data" / dir_name
    ANALYSIS_DIR = DATA_DIR / "analysis"
    DECISIONS_DIR = DATA_DIR / "decisions"
    SETTINGS_FILE = CONFIG_DIR / "settings.json"
    BID_FUNCTIONS_FILE = CONFIG_DIR / "bid_functions.json"

    logger.info("Paths: data=%s, config=%s", DATA_DIR, CONFIG_DIR)

# Para birimi haritasi
CURRENCY_MAP = {
    "US": "$", "CA": "CA$", "UK": "£", "DE": "€", "FR": "€",
    "ES": "€", "IT": "€", "JP": "¥", "AU": "A$", "MX": "MX$",
    "SE": "SEK", "PL": "PLN", "NL": "€",
}


# ============================================================================
# AYAR YUKLEYICI
# ============================================================================

def load_settings():
    """Settings'i Supabase'den yukler. Basarisizsa JSON dosyasina fallback."""
    hk = os.environ.get("HESAP_KEY", "")
    mp = os.environ.get("MARKETPLACE", "")
    if hk and mp:
        try:
            from supabase.db_client import SupabaseClient
            db = SupabaseClient()
            conn = db._conn()
            cur = conn.cursor()
            cur.execute("SELECT genel_ayarlar, esik_degerleri, asin_hedefleri, segmentasyon_kurallari, agent3_ayarlari, ozel_kurallar, negatif_keyword_kurali, yeni_keyword_kurali, harvesting_ayarlari FROM settings WHERE hesap_key = %s AND marketplace = %s", (hk, mp))
            row = cur.fetchone()
            cur.close()
            conn.close()
            if row:
                result = {}
                keys = ["genel_ayarlar", "esik_degerleri", "asin_hedefleri", "segmentasyon_kurallari", "agent3_ayarlari", "ozel_kurallar", "negatif_keyword_kurali", "yeni_keyword_kurali", "harvesting_ayarlari"]
                for i, key in enumerate(keys):
                    if row[i]:
                        result[key] = row[i] if isinstance(row[i], dict) else json.loads(row[i])
                logger.info("Settings Supabase'den yuklendi (%s/%s)", hk, mp)
                return result
        except Exception as e:
            logger.warning("Settings Supabase'den okunamadi, dosyaya fallback: %s", e)

    if SETTINGS_FILE and SETTINGS_FILE.exists():
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    raise FileNotFoundError(f"Settings bulunamadi: Supabase ve {SETTINGS_FILE}")


def load_bid_functions():
    """Bid functions'i Supabase'den yukler. Basarisizsa JSON dosyasina fallback."""
    hk = os.environ.get("HESAP_KEY", "")
    mp = os.environ.get("MARKETPLACE", "")
    if hk and mp:
        try:
            from supabase.db_client import SupabaseClient
            db = SupabaseClient()
            conn = db._conn()
            cur = conn.cursor()
            cur.execute("SELECT tanh_formulu, segment_parametreleri, genel_limitler, asin_parametreleri FROM bid_functions WHERE hesap_key = %s AND marketplace = %s", (hk, mp))
            row = cur.fetchone()
            cur.close()
            conn.close()
            if row:
                result = {}
                if row[0]: result["tanh_formulu"] = row[0] if isinstance(row[0], dict) else json.loads(row[0])
                if row[1]: result["segment_parametreleri"] = row[1] if isinstance(row[1], dict) else json.loads(row[1])
                if row[2]: result["genel_limitler"] = row[2] if isinstance(row[2], dict) else json.loads(row[2])
                if row[3]: result["asin_parametreleri"] = row[3] if isinstance(row[3], dict) else json.loads(row[3])
                logger.info("Bid functions Supabase'den yuklendi (%s/%s)", hk, mp)
                return result
        except Exception as e:
            logger.warning("Bid functions Supabase'den okunamadi, dosyaya fallback: %s", e)

    if BID_FUNCTIONS_FILE and BID_FUNCTIONS_FILE.exists():
        with open(BID_FUNCTIONS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    raise FileNotFoundError(f"Bid functions bulunamadi: Supabase ve {BID_FUNCTIONS_FILE}")


def get_hedef_acos(settings, asin):
    hedefler = settings.get("asin_hedefleri", {})
    if asin in hedefler and isinstance(hedefler[asin], dict):
        return hedefler[asin].get("hedef_acos", hedefler.get("_varsayilan_acos", 25))
    return hedefler.get("_varsayilan_acos", 25)


def get_currency(settings):
    marketplace = settings.get("genel_ayarlar", {}).get("aktif_marketplace", "US")
    return CURRENCY_MAP.get(marketplace, "$")


# ============================================================================
# VERI YUKLEYICI
# ============================================================================

def find_latest_data_file(prefix):
    if not DATA_DIR.exists():
        return None
    candidates = sorted(DATA_DIR.glob(f"*_{prefix}.json"), reverse=True)
    return candidates[0] if candidates else None


def load_json_file(filepath):
    if filepath is None or not Path(filepath).exists():
        return []
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)


def load_all_agent1_data():
    """
    Agent 1'in tum dosyalarini yukler ve organize eder.
    Supabase-first: data_loader uzerinden Supabase raw_data'dan okur.
    JSON fallback: Supabase basarisiz olursa JSON dosyalarina doner.
    """
    hk = os.environ.get("HESAP_KEY", "")
    mp = os.environ.get("MARKETPLACE", "")

    if hk and mp:
        try:
            from data_loader import load_all_agent1_data as _load_supabase
            today = datetime.utcnow().strftime("%Y-%m-%d")
            data = _load_supabase(hk, mp, str(DATA_DIR), today)
            # Kritik kontrol: SP kampanya verisi var mi?
            if data.get("sp", {}).get("campaigns"):
                logger.info("Agent 1 verileri Supabase'den yuklendi (data_loader)")
                return data
            logger.warning("Supabase'den SP kampanya verisi bos, JSON fallback")
        except Exception as e:
            logger.warning("data_loader basarisiz, JSON fallback: %s", e)

    # JSON fallback (eski yontem)
    data = {
        "portfolios": load_json_file(find_latest_data_file("portfolios")),
        "sp": {
            "campaigns": load_json_file(find_latest_data_file("sp_campaigns")),
            "keywords": load_json_file(find_latest_data_file("sp_keywords")),
            "targets": load_json_file(find_latest_data_file("sp_targets")),
            "negative_keywords": load_json_file(find_latest_data_file("sp_negative_keywords")),
            "campaign_negative_keywords": load_json_file(find_latest_data_file("sp_campaign_negative_keywords")),
            "negative_targets": load_json_file(find_latest_data_file("sp_negative_targets")),
        },
        "sb": {
            "campaigns": load_json_file(find_latest_data_file("sb_campaigns")),
            "keywords": load_json_file(find_latest_data_file("sb_keywords")),
            "targets": load_json_file(find_latest_data_file("sb_targets")),
            "negative_keywords": load_json_file(find_latest_data_file("sb_negative_keywords")),
        },
        "sd": {
            "campaigns": load_json_file(find_latest_data_file("sd_campaigns")),
            "targets": load_json_file(find_latest_data_file("sd_targets")),
        },
        "reports": {},
    }

    report_files = [
        ("sp_targeting_report", 14),
        ("sp_search_term_report", 30),
        ("sb_targeting_report", 14),
        ("sb_search_term_report", 30),
        ("sd_targeting_report", 14),
        ("sd_targeting_report", 30),
    ]

    for report_name, days in report_files:
        key = f"{report_name}_{days}d"
        filepath = find_latest_data_file(f"{report_name}_{days}d")
        data["reports"][key] = load_json_file(filepath)

    return data


# ============================================================================
# ONCEKI KARAR YUKLEYICI
# ============================================================================

def load_previous_decisions():
    """Onceki kararlari Supabase'den yukler. Fallback: JSON dosyasindan."""
    hk = os.environ.get("HESAP_KEY", "")
    mp = os.environ.get("MARKETPLACE", "")
    if hk and mp:
        try:
            from supabase.db_client import SupabaseClient
            db = SupabaseClient()
            conn = db._conn()
            cur = conn.cursor()
            # Son analiz tarihinden kararlari al
            cur.execute("""
                SELECT keyword_id, keyword_text, campaign_name, current_bid, recommended_bid,
                       bid_change_pct, segment, analysis_date, ad_type
                FROM bid_recommendations
                WHERE hesap_key = %s AND marketplace = %s
                ORDER BY analysis_date DESC LIMIT 5000
            """, (hk, mp))
            rows = cur.fetchall()
            cur.close()
            conn.close()
            if rows:
                result = {}
                for row in rows:
                    hid = str(row[0] or "")
                    if hid and hid not in result:
                        result[hid] = {
                            "hedefleme_id": hid,
                            "hedefleme": row[1] or "",
                            "kampanya": row[2] or "",
                            "onceki_bid": float(row[3] or 0),
                            "yeni_bid": float(row[4] or 0),
                            "degisim_yuzde": float(row[5] or 0),
                            "segment": row[6] or "",
                            "tarih": str(row[7] or ""),
                            "reklam_tipi": row[8] or "",
                        }
                logger.info("Onceki kararlar Supabase'den yuklendi: %d kayit", len(result))
                return result
        except Exception as e:
            logger.warning("Onceki kararlar Supabase'den okunamadi, dosyaya fallback: %s", e)

    if not DECISIONS_DIR or not DECISIONS_DIR.exists():
        return {}
    candidates = sorted(DECISIONS_DIR.glob("*_decisions.json"), reverse=True)
    if not candidates:
        return {}
    with open(candidates[0], "r", encoding="utf-8") as f:
        decisions = json.load(f)
    return {d["hedefleme_id"]: d for d in decisions}


# ============================================================================
# AKTIF KAMPANYA FILTRESI + PORTFOLIO HARITASI
# ============================================================================

def get_active_campaign_ids(campaigns):
    active = set()
    for c in campaigns:
        state = c.get("state", c.get("campaignStatus", c.get("status", ""))).upper()
        if state in ("ENABLED", "DELIVERING"):
            cid = str(c.get("campaignId", ""))
            if cid:
                active.add(cid)
    return active


def build_campaign_maps(data, settings=None):
    """Kampanya ID → isim, kampanya ID → portfolio ismi haritalarini olusturur.
    Portfolio ismi icin Agent 1'in cektigi portfolios entity listesi kullanilir.
    Response'tan gelen her portfolio: {"portfolioId": "...", "name": "...", ...}
    """
    camp_names = {}
    camp_portfolios = {}  # campaignId → portfolio NAME (veya ID)

    # Agent 1'den gelen portfolio listesi (v3 API response)
    portfolio_names = {}
    portfolios = data.get("portfolios", [])
    for p in portfolios:
        pid = str(p.get("portfolioId", ""))
        pname = p.get("name", "")
        if pid and pname:
            portfolio_names[pid] = pname

    # Settings'ten de ek/override portfolio isimleri alinabilir
    if settings:
        extra = settings.get("portfolio_isimleri", {})
        for k, v in extra.items():
            portfolio_names[str(k)] = v

    all_campaigns = (
        data["sp"]["campaigns"] +
        data["sb"]["campaigns"] +
        data["sd"]["campaigns"]
    )

    for c in all_campaigns:
        cid = str(c.get("campaignId", ""))
        camp_names[cid] = c.get("campaignName", c.get("name", "Bilinmiyor"))
        pid = c.get("portfolioId")
        if pid:
            pid_str = str(pid)
            # Portfolio ismi varsa isim, yoksa ID goster
            camp_portfolios[cid] = portfolio_names.get(pid_str, pid_str)

    return camp_names, camp_portfolios


# ============================================================================
# HEDEFLEME BIRLESTIRICISI — DUZELTILMIS KOLON ISIMLERI
# ============================================================================

def build_targeting_list(data, settings):
    """
    Tum reklam tiplerinden hedeflemeleri tek listeye birlestirir.
    KRITIK: SP → sales14d/purchases14d, SB/SD → sales/purchases
    """
    targets = []
    lookback = settings["genel_ayarlar"]["lookback_gun"]
    suffix = f"{lookback}d"

    sp_active = get_active_campaign_ids(data["sp"]["campaigns"])
    sb_active = get_active_campaign_ids(data["sb"]["campaigns"])
    sd_active = get_active_campaign_ids(data["sd"]["campaigns"])

    camp_names, camp_portfolios = build_campaign_maps(data, settings)

    # ----- SP TARGETING RAPORU -----
    # DUZELTME: sales14d, purchases14d kullaniliyor (sales7d DEGIL!)
    sp_report = data["reports"].get(f"sp_targeting_report_{suffix}", [])
    for row in sp_report:
        cid = str(row.get("campaignId", ""))
        if cid not in sp_active:
            continue
        targets.append(_build_target_row(
            "SP", row, camp_names.get(cid, ""), cid,
            camp_portfolios.get(cid, ""),
            sales_key="sales14d", orders_key="purchases14d",
        ))

    # ----- SB TARGETING RAPORU -----
    # SB: sales, purchases (window eki YOK)
    sb_report = data["reports"].get(f"sb_targeting_report_{suffix}", [])
    for row in sb_report:
        cid = str(row.get("campaignId", ""))
        if cid not in sb_active:
            continue
        targets.append(_build_target_row(
            "SB", row, camp_names.get(cid, ""), cid,
            camp_portfolios.get(cid, ""),
            sales_key="sales", orders_key="purchases",
        ))

    # ----- SD TARGETING RAPORU -----
    # SD: sales, purchases (window eki YOK)
    sd_report = data["reports"].get(f"sd_targeting_report_{suffix}", [])
    for row in sd_report:
        cid = str(row.get("campaignId", ""))
        if cid not in sd_active:
            continue
        targets.append(_build_target_row(
            "SD", row, camp_names.get(cid, ""), cid,
            camp_portfolios.get(cid, ""),
            sales_key="sales", orders_key="purchases",
        ))

    return targets


def _build_target_row(ad_type, row, camp_name, camp_id, portfolio_id,
                       sales_key="sales14d", orders_key="purchases14d"):
    """Tek hedefleme satirini standart formata donusturur.
    DUZELTME: Bid bilgisi dogrudan rapor satirindan alinir (keywordBid/bid alani).
    """
    impressions = int(float(row.get("impressions", 0)))
    clicks = int(float(row.get("clicks", 0)))
    spend = float(row.get("cost", 0))
    sales = float(row.get(sales_key, 0))
    orders = int(float(row.get(orders_key, 0)))

    acos = round((spend / sales * 100), 2) if sales > 0 else 0
    cvr = round((orders / clicks * 100), 2) if clicks > 0 else 0
    cpc = round(spend / clicks, 2) if clicks > 0 else 0

    # Hedefleme bilgisi
    hedefleme = (
        row.get("keyword") or
        row.get("keywordText") or
        row.get("targetingExpression") or
        row.get("targeting") or
        str(row.get("targetId", ""))
    )
    match_type = row.get("matchType", row.get("targetingType", ""))
    keyword_id = str(row.get("keywordId", row.get("targetId", "")))

    # DUZELTME: Bid bilgisi dogrudan rapor satirindan alinir
    # SB targeting raporunda keywordBid alani VAR (keywordId YOK)
    # SP targeting raporunda keywordBid alani olabilir
    mevcut_bid = float(row.get("keywordBid", row.get("bid", 0)) or 0)

    # Ad Group bilgisi
    ad_group_name = row.get("adGroupName", "")
    ad_group_id = str(row.get("adGroupId", ""))

    hedefleme_id = f"{ad_type}_{camp_id}_{keyword_id}_{match_type}"

    return {
        "hedefleme_id": hedefleme_id,
        "reklam_tipi": ad_type,
        "kampanya_adi": camp_name,
        "kampanya_id": camp_id,
        "portfolio_id": portfolio_id,
        "ad_group_name": ad_group_name,
        "ad_group_id": ad_group_id,
        "hedefleme": str(hedefleme),
        "match_type": match_type,
        "keyword_id": keyword_id,
        "impressions": impressions,
        "clicks": clicks,
        "spend": round(spend, 2),
        "sales": round(sales, 2),
        "orders": orders,
        "acos": acos,
        "cvr": cvr,
        "cpc": cpc,
        "mevcut_bid": mevcut_bid,
    }


# ============================================================================
# BID ESLESTIRME — DUZELTILMIS
# ============================================================================

def enrich_with_bid_data(targets, data):
    """
    Bid bilgisi rapor satirindan zaten aliniyor (_build_target_row icinde).
    Bu fonksiyon sadece bid=0 kalanlari keyword/target listesinden eslestirir (fallback).
    
    DUZELTME: SB'de keywordText ile eslestirme eklendi (keywordId raporda YOK).
    """
    # SP keyword bid haritasi: keywordId → bid
    sp_kw_bids = {}
    for kw in data["sp"]["keywords"]:
        kid = str(kw.get("keywordId", ""))
        bid = kw.get("bid", 0)
        if isinstance(bid, dict):
            bid = bid.get("amount", bid.get("value", 0))
        sp_kw_bids[kid] = float(bid) if bid else 0

    # SP target bid haritasi: targetId → bid
    sp_tgt_bids = {}
    for tgt in data["sp"]["targets"]:
        tid = str(tgt.get("targetId", ""))
        bid = tgt.get("bid", 0)
        if isinstance(bid, dict):
            bid = bid.get("amount", bid.get("value", 0))
        sp_tgt_bids[tid] = float(bid) if bid else 0

    # SB keyword bid haritasi: keywordText+matchType → bid (keywordId ile DEGIL!)
    sb_kw_bids_by_text = {}
    for kw in data["sb"]["keywords"]:
        text = kw.get("keywordText", "").lower()
        mt = kw.get("matchType", "").lower()
        bid = float(kw.get("bid", 0) or 0)
        sb_kw_bids_by_text[f"{text}|{mt}"] = bid

    # SD target bid haritasi: targetId → bid
    sd_tgt_bids = {}
    for tgt in data["sd"]["targets"]:
        tid = str(tgt.get("targetId", ""))
        bid = float(tgt.get("bid", 0) or 0)
        sd_tgt_bids[tid] = bid

    for t in targets:
        # Sadece bid=0 olanlari fallback ile doldur
        if t["mevcut_bid"] > 0:
            continue

        kid = t["keyword_id"]
        ad = t["reklam_tipi"]

        if ad == "SP":
            t["mevcut_bid"] = sp_kw_bids.get(kid, sp_tgt_bids.get(kid, 0))
        elif ad == "SB":
            # SB: keywordText + matchType ile esle
            text = t["hedefleme"].lower()
            mt = t["match_type"].lower()
            t["mevcut_bid"] = sb_kw_bids_by_text.get(f"{text}|{mt}", 0)
        elif ad == "SD":
            t["mevcut_bid"] = sd_tgt_bids.get(kid, 0)

    return targets


# ============================================================================
# ASIN ESLESTIRME
# ============================================================================

def enrich_with_asin(targets, data):
    for t in targets:
        t["asin"] = _extract_asin_from_campaign_name(t["kampanya_adi"])
    return targets


def _extract_asin_from_campaign_name(name):
    import re
    match = re.search(r'B0[A-Z0-9]{8,}', name.upper())
    return match.group(0) if match else ""


# ============================================================================
# SEGMENTASYON MOTORU
# ============================================================================

def segmentize(target, settings, previous_decisions):
    """
    Segmentasyon Akisi:
      1. Gosterim Kontrolu → GORUNMEZ
      2. Tiklama Yeterliligi → YETERSIZ_VERI
      3. Harcayan Ama Satmayan → KAN_KAYBEDEN
      4. ACoS Optimizasyonu → SUPER_STAR / TUZAK / KAZANAN / OPTIMIZE_ET / ZARAR
    """
    esikler = settings["esik_degerleri"]
    ozel = settings["ozel_kurallar"]
    gosterim_esik = esikler["gosterim_esik"]
    tiklama_esik = esikler["tiklama_esik"]

    imp = target["impressions"]
    clicks = target["clicks"]
    sales = target["sales"]
    orders = target["orders"]
    acos = target["acos"]
    cvr = target["cvr"]

    hedef_acos = get_hedef_acos(settings, target.get("asin", ""))

    # ---- EK KONTROL: Impression Takibi ----
    prev = previous_decisions.get(target["hedefleme_id"])
    if prev and ozel.get("impression_takibi", {}).get("aktif", True):
        prev_imp = prev.get("metrikler", {}).get("impressions", 0)
        prev_bid = prev.get("yeni_bid", prev.get("onceki_bid", 0))
        imp_esik = ozel["impression_takibi"]["impression_dusus_esik_yuzde"] / 100

        if prev_bid != prev.get("onceki_bid", 0) and prev_imp > gosterim_esik:
            imp_degisim = (imp - prev_imp) / prev_imp if prev_imp > 0 else 0
            if imp_degisim < -imp_esik:
                return "IMPRESSION_BEKLE", hedef_acos, \
                    f"Onceki bid degisikligi sonrasi impression {abs(imp_degisim)*100:.0f}% dustu. 1 periyot bekleniyor."

    # ADIM 1: Gosterim Kontrolu
    if imp < gosterim_esik:
        return "GORUNMEZ", hedef_acos, \
            f"Impression ({imp}) < esik ({gosterim_esik}). Bid artirarak gosterim kazandirilacak."

    # ADIM 2: Tiklama Yeterliligi
    if sales == 0 and clicks < tiklama_esik:
        return "YETERSIZ_VERI", hedef_acos, \
            f"Satis yok, tiklama ({clicks}) < esik ({tiklama_esik}). Veri yetersiz, bekleniyor."

    # ADIM 3: Harcayan Ama Satmayan
    if clicks >= tiklama_esik and sales == 0:
        return "KAN_KAYBEDEN", hedef_acos, \
            f"{clicks} tiklama, ${target['spend']:.2f} harcama, 0 satis. Bid dusuruluyor."

    # ADIM 4: ACoS Optimizasyonu
    if sales > 0:
        ss = ozel.get("super_star_koruma", {})
        if ss.get("aktif", True) and acos < ss.get("acos_esik", 10) and acos > 0:
            return "SUPER_STAR", hedef_acos, \
                f"ACoS {acos:.1f}% < %10 (Super Star). Korumali artis."

        cvr_kural = ozel.get("dusuk_cvr_tuzagi", {})
        if cvr_kural.get("aktif", True) and acos < hedef_acos and cvr < cvr_kural.get("cvr_esik", 1.0):
            return "TUZAK", hedef_acos, \
                f"ACoS {acos:.1f}% iyi ama CVR {cvr:.2f}% < %{cvr_kural.get('cvr_esik', 1.0)}. Dokunulmuyor."

        if acos < hedef_acos:
            return "KAZANAN", hedef_acos, \
                f"ACoS {acos:.1f}% < hedef %{hedef_acos}. Olcekleniyor."

        if acos < hedef_acos * 1.5:
            return "OPTIMIZE_ET", hedef_acos, \
                f"ACoS {acos:.1f}% > hedef %{hedef_acos}. Optimize ediliyor."

        return "ZARAR", hedef_acos, \
            f"ACoS {acos:.1f}% > hedef %{hedef_acos} x1.5. Sert dusus."

    return "YETERSIZ_VERI", hedef_acos, "Siniflandirilamadi, bekleniyor."


# ============================================================================
# DINAMIK BID HESAPLAMA
# ============================================================================

def get_tanh_params_for_asin(asin, bid_funcs):
    """
    ASIN bazli tanh parametrelerini yukler.

    Oncelik sirasi:
      1. Supabase asin_bid_params tablosunda aktif=True ise → ASIN parametreleri
      2. bid_functions.asin_parametreleri'nde varsa (fallback) → ASIN parametreleri
      3. Yoksa → tanh_formulu global degerleri

    Donus: (hassasiyet, max_degisim, parametre_kaynagi)
      parametre_kaynagi: "ASIN_OZEL" veya "GLOBAL" — log ve Excel raporuna yazilir.
    """
    global_params = bid_funcs.get("tanh_formulu", {})
    global_hass = global_params.get("hassasiyet", 0.8)
    global_max  = global_params.get("max_degisim", 0.20)

    if not asin:
        return global_hass, global_max, "GLOBAL"

    # Supabase asin_bid_params tablosundan oku (oncelikli)
    asin_entry = _get_asin_bid_params_cache().get(asin)

    # Fallback: bid_functions.asin_parametreleri (eski format)
    if not asin_entry:
        asin_params = bid_funcs.get("asin_parametreleri", {})
        asin_entry = asin_params.get(asin)

    if not asin_entry or not isinstance(asin_entry, dict):
        return global_hass, global_max, "GLOBAL"

    if not asin_entry.get("aktif", True):
        return global_hass, global_max, "GLOBAL (ASIN devre disi)"

    hass = asin_entry.get("hassasiyet", global_hass)
    maxd = asin_entry.get("max_degisim", global_max)
    return hass, maxd, "ASIN_OZEL"


# Cache: asin_bid_params Supabase'den bir kez okunur, session boyunca kullanilir
_asin_bid_params_cache = None

def _get_asin_bid_params_cache():
    """Supabase asin_bid_params tablosundan tum ASIN parametrelerini yukler ve cache'ler."""
    global _asin_bid_params_cache
    if _asin_bid_params_cache is not None:
        return _asin_bid_params_cache

    _asin_bid_params_cache = {}
    try:
        from supabase.db_client import SupabaseClient
        db = SupabaseClient()
        conn = db._conn()
        cur = conn.cursor()

        import os
        hk = os.environ.get("HESAP_KEY", "")
        mp = os.environ.get("MARKETPLACE", "")

        if hk and mp:
            cur.execute("""
                SELECT asin, aktif, hassasiyet, max_degisim, urun_adi
                FROM asin_bid_params
                WHERE hesap_key = %s AND marketplace = %s
            """, (hk, mp))
            for row in cur.fetchall():
                _asin_bid_params_cache[row[0]] = {
                    "aktif": row[1],
                    "hassasiyet": float(row[2]) if row[2] else 0.5,
                    "max_degisim": float(row[3]) if row[3] else 0.2,
                    "urun_adi": row[4],
                }
        cur.close()
        conn.close()
    except Exception as e:
        import logging
        logging.getLogger("analyst").warning("asin_bid_params Supabase'den okunamadi: %s", e)

    return _asin_bid_params_cache


def calculate_new_bid(segment, target, hedef_acos, settings, bid_funcs, previous_decisions):
    """
    Segment ve ASIN bazli parametrelerle yeni bid hesaplar.

    tanh parametreleri (hassasiyet, max_degisim):
      - Once bid_functions.json > asin_parametreleri'nde bu ASIN icin tanim var mi bakilir.
      - Varsa ve aktif=True ise ASIN ozel parametreler kullanilir.
      - Yoksa global tanh_formulu parametreleri kullanilir.
      - KAN_KAYBEDEN ve GORUNMEZ segmentleri tanh kullanmaz; bu ayrimi etkilemez.

    Guvenlik: genel_limitler her zaman uygulanir — ASIN ozel parametre bile asalamaz.
    """
    mevcut_bid = target["mevcut_bid"]
    asin       = target.get("asin", "")
    min_bid    = settings["genel_ayarlar"]["min_bid"]
    max_bid    = settings["genel_ayarlar"]["max_bid"]

    if mevcut_bid <= 0:
        return 0, 0, "Mevcut bid bilgisi bulunamadi"

    # ── Parametreleri yukle (ASIN ozel veya global) ──────────────────────
    hassasiyet, max_degisim, param_kaynagi = get_tanh_params_for_asin(asin, bid_funcs)

    seg_params     = bid_funcs.get("segment_parametreleri", {})
    genel_limit    = bid_funcs.get("genel_limitler", {})
    max_artis_limit = genel_limit.get("tek_seferde_max_artis", 0.30)
    max_dusus_limit = genel_limit.get("tek_seferde_max_dusus", 0.35)

    degisim = 0
    sebep   = ""

    # ── Segment bazli hesaplama ───────────────────────────────────────────
    if segment in ("YETERSIZ_VERI", "TUZAK", "IMPRESSION_BEKLE"):
        return mevcut_bid, 0, "Dokunulmadi"

    if segment == "GORUNMEZ":
        p = seg_params.get("GORUNMEZ", {})
        degisim = p.get("artis_orani", 0.10)
        sebep = f"Gosterim yok, bid +%{degisim*100:.0f}"

    elif segment == "KAN_KAYBEDEN":
        p   = seg_params.get("KAN_KAYBEDEN", {})
        min_d = p.get("min_dusus", 0.10)
        max_d = p.get("max_dusus", 0.30)
        ref   = p.get("referans_tiklama", 40)
        siddet = min(target["clicks"] / ref, 1.0)
        degisim = -(min_d + siddet * (max_d - min_d))
        sebep = f"{target['clicks']} tik 0 satis, bid {degisim*100:.1f}%"

    elif segment == "SUPER_STAR":
        p = seg_params.get("SUPER_STAR", {})
        acos      = target["acos"]
        acos_fark = (acos - hedef_acos) / hedef_acos if hedef_acos > 0 else 0
        degisim   = -math.tanh(acos_fark * hassasiyet) * max_degisim
        max_ss    = p.get("max_artis", 0.10)
        degisim   = min(degisim, max_ss)
        sebep = (
            f"Super Star ACoS {acos:.1f}%, korumali artis +%{degisim*100:.1f}"
            f" [{param_kaynagi}: hass={hassasiyet}, max={max_degisim}]"
        )

    elif segment in ("KAZANAN", "OPTIMIZE_ET", "ZARAR"):
        acos      = target["acos"]
        acos_fark = (acos - hedef_acos) / hedef_acos if hedef_acos > 0 else 0
        degisim   = -math.tanh(acos_fark * hassasiyet) * max_degisim
        sebep = (
            f"ACoS {acos:.1f}% vs hedef %{hedef_acos}, tanh {degisim*100:+.1f}%"
            f" [{param_kaynagi}: hass={hassasiyet}, max={max_degisim}]"
        )

    # ── Guvenlik siniri (ASIN parametresi bile asamaz) ────────────────────
    degisim  = max(-max_dusus_limit, min(max_artis_limit, degisim))
    yeni_bid = mevcut_bid * (1 + degisim)
    yeni_bid = max(min_bid, min(max_bid, round(yeni_bid, 2)))

    gercek_degisim = round((yeni_bid - mevcut_bid) / mevcut_bid * 100, 1) if mevcut_bid > 0 else 0

    return yeni_bid, gercek_degisim, sebep


# ============================================================================
# NEGATIF KEYWORD ADAYI TESPITI
# ============================================================================

def find_negative_candidates(data, settings):
    """
    Negatif keyword/targeting adaylarini belirler.

    KAYNAK: 30 gunluk search term raporlari (SP + SB)
    KAPSAM: Harvesting ile ayni — Auto/Broad/Phrase/Category/ASIN kampanyalarindan
    gelen search term'ler taranir.

    MANTIK: Cok tiklama alip satis yapmayan (veya cok az satis yapan) search term'ler
    negatif keyword adayi olarak isaretlenir. Amac: Bu search term'leri kaynak
    kampanyanin negatif hedeflemesine ekleyerek israf harcamayi durdurmak.

    KONTROL: Zaten negatif olarak eklenmis search term'ler atlanir.

    Kurallar (settings.negatif_keyword_kurali):
      - min_tiklama: Minimum tiklama sayisi (varsayilan: 20)
      - max_satis: Maksimum satis sayisi (varsayilan: 0 = satissiz)

    SP icin: purchases14d/sales14d
    SB icin: purchases/sales
    SD: Search term raporu yok.
    """
    kural = settings.get("negatif_keyword_kurali", {})
    min_tik = kural.get("min_tiklama", 20)
    max_satis = kural.get("max_satis", 0)

    camp_names, camp_portfolios = build_campaign_maps(data, settings)
    candidates = []
    seen = set()  # Tekrar engelleme
    skipped_existing = 0  # Zaten negatif olanlarin sayaci

    # ================================================================
    # MEVCUT NEGATIF KEYWORD/TARGET SETINI OLUSTUR
    # Amac: Zaten negatif olarak eklenmis terimleri tekrar onermemek
    # Key formati: "{campaign_id}_{keyword_text_lower}" veya global "{keyword_text_lower}"
    # ================================================================
    existing_negatives = set()  # kampanya bazli: "cid_term"
    existing_negatives_global = set()  # tum kampanyalar icin: "term"

    # SP ad group negatif keyword'ler
    for nk in data["sp"].get("negative_keywords", []):
        text = (nk.get("keywordText", "") or "").lower().strip()
        cid = str(nk.get("campaignId", ""))
        if text:
            existing_negatives.add(f"{cid}_{text}")
            existing_negatives_global.add(text)

    # SP kampanya negatif keyword'ler (kampanya seviyesi — tum ad group'lara uygulanir)
    for nk in data["sp"].get("campaign_negative_keywords", []):
        text = (nk.get("keywordText", "") or "").lower().strip()
        cid = str(nk.get("campaignId", ""))
        if text:
            existing_negatives.add(f"{cid}_{text}")
            existing_negatives_global.add(text)

    # SP negatif targeting (ASIN/kategori)
    for nt in data["sp"].get("negative_targets", []):
        expr = (nt.get("expression", []) or [])
        cid = str(nt.get("campaignId", ""))
        for e in expr:
            val = (e.get("value", "") or "").lower().strip()
            if val:
                existing_negatives.add(f"{cid}_{val}")
                existing_negatives_global.add(val)

    # SB negatif keyword'ler
    for nk in data["sb"].get("negative_keywords", []):
        text = (nk.get("keywordText", "") or "").lower().strip()
        cid = str(nk.get("campaignId", ""))
        if text:
            existing_negatives.add(f"{cid}_{text}")
            existing_negatives_global.add(text)

    logger.info("Mevcut negatif set: %d kampanya-bazli, %d global",
                len(existing_negatives), len(existing_negatives_global))

    # Negatif'e dahil edilecek match type'lar (harvesting ile ayni kaynak)
    NEG_MATCH_TYPES = {
        "AUTO",
        "BROAD",
        "PHRASE",
        "TARGETING_EXPRESSION",
        "TARGETING_EXPRESSION_PREDEFINED",
    }

    def _add_neg_candidate(ad_type, cid, row, sales_key, orders_key, kaynak, search_term):
        """Negatif adayini listeye ekler."""
        nonlocal skipped_existing

        # KONTROL: Bu search term zaten bu kampanyada negatif mi?
        term_lower = search_term.lower().strip()
        if f"{cid}_{term_lower}" in existing_negatives:
            skipped_existing += 1
            return  # Zaten negatif — atla

        orders = int(float(row.get(orders_key, 0)))
        sales = float(row.get(sales_key, 0))
        spend = float(row.get("cost", 0))
        clicks = int(float(row.get("clicks", 0)))
        impressions = int(float(row.get("impressions", 0)))
        acos = round((spend / sales * 100), 2) if sales > 0 else 0
        cvr = round((orders / clicks * 100), 2) if clicks > 0 else 0
        cpc = round(spend / clicks, 2) if clicks > 0 else 0

        # Negatif kurali: min tiklama VE max satis (cok tiklama + satis yok/az)
        if clicks < min_tik or orders > max_satis:
            return

        # Tekrar kontrolu
        dedupe_key = f"{ad_type}_{cid}_{search_term.lower()}"
        if dedupe_key in seen:
            return
        seen.add(dedupe_key)

        match_type = row.get("matchType", row.get("targetingType", row.get("keywordType", "")))

        candidates.append({
            "reklam_tipi": ad_type,
            "kampanya_adi": camp_names.get(cid, cid),
            "kampanya_id": cid,
            "portfolio_id": camp_portfolios.get(cid, ""),
            "hedefleme": search_term,
            "match_type": match_type,
            "kaynak": kaynak,
            "impressions": impressions,
            "clicks": clicks,
            "spend": round(spend, 2),
            "sales": round(sales, 2),
            "orders": orders,
            "acos": acos,
            "cvr": cvr,
            "cpc": cpc,
            "sebep": f"{clicks} tik, {orders} satis, ${spend:.2f} harcama. "
                     f"Negatif keyword olarak ekle ({kaynak}).",
        })

    # ================================================================
    # SP Search Term 30d
    # ================================================================
    sp_st_30 = data["reports"].get("sp_search_term_report_30d", [])
    for row in sp_st_30:
        mt = (row.get("matchType", "") or row.get("keywordType", "") or "").upper()
        if mt not in NEG_MATCH_TYPES:
            continue
        cid = str(row.get("campaignId", ""))
        search_term = row.get("searchTerm", "")
        _add_neg_candidate("SP", cid, row, "sales14d", "purchases14d",
                           f"Search Term ({mt})", search_term)

    # ================================================================
    # SB Search Term 30d
    # ================================================================
    sb_st_30 = data["reports"].get("sb_search_term_report_30d", [])
    for row in sb_st_30:
        mt = (row.get("matchType", "") or row.get("keywordType", "") or "").upper()
        if mt not in NEG_MATCH_TYPES:
            continue
        cid = str(row.get("campaignId", ""))
        search_term = row.get("searchTerm", "")
        _add_neg_candidate("SB", cid, row, "sales", "purchases",
                           f"Search Term ({mt})", search_term)

    logger.info("Negatif adaylar: %d (search term bazli, %d zaten negatif — atlandi)",
                len(candidates), skipped_existing)
    return candidates


# ============================================================================
# HARVESTING ADAYI TESPITI — DUZELTILMIS KOLON ISIMLERI
# ============================================================================

def find_harvesting_candidates(data, settings):
    """
    Harvesting adaylarini belirler.

    KAYNAK: 30 gunluk search term raporlari (SP + SB)
    KAPSAM: Asagidaki hedefleme tiplerinden gelen TUM search term'ler:
      - Auto targeting reklamlari
      - Broad hedeflemeli reklamlar
      - Phrase hedeflemeli reklamlar
      - Kategori targeting reklamlari (TARGETING_EXPRESSION)
      - ASIN target reklamlari (expanded asin targetler dahil)

    SD: Search term raporu yok (Amazon desteklemiyor).
    SD targeting 30d verisi ayri olarak islenir.

    KONTROL: Zaten exact keyword olarak eklenmis search term'ler atlanir.

    Harvesting kurallari (settings): min_satis, max_acos
    Amac: Iyi performans gosteren terimleri exact kampanyalara tasimak.
    """
    yeni_kw_kural = settings.get("yeni_keyword_kurali", {})
    min_satis = yeni_kw_kural.get("min_satis", 3)
    max_acos = yeni_kw_kural.get("max_acos", 15)

    camp_names, camp_portfolios = build_campaign_maps(data, settings)
    candidates = []
    seen = set()  # Tekrar engelleme
    skipped_existing = 0  # Zaten exact olanlarin sayaci

    # ================================================================
    # MEVCUT EXACT KEYWORD SETINI OLUSTUR
    # Amac: Zaten exact keyword olarak eklenmis terimleri tekrar onermemek
    # Herhangi bir kampanyada exact olarak varsa, tekrar onermeyiz
    # ================================================================
    existing_exact_keywords = set()  # global set: "term_lower"

    # SP keyword'ler — sadece EXACT match type olanlari
    for kw in data["sp"].get("keywords", []):
        mt = (kw.get("matchType", "") or "").upper()
        if mt == "EXACT":
            text = (kw.get("keywordText", "") or "").lower().strip()
            if text:
                existing_exact_keywords.add(text)

    # SP targeting clauses — exact ASIN/kategori targeting
    for tgt in data["sp"].get("targets", []):
        expr = tgt.get("expression", []) or []
        for e in expr:
            val = (e.get("value", "") or "").lower().strip()
            if val:
                existing_exact_keywords.add(val)

    # SB keyword'ler — sadece EXACT match type olanlari
    for kw in data["sb"].get("keywords", []):
        mt = (kw.get("matchType", "") or "").upper()
        if mt == "EXACT":
            text = (kw.get("keywordText", "") or "").lower().strip()
            if text:
                existing_exact_keywords.add(text)

    # SB targeting clauses
    for tgt in data["sb"].get("targets", []):
        expr = tgt.get("expression", tgt.get("expressions", [])) or []
        if isinstance(expr, list):
            for e in expr:
                val = (e.get("value", "") or "").lower().strip() if isinstance(e, dict) else ""
                if val:
                    existing_exact_keywords.add(val)

    logger.info("Mevcut exact keyword seti: %d terim", len(existing_exact_keywords))

    # Harvesting'e dahil edilecek match type'lar
    HARVEST_MATCH_TYPES = {
        "AUTO",                              # Auto targeting
        "BROAD",                             # Broad hedefleme
        "PHRASE",                            # Phrase hedefleme
        "TARGETING_EXPRESSION",              # Kategori + ASIN targeting
        "TARGETING_EXPRESSION_PREDEFINED",   # Predefined ASIN/kategori
    }

    def _add_candidate(ad_type, cid, row, sales_key, orders_key, kaynak, hedefleme_text):
        """Harvesting adayini listeye ekler (tekrar kontrolu ile)."""
        nonlocal skipped_existing

        # KONTROL: Bu search term zaten bir exact kampanyada var mi?
        term_lower = hedefleme_text.lower().strip()
        if term_lower in existing_exact_keywords:
            skipped_existing += 1
            return  # Zaten exact'te var — atla

        orders = int(float(row.get(orders_key, 0)))
        sales = float(row.get(sales_key, 0))
        spend = float(row.get("cost", 0))
        clicks = int(float(row.get("clicks", 0)))
        impressions = int(float(row.get("impressions", 0)))
        acos = round((spend / sales * 100), 2) if sales > 0 else 0
        cvr = round((orders / clicks * 100), 2) if clicks > 0 else 0

        # Harvesting kurali: min satis ve max ACoS
        if orders < min_satis or acos > max_acos or acos == 0:
            return

        # Tekrar kontrolu
        dedupe_key = f"{ad_type}_{cid}_{hedefleme_text.lower()}"
        if dedupe_key in seen:
            return
        seen.add(dedupe_key)

        match_type = row.get("matchType", row.get("targetingType", row.get("keywordType", "")))

        candidates.append({
            "reklam_tipi": ad_type,
            "kaynak_kampanya": camp_names.get(cid, cid),
            "kampanya_id": cid,
            "portfolio_id": camp_portfolios.get(cid, ""),
            "hedefleme": hedefleme_text,
            "match_type": match_type,
            "kaynak": kaynak,
            "impressions": impressions,
            "clicks": clicks,
            "spend": round(spend, 2),
            "sales": round(sales, 2),
            "orders": orders,
            "acos": acos,
            "cvr": cvr,
            "oneri": f"Exact kampanyaya tasi. {orders} satis, ACoS %{acos:.1f} ({kaynak})",
        })

    # ================================================================
    # SP Search Term 30d — Auto/Broad/Phrase/Category/ASIN targeting
    # TUM kampanyalardan (izlenen sinirlamasi YOK)
    # ================================================================
    sp_st_30 = data["reports"].get("sp_search_term_report_30d", [])
    for row in sp_st_30:
        mt = (row.get("matchType", "") or row.get("keywordType", "") or "").upper()
        if mt not in HARVEST_MATCH_TYPES:
            continue
        cid = str(row.get("campaignId", ""))
        search_term = row.get("searchTerm", "")
        _add_candidate("SP", cid, row, "sales14d", "purchases14d",
                       f"Search Term ({mt})", search_term)

    # ================================================================
    # SB Search Term 30d — Broad/Phrase/Category/ASIN targeting
    # (SB'de auto targeting yok, ama TARGETING_EXPRESSION olabilir)
    # ================================================================
    sb_st_30 = data["reports"].get("sb_search_term_report_30d", [])
    for row in sb_st_30:
        mt = (row.get("matchType", "") or row.get("keywordType", "") or "").upper()
        if mt not in HARVEST_MATCH_TYPES:
            continue
        cid = str(row.get("campaignId", ""))
        search_term = row.get("searchTerm", "")
        _add_candidate("SB", cid, row, "sales", "purchases",
                       f"Search Term ({mt})", search_term)

    # ================================================================
    # SD Targeting 30d — SD'nin harvesting icin tek kaynagi
    # (SD'de search term raporu yok, targeting raporundan alinir)
    # ================================================================
    sd_tgt_30 = data["reports"].get("sd_targeting_report_30d", [])
    for row in sd_tgt_30:
        cid = str(row.get("campaignId", ""))
        hedef_text = row.get("targetingExpression", row.get("targeting", ""))
        _add_candidate("SD", cid, row, "sales", "purchases",
                       "Targeting (SD)", hedef_text)

    logger.info("Harvesting: %d aday bulundu (%d zaten exact'te var — atlandi)",
                len(candidates), skipped_existing)
    return candidates


# ============================================================================
# SIRALAMA FONKSIYONU
# ============================================================================

def sort_bid_results(bid_results):
    """
    Siralama:
      1. Yuksek ACoS → dusuk ACoS (satis yapanlar, ACoS > 0)
      2. Harcayan ama satis yapmayan (spend > 0, sales == 0)
      3. Harcamayanlar (spend == 0)
    """
    # 3 grup olustur
    satisli = [r for r in bid_results if r["sales"] > 0]
    harcayan_satissiz = [r for r in bid_results if r["spend"] > 0 and r["sales"] == 0]
    harcamayan = [r for r in bid_results if r["spend"] == 0]

    # Satislilari yuksek ACoS'tan dusuge sirala
    satisli.sort(key=lambda x: x["acos"], reverse=True)
    # Harcayan satissizlari yuksek harcamadan dusuge
    harcayan_satissiz.sort(key=lambda x: x["spend"], reverse=True)
    # Harcamayanlari impression'a gore
    harcamayan.sort(key=lambda x: x["impressions"], reverse=True)

    return satisli + harcayan_satissiz + harcamayan


# ============================================================================
# EXCEL RAPOR OLUSTURUCU — TAMAMEN YENIDEN YAZILDI
# ============================================================================

def create_excel_reports(bid_results, negative_candidates, harvesting_candidates,
                          today, settings, currency="$"):
    """3 Excel dosyasi olusturur. Formatlama kurallari:
    - ACoS: % sembol ile (orn: 24.25%)
    - Tutarlar: Para birimi ile (orn: $1,234.56)
    - Ondalik: 2 basamak
    - ASIN ve Orders sutunlari YOK
    - Sutun sirasi: Kampanya, Reklam Tipi, Portfolio, Hedefleme, Match Type,
      Impression, CVR, Click, Spend, Sales, ACoS, CPC, Bid, Tavsiye Bid, Degisim, Sebep, Segment
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    except ImportError:
        logger.error("openpyxl kurulu degil! pip install openpyxl")
        return {}

    ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)
    files = {}

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    segment_colors = {
        "GORUNMEZ": PatternFill(start_color="E0E0E0", fill_type="solid"),
        "YETERSIZ_VERI": PatternFill(start_color="F5F5F5", fill_type="solid"),
        "KAN_KAYBEDEN": PatternFill(start_color="FFCDD2", fill_type="solid"),
        "SUPER_STAR": PatternFill(start_color="C8E6C9", fill_type="solid"),
        "TUZAK": PatternFill(start_color="FFF9C4", fill_type="solid"),
        "KAZANAN": PatternFill(start_color="A5D6A7", fill_type="solid"),
        "OPTIMIZE_ET": PatternFill(start_color="FFE0B2", fill_type="solid"),
        "ZARAR": PatternFill(start_color="EF9A9A", fill_type="solid"),
        "IMPRESSION_BEKLE": PatternFill(start_color="B3E5FC", fill_type="solid"),
    }
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, cols):
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def write_cell(ws, row, col, value, fmt=None, is_segment=False, segment_name=""):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        if fmt:
            cell.number_format = fmt
        if is_segment:
            cell.fill = segment_colors.get(segment_name, PatternFill())
        return cell

    def fmt_currency(val):
        """Para birimi formati: $1,234.56"""
        return f"{currency}{val:,.2f}"

    def fmt_pct(val):
        """Yuzde formati: 24.25%"""
        return f"{val:.2f}%"

    # ---- LISTE 1: BID TAVSIYELERI ----
    # Siralama uygula
    sorted_results = sort_bid_results(bid_results)

    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "Bid Tavsiyeleri"

    # Yeni sutun sirasi (ASIN ve Orders YOK)
    cols1 = [
        "Kampanya", "Reklam Tipi", "Portfolio", "Hedefleme", "Match Type",
        "Impression", "CVR %", "Click", "Spend", "Sales",
        "ACoS %", "CPC", "Bid", "Tavsiye Bid", "Degisim %",
        "Sebep", "Segment", "Onay",
    ]
    style_header(ws1, cols1)

    # Satir renklendirme — ACoS durumuna gore
    row_fill_green = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")   # hedef alti — iyi
    row_fill_yellow = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")   # hedef ustu, orta
    row_fill_red = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")      # zarar / satissiz

    def get_row_fill(r):
        """Satirin ACoS durumuna gore arka plan rengini belirler."""
        hedef = r.get("hedef_acos", 25)
        if r["sales"] == 0 and r["spend"] > 0:
            return row_fill_red  # Harcayan ama satissiz → kirmizi
        if r["sales"] == 0 and r["spend"] == 0:
            return None  # Harcamayan → renk yok
        if r["acos"] <= hedef:
            return row_fill_green  # Hedef ACoS altinda → yesil
        if r["acos"] <= hedef * 1.5:
            return row_fill_yellow  # Hedef ustu ama 1.5x altinda → sari
        return row_fill_red  # Hedef x1.5 ustu → kirmizi

    for idx, r in enumerate(sorted_results, 2):
        row_color = get_row_fill(r)
        c = 1
        cell = write_cell(ws1, idx, c, r["kampanya_adi"]); c += 1
        cell = write_cell(ws1, idx, c, r["reklam_tipi"]); c += 1
        cell = write_cell(ws1, idx, c, r.get("portfolio_id", "")); c += 1
        cell = write_cell(ws1, idx, c, r["hedefleme"]); c += 1
        cell = write_cell(ws1, idx, c, r["match_type"]); c += 1
        cell = write_cell(ws1, idx, c, r["impressions"], "#,##0"); c += 1
        cell = write_cell(ws1, idx, c, fmt_pct(r["cvr"])); c += 1
        cell = write_cell(ws1, idx, c, r["clicks"], "#,##0"); c += 1
        cell = write_cell(ws1, idx, c, fmt_currency(r["spend"])); c += 1
        cell = write_cell(ws1, idx, c, fmt_currency(r["sales"])); c += 1
        cell = write_cell(ws1, idx, c, fmt_pct(r["acos"])); c += 1
        cell = write_cell(ws1, idx, c, fmt_currency(r["cpc"])); c += 1
        cell = write_cell(ws1, idx, c, fmt_currency(r["mevcut_bid"])); c += 1
        cell = write_cell(ws1, idx, c, fmt_currency(r["yeni_bid"])); c += 1
        cell = write_cell(ws1, idx, c, f"{r['degisim_yuzde']:+.1f}%"); c += 1
        cell = write_cell(ws1, idx, c, r["sebep"]); c += 1
        cell = write_cell(ws1, idx, c, r["segment"], is_segment=True, segment_name=r["segment"]); c += 1
        cell = write_cell(ws1, idx, c, "")  # Onay kolonu — bos (kullanici Y yazacak)

        # Satir rengi uygula (segment ve onay hucresine segment/bos rengi kalir)
        if row_color:
            for col_idx in range(1, 17):  # A-P sutunlari (segment ve onay haric)
                ws1.cell(row=idx, column=col_idx).fill = row_color

    # Kolon genislikleri
    col_widths = {
        "A": 35, "B": 12, "C": 20, "D": 30, "E": 12,
        "F": 12, "G": 10, "H": 10, "I": 14, "J": 14,
        "K": 10, "L": 10, "M": 10, "N": 14, "O": 10,
        "P": 50, "Q": 18, "R": 8,
    }
    for col_letter, width in col_widths.items():
        ws1.column_dimensions[col_letter].width = width

    path1 = str(ANALYSIS_DIR / f"{today}_bid_recommendations.xlsx")
    wb1.save(path1)
    files["bid_recommendations"] = path1
    logger.info("Bid tavsiyeleri kaydedildi: %s (%d satir)", path1, len(sorted_results))

    # ---- LISTE 2: NEGATIF ADAYLAR ----
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Negatif Adaylar"
    cols2 = [
        "Kampanya", "Reklam Tipi", "Portfolio", "Search Term", "Match Type",
        "Kaynak", "Impression", "Click", "Spend", "Sales", "CVR %", "CPC",
        "Sebep", "Onay",
    ]
    style_header(ws2, cols2)

    # Negatif adaylar icin kirmizinin 2 tonu
    # Koyu kirmizi: Satis=0 VE harcama var (tamamen israf)
    # Acik kirmizi: Satis var ama ACoS cok yuksek / CVR cok dusuk
    neg_fill_dark = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")   # koyu kirmizi
    neg_fill_light = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")   # acik kirmizi

    def get_neg_fill(r):
        """Negatif adayin siddetine gore renk."""
        sales = float(r.get("sales", 0))
        spend = float(r.get("spend", 0))
        if sales == 0 and spend > 0:
            return neg_fill_dark   # Tamamen israf — koyu kirmizi
        return neg_fill_light      # Kotu performans ama satis var — acik kirmizi

    for idx, r in enumerate(negative_candidates, 2):
        neg_color = get_neg_fill(r)
        c = 1
        write_cell(ws2, idx, c, r["kampanya_adi"]); c += 1
        write_cell(ws2, idx, c, r["reklam_tipi"]); c += 1
        write_cell(ws2, idx, c, r.get("portfolio_id", "")); c += 1
        write_cell(ws2, idx, c, r["hedefleme"]); c += 1
        write_cell(ws2, idx, c, r["match_type"]); c += 1
        write_cell(ws2, idx, c, r.get("kaynak", "")); c += 1
        write_cell(ws2, idx, c, r["impressions"], "#,##0"); c += 1
        write_cell(ws2, idx, c, r["clicks"], "#,##0"); c += 1
        write_cell(ws2, idx, c, fmt_currency(r["spend"])); c += 1
        write_cell(ws2, idx, c, fmt_currency(r["sales"])); c += 1
        write_cell(ws2, idx, c, fmt_pct(r["cvr"])); c += 1
        write_cell(ws2, idx, c, fmt_currency(r["cpc"])); c += 1
        write_cell(ws2, idx, c, r.get("sebep", "")); c += 1
        write_cell(ws2, idx, c, "")  # Onay kolonu

        # Renk uygula (Onay kolonu haric)
        if neg_color:
            for col_idx in range(1, 14):  # A-M sutunlari
                ws2.cell(row=idx, column=col_idx).fill = neg_color

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 30
    ws2.column_dimensions["F"].width = 20
    ws2.column_dimensions["M"].width = 50
    ws2.column_dimensions["N"].width = 8

    path2 = str(ANALYSIS_DIR / f"{today}_negative_candidates.xlsx")
    wb2.save(path2)
    files["negative_candidates"] = path2

    # ---- LISTE 3: HARVESTING ADAYLAR ----
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active
    ws3.title = "Harvesting Adaylar"
    cols3 = [
        "Kaynak Kampanya", "Reklam Tipi", "Portfolio", "Hedefleme", "Match Type",
        "Kaynak", "Impression", "Click", "Spend", "Sales", "ACoS %", "CVR %",
        "Oneri", "Onay",
    ]
    style_header(ws3, cols3)

    # Harvesting adaylari icin yesilin 2 tonu
    # Koyu yesil: ACoS cok dusuk (hedef ACoS'un yarisi veya altinda) — mukemmel performans
    # Acik yesil: ACoS iyi ama hedef ACoS'a yakin — iyi performans
    harvest_fill_dark = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")    # koyu yesil
    harvest_fill_light = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")   # acik yesil

    harvest_max_acos = settings.get("yeni_keyword_kurali", {}).get("max_acos", 15)

    def get_harvest_fill(r):
        """Harvesting adayinin performansina gore renk."""
        acos = r.get("acos", 0)
        if acos <= harvest_max_acos / 2:
            return harvest_fill_dark   # ACoS hedefin yarisi veya alti — koyu yesil
        return harvest_fill_light      # ACoS iyi ama yuksek tarafta — acik yesil

    for idx, r in enumerate(harvesting_candidates, 2):
        h_color = get_harvest_fill(r)
        c = 1
        write_cell(ws3, idx, c, r["kaynak_kampanya"]); c += 1
        write_cell(ws3, idx, c, r["reklam_tipi"]); c += 1
        write_cell(ws3, idx, c, r.get("portfolio_id", "")); c += 1
        write_cell(ws3, idx, c, r["hedefleme"]); c += 1
        write_cell(ws3, idx, c, r["match_type"]); c += 1
        write_cell(ws3, idx, c, r.get("kaynak", "")); c += 1
        write_cell(ws3, idx, c, r["impressions"], "#,##0"); c += 1
        write_cell(ws3, idx, c, r["clicks"], "#,##0"); c += 1
        write_cell(ws3, idx, c, fmt_currency(r["spend"])); c += 1
        write_cell(ws3, idx, c, fmt_currency(r["sales"])); c += 1
        write_cell(ws3, idx, c, fmt_pct(r["acos"])); c += 1
        write_cell(ws3, idx, c, fmt_pct(r["cvr"])); c += 1
        write_cell(ws3, idx, c, r.get("oneri", "")); c += 1
        write_cell(ws3, idx, c, "")  # Onay kolonu

        # Renk uygula (Onay kolonu haric)
        if h_color:
            for col_idx in range(1, 14):  # A-M sutunlari
                ws3.cell(row=idx, column=col_idx).fill = h_color

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 30
    ws3.column_dimensions["F"].width = 20
    ws3.column_dimensions["M"].width = 50
    ws3.column_dimensions["N"].width = 8

    path3 = str(ANALYSIS_DIR / f"{today}_harvesting_candidates.xlsx")
    wb3.save(path3)
    files["harvesting_candidates"] = path3

    return files


# ============================================================================
# JSON → EXCEL DONUSTURUCU (Ham Veri)
# ============================================================================

def convert_raw_data_to_excel(data, today):
    try:
        import openpyxl
    except ImportError:
        return {}

    ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)
    files = {}

    datasets = [
        ("sp_campaigns", data["sp"]["campaigns"]),
        ("sp_keywords", data["sp"]["keywords"]),
        ("sp_targets", data["sp"]["targets"]),
        ("sb_campaigns", data["sb"]["campaigns"]),
        ("sb_keywords", data["sb"]["keywords"]),
        ("sb_targets", data["sb"]["targets"]),
        ("sd_campaigns", data["sd"]["campaigns"]),
        ("sd_targets", data["sd"]["targets"]),
    ]

    for data_key, items in datasets:
        if not items:
            continue
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = data_key

        if isinstance(items[0], dict):
            headers = list(items[0].keys())
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            for row_idx, item in enumerate(items, 2):
                for col, h in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col, value=str(item.get(h, "")))

        path = str(ANALYSIS_DIR / f"{today}_{data_key}.xlsx")
        wb.save(path)
        files[data_key] = path

    return files


# ============================================================================
# KARAR GECMISI KAYIT
# ============================================================================

def save_decisions(bid_results, today):
    DECISIONS_DIR.mkdir(parents=True, exist_ok=True)

    decisions = []
    for r in bid_results:
        decisions.append({
            "tarih": today,
            "hedefleme_id": r["hedefleme_id"],
            "reklam_tipi": r["reklam_tipi"],
            "hedefleme": r["hedefleme"],
            "kampanya": r["kampanya_adi"],
            "portfolio_id": r.get("portfolio_id", ""),
            "asin": r.get("asin", ""),
            "segment": r["segment"],
            "onceki_bid": r["mevcut_bid"],
            "yeni_bid": r["yeni_bid"],
            "degisim_yuzde": r["degisim_yuzde"],
            "sebep": r["sebep"],
            "metrikler": {
                "impressions": r["impressions"],
                "clicks": r["clicks"],
                "spend": r["spend"],
                "sales": r["sales"],
                "orders": r["orders"],
                "acos": r["acos"],
                "cvr": r["cvr"],
                "cpc": r["cpc"],
            },
            "karar_durumu": "ONAY_BEKLIYOR",
        })

    filepath = DECISIONS_DIR / f"{today}_decisions.json"
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(decisions, f, ensure_ascii=False, indent=2)

    logger.info("Karar gecmisi kaydedildi: %s (%d karar)", filepath, len(decisions))
    return str(filepath)


# ============================================================================
# ON KONTROL
# ============================================================================

def preflight_check(today):
    """
    Supabase-first: Önce Supabase'de kritik veriler var mı kontrol et.
    Fallback: JSON dosyalarını kontrol et.
    """
    hk = os.environ.get("HESAP_KEY", "")
    mp = os.environ.get("MARKETPLACE", "")

    # Supabase-first kontrol
    if hk and mp:
        try:
            from data_loader import _fetch_count, check_report_exists
            sp_camp = _fetch_count("campaigns", hk, mp, "AND ad_type = 'SP'")
            sp_report = check_report_exists(hk, mp, "targeting_reports", "SP", today)

            if sp_camp > 0 and sp_report:
                logger.info("Preflight Supabase: SP kampanya=%d, targeting rapor=VAR", sp_camp)
                uyarilar = []
                # Opsiyonel kontroller
                sb_camp = _fetch_count("campaigns", hk, mp, "AND ad_type = 'SB'")
                sd_camp = _fetch_count("campaigns", hk, mp, "AND ad_type = 'SD'")
                if sb_camp == 0:
                    uyarilar.append("SB kampanya yok — SB analizi kisitli olabilir.")
                if sd_camp == 0:
                    uyarilar.append("SD kampanya yok — SD analizi kisitli olabilir.")
                return True, f"On kontrol gecti (Supabase). SP={sp_camp} kampanya. Tarih: {today}", uyarilar

            if sp_camp == 0:
                logger.warning("Preflight Supabase: SP kampanya=0, JSON fallback deneniyor")
            if not sp_report:
                logger.warning("Preflight Supabase: SP targeting raporu yok, JSON fallback deneniyor")
        except Exception as e:
            logger.warning("Preflight Supabase basarisiz, JSON fallback: %s", e)

    # JSON fallback
    kritik_dosyalar = [
        f"{today}_sp_campaigns.json",
        f"{today}_sp_targeting_report_14d.json",
    ]
    opsiyonel_dosyalar = [
        f"{today}_portfolios.json",
        f"{today}_sp_keywords.json",
        f"{today}_sp_targets.json",
        f"{today}_sp_negative_keywords.json",
        f"{today}_sp_campaign_negative_keywords.json",
        f"{today}_sp_negative_targets.json",
        f"{today}_sp_search_term_report_30d.json",
        f"{today}_sb_campaigns.json",
        f"{today}_sb_keywords.json",
        f"{today}_sb_negative_keywords.json",
        f"{today}_sb_targeting_report_14d.json",
        f"{today}_sb_search_term_report_30d.json",
        f"{today}_sd_campaigns.json",
        f"{today}_sd_targeting_report_14d.json",
        f"{today}_sd_targeting_report_30d.json",
    ]

    eksik_kritik = []
    eksik_opsiyonel = []

    for fname in kritik_dosyalar:
        if not (DATA_DIR / fname).exists():
            eksik_kritik.append(fname)

    for fname in opsiyonel_dosyalar:
        if not (DATA_DIR / fname).exists():
            eksik_opsiyonel.append(fname)

    uyarilar = []

    if eksik_kritik:
        mesaj = (
            f"HATA: Agent 1 bugun ({today}) henuz calistirilmamis veya kritik dosyalar eksik.\n"
            f"Eksik kritik dosyalar: {', '.join(eksik_kritik)}\n"
            f"Lutfen once Agent 1'i calistirin."
        )
        return False, mesaj, []

    if eksik_opsiyonel:
        for f in eksik_opsiyonel:
            if "sb_" in f:
                uyarilar.append(f"SB dosyasi eksik: {f} — SB analizi kisitli olabilir.")
            elif "sd_" in f:
                uyarilar.append(f"SD dosyasi eksik: {f} — SD analizi kisitli olabilir.")
            else:
                uyarilar.append(f"Opsiyonel dosya eksik: {f}")

    mevcut = len(kritik_dosyalar) + len(opsiyonel_dosyalar) - len(eksik_opsiyonel)
    toplam = len(kritik_dosyalar) + len(opsiyonel_dosyalar)
    mesaj = f"On kontrol gecti (JSON fallback). {mevcut}/{toplam} dosya mevcut. Tarih: {today}"

    return True, mesaj, uyarilar


# ============================================================================
# ANA FONKSIYON
# ============================================================================

def save_error_log(hata_tipi, hata_mesaji, traceback_str=None, adim=None,
                   extra=None, session_id=None):
    """Agent 2 hata logu — lokal + Supabase dual-write."""
    log_dir = DATA_DIR / "logs"
    dir_name = DATA_DIR.name  # "vigowood_eu_UK"
    parts = dir_name.rsplit("_", 1)
    hk = parts[0] if len(parts) == 2 else ""
    mp = parts[1] if len(parts) == 2 else ""
    return _central_save_error_log(
        hata_tipi, hata_mesaji, log_dir,
        traceback_str=traceback_str, adim=adim, extra=extra,
        session_id=session_id, agent_name="agent2",
        hesap_key=hk, marketplace=mp)


def run_analysis(hesap_key, marketplace):
    today = datetime.utcnow().strftime("%Y-%m-%d")
    init_paths(hesap_key, marketplace)
    logger.info("=== AGENT 2 v5 ANALIZ BASLADI — %s/%s — %s ===", hesap_key, marketplace, today)

    # Dashboard: Agent 2 basliyor
    _session_id = MAESTRO_SESSION_ID or f"direct_{today}_{hesap_key}_{marketplace}"
    _dashboard_status("agent2", "running")
    _dashboard_pipeline(_session_id, hesap_key, marketplace, "agent2", "running")
    _save_log("info", f"Agent 2 basliyor: {hesap_key}/{marketplace}", "agent2", hesap_key, marketplace, _session_id)

    try:
        result = _run_analysis_impl(today, hesap_key, marketplace)
        # Dashboard: Agent 2 tamamlandi
        _final = "completed" if result.get("durum") != "BASARISIZ" else "failed"
        _dashboard_status("agent2", _final, {
            "tasks": result.get("toplam_hedef", 0) or result.get("toplam_hedefleme", 0),
            "errors_7d": 1 if _final == "failed" else 0,
        })
        _dashboard_pipeline(_session_id, hesap_key, marketplace, "agent2", _final)
        _save_log("info" if _final == "completed" else "error",
                  f"Agent 2 {'tamamlandi' if _final == 'completed' else 'basarisiz'}: {result.get('durum', '')}",
                  "agent2", hesap_key, marketplace, _session_id)
        return result
    except Exception as e:
        tb = traceback.format_exc()
        hata_tipi = type(e).__name__
        hata_mesaji = str(e)
        logger.error("BEKLENMEYEN HATA [%s]: %s", hata_tipi, hata_mesaji)
        save_error_log(hata_tipi, hata_mesaji, tb, adim="run_analysis",
                       session_id=MAESTRO_SESSION_ID)
        # Dashboard: Agent 2 hata
        _dashboard_status("agent2", "failed")
        _dashboard_pipeline(_session_id, hesap_key, marketplace, "agent2", "failed", hata_mesaji[:500])
        _save_log("error", f"Agent 2 hatasi: {hata_mesaji[:200]}", "agent2", hesap_key, marketplace, _session_id, error_type=hata_tipi)
        return {"tarih": today, "durum": "BASARISIZ", "hata": hata_mesaji}


def _run_analysis_impl(today, hesap_key="", marketplace=""):

    # 0. On kontrol
    gecti, mesaj, on_uyarilar = preflight_check(today)
    logger.info("On kontrol: %s", mesaj)

    if not gecti:
        logger.error(mesaj)
        save_error_log("Preflight", mesaj, "", adim="preflight_check",
                       session_id=MAESTRO_SESSION_ID)
        return {"tarih": today, "durum": "BASARISIZ", "hata": mesaj}

    if on_uyarilar:
        for u in on_uyarilar:
            logger.warning(u)

    # 1. Ayarlar
    settings = load_settings()
    bid_funcs = load_bid_functions()
    currency = get_currency(settings)
    logger.info("Ayarlar yuklendi. Para birimi: %s", currency)

    # 2. Verileri yukle
    data = load_all_agent1_data()
    logger.info("Agent 1 verileri yuklendi.")

    # 3. JSON → Excel
    raw_excel = convert_raw_data_to_excel(data, today)
    logger.info("JSON → Excel: %d dosya", len(raw_excel))

    # 4. Hedefleme listesi (aktif kampanyalar, dogru kolon isimleri)
    targets = build_targeting_list(data, settings)
    logger.info("Hedefleme listesi: %d hedefleme", len(targets))

    # 5. Bid fallback eslestirme (raporda bid=0 kalanlar icin)
    targets = enrich_with_bid_data(targets, data)

    # 6. ASIN
    targets = enrich_with_asin(targets, data)

    # 7. Onceki kararlar
    prev_decisions = load_previous_decisions()
    logger.info("Onceki kararlar: %d kayit", len(prev_decisions))

    # 8. Segmentasyon + Bid hesaplama
    bid_results = []
    segment_counts = {}

    for t in targets:
        segment, hedef_acos, sebep = segmentize(t, settings, prev_decisions)
        yeni_bid, degisim_yuzde, bid_sebep = calculate_new_bid(
            segment, t, hedef_acos, settings, bid_funcs, prev_decisions
        )

        t["segment"] = segment
        t["yeni_bid"] = yeni_bid
        t["degisim_yuzde"] = degisim_yuzde
        t["sebep"] = sebep if segment in ("YETERSIZ_VERI", "TUZAK", "IMPRESSION_BEKLE") else bid_sebep
        t["hedef_acos"] = hedef_acos

        bid_results.append(t)
        segment_counts[segment] = segment_counts.get(segment, 0) + 1

    logger.info("Segmentasyon: %s", json.dumps(segment_counts, indent=2))

    # 9. Negatif adaylar (search term bazli — 30d)
    neg_candidates = find_negative_candidates(data, settings)
    logger.info("Negatif adaylar: %d", len(neg_candidates))

    # 10. Harvesting
    harvest_candidates = find_harvesting_candidates(data, settings)
    logger.info("Harvesting adaylari: %d", len(harvest_candidates))

    # 11. Excel raporlari (siralama dahil)
    excel_files = create_excel_reports(bid_results, neg_candidates, harvest_candidates,
                                        today, settings, currency=currency)
    logger.info("Excel raporlari: %s", list(excel_files.keys()))

    # 12. Karar gecmisi
    decisions_file = save_decisions(bid_results, today)

    # 13. Ozet
    summary = {
        "tarih": today,
        "durum": "TAMAMLANDI",
        "on_kontrol_uyarilari": on_uyarilar,
        "toplam_hedefleme": len(targets),
        "segment_dagilimi": segment_counts,
        "bid_tavsiye_sayisi": sum(1 for r in bid_results if r["degisim_yuzde"] != 0),
        "dokunulmayan_sayisi": sum(1 for r in bid_results if r["degisim_yuzde"] == 0),
        "negatif_aday_sayisi": len(neg_candidates),
        "harvesting_aday_sayisi": len(harvest_candidates),
        "reklam_tipi_dagilimi": {
            "SP": sum(1 for r in bid_results if r["reklam_tipi"] == "SP"),
            "SB": sum(1 for r in bid_results if r["reklam_tipi"] == "SB"),
            "SD": sum(1 for r in bid_results if r["reklam_tipi"] == "SD"),
        },
        "para_birimi": currency,
        "dosyalar": {
            "bid_tavsiyeleri": excel_files.get("bid_recommendations", ""),
            "negatif_adaylar": excel_files.get("negative_candidates", ""),
            "harvesting_adaylar": excel_files.get("harvesting_candidates", ""),
            "karar_gecmisi": decisions_file,
            "ham_veri_excel": raw_excel,
        },
    }

    # ---- Supabase Sync ----
    _sync_agent2_to_supabase(hesap_key, marketplace, today,
                              bid_results, neg_candidates, harvest_candidates)

    logger.info("=== AGENT 2 v5 ANALIZ TAMAMLANDI ===")
    return summary


def _sync_agent2_to_supabase(hesap_key, marketplace, today,
                              bid_results, neg_candidates, harvest_candidates):
    """Agent 2 analiz sonuclarini Supabase'e yaz."""
    try:
        import sys as _sys
        _project_root = str(BASE_DIR)
        if _project_root not in _sys.path:
            _sys.path.insert(0, _project_root)
        from supabase.db_client import SupabaseClient
        db = SupabaseClient()
    except Exception as e:
        logger.error("Supabase sync import hatasi: %s", e)
        save_error_log("InternalError", f"Supabase sync import: {e}",
                       traceback.format_exc(), adim="supabase_sync_init",
                       session_id=MAESTRO_SESSION_ID)
        return

    # Portfolio bilgisini Supabase'ten al (tek guvenilir kaynak)
    # campaign_name uzerinden eslestirme (bid_recommendations'da campaign_id NULL olabiliyor)
    portfolio_by_name = {}  # campaign_name -> portfolio_name
    portfolio_by_id = {}    # campaign_id -> portfolio_name
    try:
        camp_rows = db._fetch_all(
            "SELECT campaign_id, name, portfolio_id FROM campaigns WHERE hesap_key=%s AND marketplace=%s AND portfolio_id IS NOT NULL",
            (hesap_key, marketplace))
        pf_rows = db._fetch_all(
            "SELECT portfolio_id, name FROM portfolios WHERE hesap_key=%s AND marketplace=%s",
            (hesap_key, marketplace))
        pid_to_name = {str(r[0]): r[1] for r in pf_rows if r[0] and r[1]}

        for r in camp_rows:
            cid, cname, pid = str(r[0] or ""), r[1] or "", str(r[2] or "")
            pf_name = pid_to_name.get(pid, "")
            if pf_name:
                if cid:
                    portfolio_by_id[cid] = pf_name
                if cname:
                    portfolio_by_name[cname] = pf_name
        logger.info("Portfolio lookup: %d by id, %d by name", len(portfolio_by_id), len(portfolio_by_name))
    except Exception as e:
        logger.warning("Portfolio lookup hatasi (devam eder): %s", e)

    def _get_portfolio(record):
        """campaign_id veya kampanya_adi uzerinden portfolio ismini dondurur."""
        cid = str(record.get("campaign_id", "") or "")
        if cid and cid in portfolio_by_id:
            return portfolio_by_id[cid]
        cname = record.get("kampanya_adi") or record.get("kampanya") or record.get("kaynak_kampanya") or ""
        return portfolio_by_name.get(cname, "")

    # Bugunun eski verilerini sil (lokaldeki sil-yeniden-yaz davranisiyla ayni)
    try:
        for tbl in ["bid_recommendations", "negative_candidates", "harvesting_candidates"]:
            deleted = db._execute(
                f"DELETE FROM {tbl} WHERE hesap_key=%s AND marketplace=%s AND analysis_date=%s",
                (hesap_key, marketplace, today))
            if deleted:
                logger.info("Supabase: %s — %s/%s/%s eski %d kayit silindi",
                            tbl, hesap_key, marketplace, today, deleted)
    except Exception as e:
        logger.warning("Eski veri silme hatasi (devam eder): %s", e)

    try:
        # Bid tavsiyeleri
        bid_data = []
        for r in bid_results:
            bid_data.append({
                "reklam_tipi": r.get("reklam_tipi", "SP"),
                "campaign_id": r.get("campaign_id"),
                "kampanya": r.get("kampanya_adi"),
                "ad_group_id": r.get("ad_group_id"),
                "keyword_id": r.get("keyword_id"),
                "target_id": r.get("target_id"),
                "hedefleme": r.get("hedefleme"),
                "match_type": r.get("match_type"),
                "segment": r.get("segment"),
                "bid": r.get("mevcut_bid"),
                "tavsiye_bid": r.get("yeni_bid"),
                "degisim_yuzde": r.get("degisim_yuzde"),
                "impressions": r.get("impressions"),
                "clicks": r.get("clicks"),
                "spend": r.get("spend"),
                "sales": r.get("sales"),
                "orders": r.get("orders"),
                "acos": r.get("acos"),
                "cvr": r.get("cvr"),
                "cpc": r.get("cpc"),
                "portfolio": _get_portfolio(r),
                "reason": r.get("sebep", ""),
                "karar_durumu": "PENDING",
            })
        if bid_data:
            db.insert_bid_recommendations(hesap_key, marketplace, today, bid_data)
            logger.info("Supabase: %d bid tavsiyesi yazildi", len(bid_data))

        # Negatif adaylar
        neg_data = []
        for n in neg_candidates:
            neg_data.append({
                "reklam_tipi": n.get("reklam_tipi", "SP"),
                "campaign_id": n.get("campaign_id"),
                "kampanya": n.get("kampanya_adi") or n.get("kampanya"),
                "ad_group_id": n.get("ad_group_id"),
                "hedefleme": n.get("search_term") or n.get("hedefleme"),
                "tip": n.get("tip", "KEYWORD"),
                "sebep": n.get("sebep"),
                "portfolio": _get_portfolio(n),
                "impressions": n.get("impressions"),
                "clicks": n.get("clicks"),
                "spend": n.get("spend") or n.get("cost"),
                "sales": n.get("sales"),
                "acos": n.get("acos"),
                "cvr": n.get("cvr", 0),
                "cpc": n.get("cpc", 0),
                "karar_durumu": "PENDING",
            })
        if neg_data:
            db.insert_negative_candidates(hesap_key, marketplace, today, neg_data)
            logger.info("Supabase: %d negatif aday yazildi", len(neg_data))

        # Harvesting adaylar
        harv_data = []
        for h in harvest_candidates:
            harv_data.append({
                "reklam_tipi": h.get("reklam_tipi", "SP"),
                "campaign_id": h.get("campaign_id"),
                "kaynak_kampanya": h.get("kampanya_adi") or h.get("kaynak_kampanya"),
                "ad_group_id": h.get("ad_group_id"),
                "search_term": h.get("search_term"),
                "hedefleme": h.get("hedefleme"),
                "tip": h.get("tip", "KEYWORD"),
                "match_type": h.get("match_type"),
                "suggested_bid": h.get("suggested_bid"),
                "portfolio": _get_portfolio(h),
                "cvr": h.get("cvr", 0),
                "recommendation": h.get("oneri", "") or h.get("recommendation", ""),
                "impressions": h.get("impressions"),
                "clicks": h.get("clicks"),
                "spend": h.get("spend") or h.get("cost"),
                "sales": h.get("sales"),
                "orders": h.get("orders"),
                "acos": h.get("acos"),
                "karar_durumu": "PENDING",
            })
        if harv_data:
            db.insert_harvesting_candidates(hesap_key, marketplace, today, harv_data)
            logger.info("Supabase: %d harvesting aday yazildi", len(harv_data))

    except Exception as e:
        logger.error("Supabase sync hatasi (analiz devam eder): %s", e)
        save_error_log("InternalError", f"Supabase sync: {e}",
                       traceback.format_exc(), adim="supabase_sync",
                       session_id=MAESTRO_SESSION_ID)
        try:
            db.insert_error_log(hesap_key, marketplace, "agent2", {
                "hata_tipi": "InternalError",
                "hata_mesaji": f"Supabase sync hatasi: {e}"[:500],
                "adim": "supabase_sync",
            })
        except Exception as e:
            logger.warning("Supabase yazim hatasi: %s", e)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Kullanim: python agent2/analyst.py <hesap_key> <marketplace>")
        print("Ornek:    python agent2/analyst.py vigowood_na US")
        sys.exit(1)
    os.environ.setdefault("HESAP_KEY", sys.argv[1])
    os.environ.setdefault("MARKETPLACE", sys.argv[2])
    result = run_analysis(sys.argv[1], sys.argv[2])
    print(json.dumps(result, indent=2, ensure_ascii=False))
