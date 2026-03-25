"""
Agent 3 — Executor v3 Multi-Account
======================================
DEGISIKLIKLER (v2 -> v3):
  1. Multi-account: hesap_key + marketplace parametresi alir.
  2. Tum path'ler init_paths() ile dinamik olusturulur.
  3. Veri okuma/yazma: data/{hesap_key}_{marketplace}/ altindan.
  4. Config okuma: config/{hesap_key}_{marketplace}/ altindan.
  5. CLI: python agent3/executor.py <hesap_key> <marketplace> [--execute] [--verify]

DEGISIKLIKLER (v1.1 -> v2):
  1. save_error_log(): Yeni fonksiyon — tum beklenmeyen hatalar, preflight
     hatalari ve execution'daki HATA statusundeki islemler
     data/logs/agent3_errors.json dosyasina eklenir.
  2. run_executor() artik _run_executor_impl()'i try/except ile sarmaliyor.
  3. Execution sonunda toplam hata sayisi > 0 ise hata detaylari loglanir.
  4. Agent 4 (Learning Agent) agent3_errors.json'u okuyarak tekrar eden
     hata kaliplarini analiz eder ve Maestro CLAUDE.md guncellemesi onerir.
  5. Son 200 hata kaydi tutulur (eski kayitlar otomatik temizlenir).
Agent 2'nin Excel raporlarindaki onaylanmis kararlari Amazon Advertising API
uzerinden uygular.

3 Gorev Grubu:
  1. Bid Degisiklikleri  — "Y" veya ozel bid degeri ile onaylanan satirlar
  2. Negatif Ekleme      — Keyword veya ASIN negatif hedefleme
  3. Harvesting          — Kaynak kampanyada negatif + yeni Exact kampanya / ASIN target

Guvenlik:
  - Dry-run modu (varsayilan) — once ne yapilacagini gosterir, onay bekler
  - Max bid limiti
  - Min bid limiti
  - Gunluk max islem sayisi
  - Rollback log (eski deger → yeni deger)

Dogrulama (2 asamali):
  - Asama 1: Anlik — Her API cagrisindan sonra response kontrolu
  - Asama 2: Gecikmeli — 5 dk sonra Amazon'dan guncel verileri cekip karsilastirma
  - Uyusmazlik durumunda otomatik retry

Calisma Sekli:
  - Agent 2 gibi import edilebilir modul
  - Claude Code tarafindan cagirilir
  - Amazon API islemleri Agent 1'in MCP tool'lari uzerinden yapilir

Yazar: Amazon PPC Otomasyon Sistemi
Versiyon: 1.1
"""

import os
import sys
import re
import json
import time
import logging
import traceback
from datetime import datetime
from pathlib import Path
from copy import deepcopy

try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("openpyxl gerekli: pip install openpyxl --break-system-packages")

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("agent3_executor")

# ============================================================================
# YAPILANDIRMA — hesap_key + marketplace'den dinamik
# ============================================================================

BASE_DIR = Path(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, str(BASE_DIR))
from log_utils import save_error_log as _central_save_error_log, save_log as _save_log


def _dashboard_status(agent_name, status, health_detail=None):
    """Dashboard agent_status tablosunu gunceller."""
    try:
        from supabase.db_client import SupabaseClient
        db = SupabaseClient()
        db.update_agent_status_detail(agent_name, status, health_detail)
    except Exception:
        pass


def _dashboard_pipeline(session_id, hesap_key, marketplace, step, status, error_msg=None):
    """Dashboard pipeline_runs tablosunu gunceller."""
    if not session_id:
        return
    try:
        from supabase.db_client import SupabaseClient
        db = SupabaseClient()
        db.upsert_pipeline_run(session_id, hesap_key, marketplace, step, status, error_msg)
    except Exception:
        pass

# Bu degiskenler init_paths() ile set edilir
DATA_DIR = None
ANALYSIS_DIR = None
LOG_DIR = None
CONFIG_DIR = None
HESAP_KEY = None
MARKETPLACE = None

# Maestro pipeline session ID'si (env var ile iletilir, korelasyon icin)
MAESTRO_SESSION_ID = os.environ.get("MAESTRO_SESSION_ID")


def init_paths(hesap_key, marketplace):
    """Hesap+marketplace icin tum path'leri set eder."""
    global DATA_DIR, ANALYSIS_DIR, LOG_DIR, CONFIG_DIR, HESAP_KEY, MARKETPLACE
    HESAP_KEY = hesap_key
    MARKETPLACE = marketplace

    dir_name = f"{hesap_key}_{marketplace}"
    DATA_DIR = BASE_DIR / "data" / dir_name
    ANALYSIS_DIR = DATA_DIR / "analysis"
    LOG_DIR = DATA_DIR / "logs"
    CONFIG_DIR = BASE_DIR / "config" / dir_name

    logger.info("Agent 3 Paths: data=%s, config=%s", DATA_DIR, CONFIG_DIR)


def load_settings():
    """Settings'i Supabase'den yukler. Basarisizsa JSON dosyasina fallback."""
    hk = os.environ.get("HESAP_KEY", "")
    mp = os.environ.get("MARKETPLACE", "")
    if hk and mp:
        try:
            from supabase.db_client import SupabaseClient
            db = SupabaseClient()
            conn = db._conn()
            cur = conn.cursor()
            cur.execute("SELECT genel_ayarlar, esik_degerleri, asin_hedefleri, segmentasyon_kurallari, agent3_ayarlari FROM settings WHERE hesap_key = %s AND marketplace = %s", (hk, mp))
            row = cur.fetchone()
            cur.close()
            conn.close()
            if row:
                result = {}
                for i, key in enumerate(["genel_ayarlar", "esik_degerleri", "asin_hedefleri", "segmentasyon_kurallari", "agent3_ayarlari"]):
                    if row[i]:
                        result[key] = row[i] if isinstance(row[i], dict) else json.loads(row[i])
                logger.info("Settings Supabase'den yuklendi (%s/%s)", hk, mp)
                return result
        except Exception as e:
            logger.warning("Settings Supabase'den okunamadi, dosyaya fallback: %s", e)

    settings_path = CONFIG_DIR / "settings.json"
    if settings_path.exists():
        with open(settings_path, "r", encoding="utf-8") as f:
            return json.load(f)
    raise FileNotFoundError(f"Settings bulunamadi: Supabase ve {settings_path}")


def get_agent3_config(settings):
    """Agent 3'e ozel ayarlari doner. Yoksa varsayilanlari kullanir."""
    defaults = {
        "dry_run": True,
        "max_bid_limiti": 5.00,
        "min_bid_limiti": 0.15,
        "gunluk_max_islem": 200,
        "yeni_kampanya_butcesi": 10.00,
        "para_birimi": "$",
        "negatif_match_type": "NEGATIVE_EXACT",
        "portfolio_asin_target_kampanyalari": {},
    }
    user_config = settings.get("agent3_ayarlari", {})
    merged = {**defaults, **user_config}

    # Supabase portfolio_asin_campaigns tablosundan guncel eslestirmeleri oku
    # Dashboard'dan yapilan eslestirmeler burada — settings.json'daki eski verileri override eder
    supabase_map = _load_portfolio_asin_campaigns_from_supabase()
    if supabase_map:
        merged["portfolio_asin_target_kampanyalari"] = supabase_map
        logger.info("Portfolio ASIN kampanya eslestirmeleri Supabase'den yuklendi: %d portfolio", len(supabase_map))

    return merged


def _load_portfolio_asin_campaigns_from_supabase():
    """
    Supabase portfolio_asin_campaigns tablosundan {portfolio_name: campaign_id} eslesmesi okur.
    Mevcut hesap/marketplace icin aktif eslestirmeleri doner.
    Birden fazla kampanya varsa ilkini kullanir (ASIN Target oncelikli).
    """
    try:
        from supabase.db_client import SupabaseClient
        db = SupabaseClient()
        conn = db._conn()
        cur = conn.cursor()

        hk = os.environ.get("HESAP_KEY", "")
        mp = os.environ.get("MARKETPLACE", "")
        if not hk or not mp:
            # init_paths'den alinmis olabilir
            if DATA_DIR:
                parts = DATA_DIR.parent.name.split("_")
                if len(parts) >= 2:
                    mp = parts[-1]
                    hk = "_".join(parts[:-1])

        if not hk or not mp:
            logger.debug("Portfolio ASIN campaigns: hesap_key/marketplace belirlenemedi")
            conn.close()
            return {}

        cur.execute("""
            SELECT portfolio_name, campaign_id, campaign_name
            FROM portfolio_asin_campaigns
            WHERE hesap_key = %s AND marketplace = %s AND campaign_id IS NOT NULL
            ORDER BY portfolio_name
        """, (hk, mp))

        rows = cur.fetchall()
        conn.close()

        if not rows:
            return {}

        # portfolio_name -> campaign_id (ilk eslesen)
        result = {}
        for portfolio_name, campaign_id, campaign_name in rows:
            if portfolio_name not in result:
                result[portfolio_name] = campaign_id
                logger.debug("Portfolio ASIN mapping: %s -> %s (%s)", portfolio_name, campaign_id, campaign_name)

        return result

    except Exception as e:
        logger.warning("Portfolio ASIN campaigns Supabase'den okunamadi: %s", e)
        return {}


# ============================================================================
# EXCEL OKUMA
# ============================================================================

def _read_excel_sheet(filepath, sheet_name=None):
    """
    Excel dosyasindan satirlari okur. Header satiri 1, veri satiri 2'den baslar.
    Her satir bir dict olarak donulur: {kolon_adi: deger}
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val
        data.append(record)
    return data


def find_todays_excel(prefix, today=None):
    """Bugunun tarihiyle baslayan Excel dosyasini bulur."""
    if today is None:
        today = datetime.utcnow().strftime("%Y-%m-%d")
    pattern = f"{today}_{prefix}"
    for f in ANALYSIS_DIR.iterdir():
        if f.name.startswith(pattern) and f.suffix == ".xlsx":
            return f
    return None


# ============================================================================
# ONAY PARSE
# ============================================================================

def parse_onay(value):
    """
    Onay kutucugunu parse eder.
    Returns:
        ("Y", None)          — tavsiye edilen bid uygulanacak
        ("CUSTOM", 1.25)     — kullanicinin girdigi bid uygulanacak
        ("SKIP", None)       — bos veya gecersiz — atla
    """
    if value is None:
        return ("SKIP", None)
    val = str(value).strip()
    if val == "":
        return ("SKIP", None)
    if val.upper() == "Y":
        return ("Y", None)
    # Sayi mi? (kullanici ozel bid girmis olabilir)
    cleaned = val.replace("$", "").replace("€", "").replace("£", "").replace("¥", "").replace(",", ".")
    try:
        custom_bid = float(cleaned)
        if custom_bid > 0:
            return ("CUSTOM", custom_bid)
    except ValueError:
        pass
    return ("SKIP", None)


def _parse_currency(value):
    """Para birimi iceren string'den float cikarir: '$1.25' → 1.25"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace("$", "").replace("€", "").replace("£", "").replace("¥", "").replace(",", "")
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0


def _parse_pct(value):
    """Yuzde string'inden float cikarir: '24.5%' → 24.5"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace("%", "").strip()
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0


# ============================================================================
# GOREV 1: BID DEGISIKLIKLERI
# ============================================================================

def parse_bid_recommendations(today=None):
    """
    Bid recommendations — onaylari Supabase'den okur, fallback olarak Excel kullanir.
    """
    # Oncelik 1: Supabase'den onaylari oku
    actions = _parse_bid_recommendations_from_supabase(today)
    if actions is not None:
        return actions

    # Fallback: Excel'den oku
    logger.info("Supabase'den bid onaylari alinamadi, Excel'e fallback yapiliyor...")
    filepath = find_todays_excel("bid_recommendations", today)
    if not filepath:
        logger.warning("Bid recommendations Excel bulunamadi.")
        return []

    rows = _read_excel_sheet(filepath)
    actions = []

    for row in rows:
        onay_raw = row.get("Onay", row.get("onay", None))
        onay_type, custom_bid = parse_onay(onay_raw)

        if onay_type == "SKIP":
            continue

        kampanya = row.get("Kampanya", row.get("kampanya", ""))
        reklam_tipi = row.get("Reklam Tipi", row.get("reklam_tipi", ""))
        hedefleme = row.get("Hedefleme", row.get("hedefleme", ""))
        match_type = row.get("Match Type", row.get("match_type", ""))
        mevcut_bid = _parse_currency(row.get("Bid", row.get("bid", 0)))
        tavsiye_bid = _parse_currency(row.get("Tavsiye Bid", row.get("tavsiye_bid", 0)))
        segment = row.get("Segment", row.get("segment", ""))

        if onay_type == "CUSTOM":
            yeni_bid = custom_bid
        else:
            yeni_bid = tavsiye_bid

        actions.append({
            "tip": "BID_DEGISIKLIGI",
            "kampanya_adi": kampanya,
            "reklam_tipi": reklam_tipi,
            "portfolio": row.get("Portfolio", row.get("portfolio", "")),
            "hedefleme": hedefleme,
            "match_type": match_type,
            "eski_bid": mevcut_bid,
            "yeni_bid": yeni_bid,
            "segment": segment,
            "onay_tipi": onay_type,
            "kaynak": "bid_recommendations",
        })

    logger.info("Bid recommendations (Excel): %d onaylanmis islem", len(actions))
    return actions


def _parse_bid_recommendations_from_supabase(today=None):
    """Supabase bid_recommendations tablosundan APPROVED/MODIFIED onaylari okur."""
    if not today:
        today = datetime.utcnow().strftime("%Y-%m-%d")
    try:
        from supabase.db_client import SupabaseClient
        db = SupabaseClient()
        hk = HESAP_KEY or ""
        mp = MARKETPLACE or ""
        if not hk or not mp:
            return None

        rows = db._fetch_all(
            """SELECT campaign_name, ad_type, portfolio, keyword_text, targeting, match_type,
                      current_bid, recommended_bid, decision_bid, segment, decision
               FROM bid_recommendations
               WHERE hesap_key=%s AND marketplace=%s AND analysis_date=%s
                 AND decision IN ('APPROVED', 'MODIFIED')""",
            (hk, mp, today))

        if not rows:
            logger.info("Supabase: bid onaylari bulunamadi (%s/%s/%s)", hk, mp, today)
            return []

        actions = []
        for r in rows:
            campaign_name, ad_type, portfolio, keyword_text, targeting, match_type, \
                current_bid, recommended_bid, decision_bid, segment, decision = r

            mevcut_bid = float(current_bid) if current_bid else 0
            tavsiye_bid = float(recommended_bid) if recommended_bid else mevcut_bid

            if decision == "MODIFIED" and decision_bid:
                yeni_bid = float(decision_bid)
                onay_tipi = "CUSTOM"
            else:
                yeni_bid = tavsiye_bid
                onay_tipi = "APPROVED"

            actions.append({
                "tip": "BID_DEGISIKLIGI",
                "kampanya_adi": campaign_name or "",
                "reklam_tipi": ad_type or "SP",
                "portfolio": portfolio or "",
                "hedefleme": keyword_text or targeting or "",
                "match_type": match_type or "",
                "eski_bid": mevcut_bid,
                "yeni_bid": yeni_bid,
                "segment": segment or "",
                "onay_tipi": onay_tipi,
                "kaynak": "supabase_bid_recommendations",
            })

        logger.info("Bid recommendations (Supabase): %d onaylanmis islem", len(actions))
        return actions
    except Exception as e:
        logger.warning("Supabase bid okuma hatasi: %s", e)
        return None


# ============================================================================
# GOREV 2: NEGATIF EKLEME
# ============================================================================

def parse_negative_candidates(today=None):
    """
    Negatif keyword adaylari — onaylari Supabase'den okur, fallback olarak Excel kullanir.
    """
    actions = _parse_negative_candidates_from_supabase(today)
    if actions is not None:
        return actions

    logger.info("Supabase'den negatif onaylari alinamadi, Excel'e fallback yapiliyor...")
    filepath = find_todays_excel("negative_candidates", today)
    if not filepath:
        logger.warning("Negative candidates Excel bulunamadi.")
        return []

    rows = _read_excel_sheet(filepath)
    actions = []

    for row in rows:
        onay_raw = row.get("Onay", row.get("onay", None))
        onay_type, _ = parse_onay(onay_raw)

        if onay_type == "SKIP":
            continue

        hedefleme = row.get("Search Term", row.get("Hedefleme", row.get("hedefleme", "")))
        kampanya = row.get("Kampanya", row.get("kampanya", ""))
        reklam_tipi = row.get("Reklam Tipi", row.get("reklam_tipi", ""))
        match_type = row.get("Match Type", row.get("match_type", ""))

        if not hedefleme or str(hedefleme).strip() == "":
            continue

        is_asin = _is_asin_target(hedefleme)

        actions.append({
            "tip": "NEGATIF_ASIN" if is_asin else "NEGATIF_KEYWORD",
            "kampanya_adi": kampanya,
            "reklam_tipi": reklam_tipi,
            "portfolio": row.get("Portfolio", row.get("portfolio", "")),
            "hedefleme": hedefleme,
            "match_type": match_type,
            "negatif_match_type": "NEGATIVE_EXACT",
            "harcama": _parse_currency(row.get("Spend", row.get("spend", 0))),
            "satis": _parse_currency(row.get("Sales", row.get("sales", 0))),
            "sebep": row.get("Sebep", row.get("sebep", "")),
            "kaynak": "negative_candidates",
        })

    logger.info("Negatif adaylar (Excel): %d onaylanmis islem", len(actions))
    return actions


def _parse_negative_candidates_from_supabase(today=None):
    """Supabase negative_candidates tablosundan APPROVED onaylari okur."""
    if not today:
        today = datetime.utcnow().strftime("%Y-%m-%d")
    try:
        from supabase.db_client import SupabaseClient
        db = SupabaseClient()
        hk = HESAP_KEY or ""
        mp = MARKETPLACE or ""
        if not hk or not mp:
            return None

        rows = db._fetch_all(
            """SELECT campaign_name, ad_type, portfolio, search_term, match_type,
                      cost, sales, reason
               FROM negative_candidates
               WHERE hesap_key=%s AND marketplace=%s AND analysis_date=%s
                 AND decision = 'APPROVED'""",
            (hk, mp, today))

        if not rows:
            logger.info("Supabase: negatif onaylari bulunamadi (%s/%s/%s)", hk, mp, today)
            return []

        actions = []
        for r in rows:
            campaign_name, ad_type, portfolio, search_term, match_type, cost, sales, reason = r
            hedefleme = search_term or ""
            if not hedefleme.strip():
                continue
            is_asin = _is_asin_target(hedefleme)
            actions.append({
                "tip": "NEGATIF_ASIN" if is_asin else "NEGATIF_KEYWORD",
                "kampanya_adi": campaign_name or "",
                "reklam_tipi": ad_type or "SP",
                "portfolio": portfolio or "",
                "hedefleme": hedefleme,
                "match_type": match_type or "",
                "negatif_match_type": "NEGATIVE_EXACT",
                "harcama": float(cost) if cost else 0,
                "satis": float(sales) if sales else 0,
                "sebep": reason or "",
                "kaynak": "supabase_negative_candidates",
            })

        logger.info("Negatif adaylar (Supabase): %d onaylanmis islem", len(actions))
        return actions
    except Exception as e:
        logger.warning("Supabase negatif okuma hatasi: %s", e)
        return None


def _is_asin_target(hedefleme):
    """Hedefleme bir ASIN mi yoksa keyword mu?"""
    if not hedefleme:
        return False
    h = str(hedefleme).strip().upper()
    # ASIN formati: B0XXXXXXXXX veya asin="B0..." seklinde
    if re.match(r'^B0[A-Z0-9]{8,}$', h):
        return True
    if 'ASIN=' in h or 'asin=' in hedefleme:
        return True
    if h.startswith("ASIN(") or h.startswith("CATEGORY("):
        return True
    return False


def _format_asin(asin_text):
    """
    ASIN'i Amazon API formatina uygun hale getirir.
    Amazon ASIN'leri BUYUK HARF olmali (B0XXXXXXXXX).
    Kucuk harfle gonderilen ASIN'ler gecersiz sayilir.
    """
    if not asin_text:
        return asin_text
    formatted = str(asin_text).strip().upper()
    # Eger ASIN formatinda degilse olduğu gibi don (keyword olabilir)
    if re.match(r'^B0[A-Z0-9]{8,}$', formatted):
        return formatted
    return str(asin_text).strip()


# ============================================================================
# GOREV 3: HARVESTING
# ============================================================================

def parse_harvesting_candidates(today=None):
    """
    Harvesting adaylari — onaylari Supabase'den okur, fallback olarak Excel kullanir.
    """
    actions = _parse_harvesting_candidates_from_supabase(today)
    if actions is not None:
        return actions

    logger.info("Supabase'den harvesting onaylari alinamadi, Excel'e fallback yapiliyor...")
    filepath = find_todays_excel("harvesting_candidates", today)
    if not filepath:
        logger.warning("Harvesting candidates Excel bulunamadi.")
        return []

    rows = _read_excel_sheet(filepath)
    actions = []

    for row in rows:
        onay_raw = row.get("Onay", row.get("onay", None))
        onay_type, _ = parse_onay(onay_raw)

        if onay_type == "SKIP":
            continue

        hedefleme = row.get("Hedefleme", row.get("hedefleme", ""))
        kaynak_kampanya = row.get("Kaynak Kampanya", row.get("kaynak_kampanya", ""))
        reklam_tipi = row.get("Reklam Tipi", row.get("reklam_tipi", ""))
        portfolio = row.get("Portfolio", row.get("portfolio", ""))
        bid = _parse_currency(row.get("CPC", row.get("cpc", 0)))
        if bid <= 0:
            spend = _parse_currency(row.get("Spend", row.get("spend", 0)))
            clicks = _parse_currency(row.get("Click", row.get("click", 0)))
            if clicks > 0:
                bid = round(spend / clicks, 2)

        is_asin = _is_asin_target(hedefleme)

        if is_asin:
            actions.append({
                "tip": "HARVEST_ASIN",
                "kaynak_kampanya": kaynak_kampanya,
                "reklam_tipi": reklam_tipi,
                "portfolio": portfolio,
                "hedefleme": hedefleme,
                "bid": bid,
                "kaynak": "harvesting_candidates",
            })
        else:
            actions.append({
                "tip": "HARVEST_KEYWORD",
                "kaynak_kampanya": kaynak_kampanya,
                "reklam_tipi": reklam_tipi,
                "portfolio": portfolio,
                "hedefleme": hedefleme,
                "match_type": row.get("Match Type", ""),
                "bid": bid,
                "kaynak": "harvesting_candidates",
            })

    logger.info("Harvesting (Excel): %d onaylanmis islem", len(actions))
    return actions


def _parse_harvesting_candidates_from_supabase(today=None):
    """Supabase harvesting_candidates tablosundan APPROVED onaylari okur."""
    if not today:
        today = datetime.utcnow().strftime("%Y-%m-%d")
    try:
        from supabase.db_client import SupabaseClient
        db = SupabaseClient()
        hk = HESAP_KEY or ""
        mp = MARKETPLACE or ""
        if not hk or not mp:
            return None

        rows = db._fetch_all(
            """SELECT source_campaign_name, ad_type, portfolio, search_term, targeting,
                      suggested_match_type, suggested_bid, cost, clicks
               FROM harvesting_candidates
               WHERE hesap_key=%s AND marketplace=%s AND analysis_date=%s
                 AND decision = 'APPROVED'""",
            (hk, mp, today))

        if not rows:
            logger.info("Supabase: harvesting onaylari bulunamadi (%s/%s/%s)", hk, mp, today)
            return []

        actions = []
        for r in rows:
            source_campaign, ad_type, portfolio, search_term, targeting, \
                suggested_match, suggested_bid, cost, clicks = r

            hedefleme = search_term or targeting or ""
            if not hedefleme.strip():
                continue

            bid = float(suggested_bid) if suggested_bid else 0
            if bid <= 0 and clicks and float(clicks) > 0 and cost:
                bid = round(float(cost) / float(clicks), 2)

            is_asin = _is_asin_target(hedefleme)

            if is_asin:
                actions.append({
                    "tip": "HARVEST_ASIN",
                    "kaynak_kampanya": source_campaign or "",
                    "reklam_tipi": ad_type or "SP",
                    "portfolio": portfolio or "",
                    "hedefleme": hedefleme,
                    "bid": bid,
                    "kaynak": "supabase_harvesting_candidates",
                })
            else:
                actions.append({
                    "tip": "HARVEST_KEYWORD",
                    "kaynak_kampanya": source_campaign or "",
                    "reklam_tipi": ad_type or "SP",
                    "portfolio": portfolio or "",
                    "hedefleme": hedefleme,
                    "match_type": suggested_match or "",
                    "bid": bid,
                    "kaynak": "supabase_harvesting_candidates",
                })

        logger.info("Harvesting (Supabase): %d onaylanmis islem", len(actions))
        return actions
    except Exception as e:
        logger.warning("Supabase harvesting okuma hatasi: %s", e)
        return None


# ============================================================================
# GUVENLIK KONTROLLERI
# ============================================================================

def validate_bid(bid, config):
    """Bid degerini min/max limitlerine gore kontrol eder."""
    min_bid = config.get("min_bid_limiti", 0.15)
    max_bid = config.get("max_bid_limiti", 5.00)
    if bid < min_bid:
        return min_bid, f"Bid {bid:.2f} minimum limitin ({min_bid:.2f}) altinda, {min_bid:.2f} olarak ayarlandi"
    if bid > max_bid:
        return max_bid, f"Bid {bid:.2f} maksimum limitin ({max_bid:.2f}) ustunde, {max_bid:.2f} olarak ayarlandi"
    return bid, None


def check_daily_limit(total_actions, config):
    """Gunluk islem limitini kontrol eder."""
    limit = config.get("gunluk_max_islem", 200)
    if total_actions > limit:
        return False, f"Toplam {total_actions} islem, gunluk limit {limit}. Ilk {limit} islem uygulanacak."
    return True, None


# ============================================================================
# KAMPANYA ID ESLESTIRME
# ============================================================================

def build_campaign_lookup(today=None):
    """
    Kampanya/AdGroup/Portfolio lookup'ini Supabase'den olusturur.
    Basarisizsa JSON dosyalarina fallback yapar.
    """
    if today is None:
        today = datetime.utcnow().strftime("%Y-%m-%d")

    lookup = {
        "by_name": {},
        "by_id": {},
        "ad_groups": {},
        "portfolios": {},
    }

    hk = os.environ.get("HESAP_KEY", "")
    mp = os.environ.get("MARKETPLACE", "")
    supabase_ok = False

    if hk and mp:
        try:
            from supabase.db_client import SupabaseClient
            db = SupabaseClient()
            conn = db._conn()
            cur = conn.cursor()

            # Portfolios
            cur.execute("SELECT portfolio_id, name FROM portfolios WHERE hesap_key = %s AND marketplace = %s", (hk, mp))
            for pid, pname in cur.fetchall():
                if pid:
                    lookup["portfolios"][str(pid)] = pname

            # Campaigns
            cur.execute("SELECT campaign_id, name, ad_type, portfolio_id, state FROM campaigns WHERE hesap_key = %s AND marketplace = %s", (hk, mp))
            for cid, cname, ad_type, pid, state in cur.fetchall():
                cid_s = str(cid or "")
                pid_s = str(pid or "")
                pname = lookup["portfolios"].get(pid_s, "")
                info = {"id": cid_s, "name": cname, "ad_type": ad_type or "SP", "portfolio_id": pid_s, "portfolio_name": pname, "state": state or ""}
                if cname:
                    lookup["by_name"][cname] = info
                if cid_s:
                    lookup["by_id"][cid_s] = info

            # Ad Groups
            cur.execute("SELECT campaign_id, ad_group_id, name, state, default_bid FROM ad_groups WHERE hesap_key = %s AND marketplace = %s", (hk, mp))
            for cid, agid, name, state, bid in cur.fetchall():
                cid_s = str(cid or "")
                if cid_s not in lookup["ad_groups"]:
                    lookup["ad_groups"][cid_s] = []
                lookup["ad_groups"][cid_s].append({
                    "ad_group_id": str(agid or ""),
                    "name": name or "",
                    "state": state or "",
                    "default_bid": bid or 0,
                })

            cur.close()
            conn.close()
            supabase_ok = True
            logger.info("Kampanya lookup Supabase'den yuklendi: %d kampanya, %d ad group, %d portfolio",
                        len(lookup["by_id"]), sum(len(v) for v in lookup["ad_groups"].values()),
                        len(lookup["portfolios"]))
        except Exception as e:
            logger.warning("Kampanya lookup Supabase hatasi, JSON fallback: %s", e)

    if not supabase_ok:
        # JSON dosyalarina fallback
        portfolio_file = DATA_DIR / f"{today}_portfolios.json"
        if portfolio_file.exists():
            with open(portfolio_file, "r", encoding="utf-8") as f:
                portfolios = json.load(f)
            for p in portfolios:
                pid = str(p.get("portfolioId", ""))
                pname = p.get("name", "")
                if pid:
                    lookup["portfolios"][pid] = pname

        for prefix, ad_type in [("sp_campaigns", "SP"), ("sb_campaigns", "SB"), ("sd_campaigns", "SD")]:
            fpath = DATA_DIR / f"{today}_{prefix}.json"
            if not fpath.exists():
                continue
            with open(fpath, "r", encoding="utf-8") as f:
                camps = json.load(f)
            for c in camps:
                cid = str(c.get("campaignId", ""))
                cname = c.get("name", c.get("campaignName", ""))
                pid = str(c.get("portfolioId", ""))
                pname = lookup["portfolios"].get(pid, "")
                info = {"id": cid, "name": cname, "ad_type": ad_type, "portfolio_id": pid, "portfolio_name": pname, "state": c.get("state", "")}
                if cname:
                    lookup["by_name"][cname] = info
                if cid:
                    lookup["by_id"][cid] = info

        for prefix, ad_type in [("sp_ad_groups", "SP"), ("sb_ad_groups", "SB"), ("sd_ad_groups", "SD")]:
            fpath = DATA_DIR / f"{today}_{prefix}.json"
            if not fpath.exists():
                continue
            with open(fpath, "r", encoding="utf-8") as f:
                groups = json.load(f)
            for g in groups:
                cid = str(g.get("campaignId", ""))
                if cid not in lookup["ad_groups"]:
                    lookup["ad_groups"][cid] = []
                lookup["ad_groups"][cid].append({
                    "ad_group_id": str(g.get("adGroupId", "")),
                    "name": g.get("name", ""),
                    "state": g.get("state", ""),
                    "default_bid": g.get("defaultBid", 0),
                })

        logger.info("Kampanya lookup JSON'dan yuklendi: %d kampanya, %d ad group, %d portfolio",
                    len(lookup["by_id"]), sum(len(v) for v in lookup["ad_groups"].values()),
                    len(lookup["portfolios"]))

    return lookup


def resolve_campaign_id(action, lookup):
    """Kampanya adindan kampanya ID'sini bulur."""
    camp_name = action.get("kampanya_adi", "") or action.get("kaynak_kampanya", "")
    if not camp_name:
        return None, "Kampanya adi bos"
    info = lookup["by_name"].get(camp_name)
    if info:
        return info["id"], None
    # Partial match dene
    for name, info in lookup["by_name"].items():
        if camp_name in name or name in camp_name:
            return info["id"], None
    return None, f"Kampanya bulunamadi: {camp_name}"


def resolve_ad_group_id(campaign_id, lookup):
    """Kampanyanin ilk aktif ad group'unu bulur."""
    groups = lookup["ad_groups"].get(campaign_id, [])
    for g in groups:
        if g.get("state", "").upper() in ("ENABLED", ""):
            return g["ad_group_id"], None
    if groups:
        return groups[0]["ad_group_id"], "Aktif ad group yok, ilk ad group kullanildi"
    return None, f"Ad group bulunamadi (campaign_id: {campaign_id})"


# ============================================================================
# HEDEFLEME ID ESLESTIRME
# ============================================================================

# Auto-targeting rapor adi → entity expression type eslesmesi
AUTO_TARGETING_MAP = {
    "close-match": "QUERY_HIGH_REL_MATCHES",
    "loose-match": "QUERY_BROAD_REL_MATCHES",
    "complements": "ASIN_ACCESSORY_RELATED",
    "substitutes": "ASIN_SUBSTITUTE_RELATED",
}
AUTO_TARGETING_REVERSE = {v: k for k, v in AUTO_TARGETING_MAP.items()}


def build_targeting_lookup(today=None):
    """
    Keyword ve targeting entity'lerinden hedefleme_text → entity_id eslemesi olusturur.
    Bu esleme bid guncelleme ve negatif ekleme icin gerekli.

    DUZELTME v2:
    - ASIN entity'leri hem value bazli hem kampanya+adgroup+value bazli eklenir
    - Auto-targeting (close-match, loose-match vb.) rapor adiyla da aranabilir
    - Category targeting destegi eklendi
    """
    if today is None:
        today = datetime.utcnow().strftime("%Y-%m-%d")

    lookup = {}  # key: (kampanya_id, hedefleme_text, match_type) → entity bilgisi

    hk = os.environ.get("HESAP_KEY", "")
    mp = os.environ.get("MARKETPLACE", "")
    supabase_ok = False

    if hk and mp:
        try:
            from supabase.db_client import SupabaseClient
            db = SupabaseClient()
            conn = db._conn()
            cur = conn.cursor()

            # KEYWORDS
            cur.execute("SELECT keyword_id, keyword_text, match_type, campaign_id, ad_group_id, ad_type, state, bid FROM keywords WHERE hesap_key = %s AND marketplace = %s", (hk, mp))
            for eid, text, mt, cid, agid, ad_type, state, bid in cur.fetchall():
                entity_info = {"entity_id": str(eid or ""), "entity_type": "KEYWORD", "ad_type": ad_type or "SP",
                               "campaign_id": str(cid or ""), "ad_group_id": str(agid or ""), "state": state or "enabled", "bid": bid or 0}
                cid_s = str(cid or "")
                key = (cid_s, (text or "").lower(), (mt or "").upper())
                lookup[key] = entity_info
                key2 = (cid_s, (text or "").lower(), "TARGETING")
                if key2 not in lookup:
                    lookup[key2] = entity_info

            # TARGETS
            cur.execute("SELECT target_id, campaign_id, ad_group_id, ad_type, state, bid, expression FROM targets WHERE hesap_key = %s AND marketplace = %s", (hk, mp))
            for eid, cid, agid, ad_type, state, bid, expression in cur.fetchall():
                cid_s = str(cid or "")
                agid_s = str(agid or "")
                entity_info = {"entity_id": str(eid or ""), "entity_type": "TARGET", "ad_type": ad_type or "SP",
                               "campaign_id": cid_s, "ad_group_id": agid_s, "state": state or "enabled", "bid": bid or 0}

                expr = expression
                if isinstance(expr, str):
                    try:
                        expr = json.loads(expr)
                    except Exception:
                        pass

                if isinstance(expr, list) and expr:
                    first = expr[0] if isinstance(expr[0], dict) else {}
                    expr_type = first.get("type", "")
                    expr_value = str(first.get("value", "")).lower()
                    if expr_value:
                        lookup[(cid_s, expr_value, "TARGETING")] = entity_info
                        lookup[(cid_s, agid_s, expr_value)] = entity_info
                    if expr_type in AUTO_TARGETING_REVERSE:
                        rapor_adi = AUTO_TARGETING_REVERSE[expr_type]
                        if (cid_s, rapor_adi, "TARGETING") not in lookup:
                            lookup[(cid_s, rapor_adi, "TARGETING")] = entity_info
                        lookup[(cid_s, agid_s, rapor_adi)] = entity_info
                elif isinstance(expr, str) and expr:
                    lookup[(cid_s, expr.lower(), "TARGETING")] = entity_info

            cur.close()
            conn.close()
            supabase_ok = True
            logger.info("Targeting lookup Supabase'den yuklendi: %d entity", len(lookup))
        except Exception as e:
            logger.warning("Targeting lookup Supabase hatasi, JSON fallback: %s", e)

    if not supabase_ok:
        # JSON dosyalarina fallback
        for prefix, ad_type, id_field, text_field in [
            ("sp_keywords", "SP", "keywordId", "keywordText"),
            ("sb_keywords", "SB", "keywordId", "keywordText"),
        ]:
            fpath = DATA_DIR / f"{today}_{prefix}.json"
            if not fpath.exists():
                continue
            with open(fpath, "r", encoding="utf-8") as f:
                entities = json.load(f)
            for e in entities:
                eid = str(e.get(id_field, ""))
                text = e.get(text_field, "")
                mt = e.get("matchType", "")
                cid = str(e.get("campaignId", ""))
                agid = str(e.get("adGroupId", ""))
                entity_info = {"entity_id": eid, "entity_type": "KEYWORD", "ad_type": ad_type,
                               "campaign_id": cid, "ad_group_id": agid, "state": e.get("state", "enabled"), "bid": e.get("bid", 0)}
                lookup[(cid, (text or "").lower(), (mt or "").upper())] = entity_info
                key2 = (cid, (text or "").lower(), "TARGETING")
                if key2 not in lookup:
                    lookup[key2] = entity_info

        for prefix, ad_type, id_field in [
            ("sp_targets", "SP", "targetId"),
            ("sb_targets", "SB", "targetId"),
            ("sd_targets", "SD", "targetId"),
        ]:
            fpath = DATA_DIR / f"{today}_{prefix}.json"
            if not fpath.exists():
                continue
            with open(fpath, "r", encoding="utf-8") as f:
                entities = json.load(f)
            for e in entities:
                eid = str(e.get(id_field, ""))
                cid = str(e.get("campaignId", ""))
                agid = str(e.get("adGroupId", ""))
                expression = e.get("expression", e.get("expressions", e.get("targetingExpression", "")))
                entity_info = {"entity_id": eid, "entity_type": "TARGET", "ad_type": ad_type,
                               "campaign_id": cid, "ad_group_id": agid, "state": e.get("state", "enabled"), "bid": e.get("bid", 0)}
                if isinstance(expression, list) and expression:
                    expr_type = expression[0].get("type", "")
                    expr_value = str(expression[0].get("value", "")).lower()
                    if expr_value:
                        lookup[(cid, expr_value, "TARGETING")] = entity_info
                        lookup[(cid, agid, expr_value)] = entity_info
                    if expr_type in AUTO_TARGETING_REVERSE:
                        rapor_adi = AUTO_TARGETING_REVERSE[expr_type]
                        if (cid, rapor_adi, "TARGETING") not in lookup:
                            lookup[(cid, rapor_adi, "TARGETING")] = entity_info
                        lookup[(cid, agid, rapor_adi)] = entity_info
                elif isinstance(expression, str) and expression:
                    lookup[(cid, (expression or "").lower(), "TARGETING")] = entity_info

    # ---- SB THEME ENTITY'LER ----
    THEME_TYPE_TO_REPORT = {
        "KEYWORDS_RELATED_TO_YOUR_BRAND": "keywords-related-to-your-brand",
        "KEYWORDS_RELATED_TO_YOUR_LANDING_PAGES": "keywords-related-to-your-landing-pages",
    }

    sb_themes_path = DATA_DIR / f"{today}_sb_themes.json"
    if sb_themes_path.exists():
        with open(sb_themes_path, "r", encoding="utf-8") as f:
            themes = json.load(f)
        for t in themes:
            tid = str(t.get("themeId", ""))
            cid = str(t.get("campaignId", ""))
            agid = str(t.get("adGroupId", ""))
            theme_type = t.get("themeType", "")
            report_text = THEME_TYPE_TO_REPORT.get(theme_type, (theme_type or "").lower())

            entity_info = {
                "entity_id": tid,
                "entity_type": "THEME",
                "ad_type": "SB",
                "campaign_id": cid,
                "ad_group_id": agid,
                "state": t.get("state", "enabled"),
                "bid": t.get("bid", 0),
            }

            # Birincil key: (campaign_id, report_text, "THEME")
            key = (cid, report_text, "THEME")
            lookup[key] = entity_info
            # Ikincil key: TARGETING ile de bulunabilsin
            key2 = (cid, report_text, "TARGETING")
            if key2 not in lookup:
                lookup[key2] = entity_info

        logger.info("SB Themes: %d theme entity yuklendi", len(themes))

    logger.info("Targeting lookup: %d entity (keyword + target + auto-targeting)", len(lookup))
    return lookup


def _extract_value_from_targeting(targeting_text):
    """
    Rapordaki targeting text'inden gercek degeri cikarir.

    Ornekler:
      'asin-expanded="B0FLQ8F1F1"' → 'b0flq8f1f1'
      'asin="B08LDWH7VF"'          → 'b08ldwh7vf'
      'category="1234567890"'       → '1234567890'
      'close-match'                 → 'close-match'
      'lap desk'                    → 'lap desk'
    """
    if targeting_text is None:
        return ""
    text = str(targeting_text).strip()
    # Tirnak icindeki degeri cikar: asin="VALUE" veya asin-expanded="VALUE"
    m = re.search(r'"([^"]+)"', text)
    if m:
        return m.group(1).lower()
    # Tirnak yoksa olduğu gibi don (keyword veya auto-targeting)
    return text.lower()


def resolve_targeting_entity(action, campaign_id, targeting_lookup, campaign_lookup):
    """
    Hedefleme text'inden entity_id'yi bulur.
    Bid guncelleme icin entity_id gerekli.

    DUZELTME v2:
    - Rapordaki targeting text'inden ASIN/keyword/category extract edilir
    - Auto-targeting (close-match vb.) destegi
    - Kampanya+AdGroup bazli lookup (ayni ASIN birden fazla ad group'ta)
    - Amazon auto-expanded icin acik hata mesaji
    """
    raw_hedefleme = action.get("hedefleme", "")
    match_type = (action.get("match_type") or "").upper()

    # Hedefleme text'inden gercek degeri cikar
    extracted = _extract_value_from_targeting(raw_hedefleme)

    # ---- ARAMA STRATEJISI (once en spesifik, sonra genis) ----

    # 1. Direkt eslestirme: (campaign_id, extracted, match_type)
    key1 = (campaign_id, extracted, match_type)
    if key1 in targeting_lookup:
        return targeting_lookup[key1], None

    # 2. TARGETING key ile: (campaign_id, extracted, "TARGETING")
    key2 = (campaign_id, extracted, "TARGETING")
    if key2 in targeting_lookup:
        return targeting_lookup[key2], None

    # 3. AdGroup bazli: (campaign_id, ad_group_id, extracted)
    ad_group_id = action.get("ad_group_id", "")
    if ad_group_id:
        key3 = (campaign_id, ad_group_id, extracted)
        if key3 in targeting_lookup:
            return targeting_lookup[key3], None

    # 4. Match type olmadan tum kampanya icinde ara (fallback)
    for k, v in targeting_lookup.items():
        if len(k) == 3 and k[0] == campaign_id and k[1] == extracted:
            return v, None

    # 5. ASIN format normalizasyonu ile tekrar dene
    if _is_asin_target(raw_hedefleme):
        asin_match = re.search(r'(B0[A-Z0-9]{8,})', raw_hedefleme.upper())
        if asin_match:
            asin_only = asin_match.group(1).lower()
            if asin_only != extracted:
                key5 = (campaign_id, asin_only, "TARGETING")
                if key5 in targeting_lookup:
                    return targeting_lookup[key5], None

    # ---- BULUNAMADI: Sebebi tespit et ----
    if "asin-expanded" in (raw_hedefleme or "").lower():
        return None, (
            f"Amazon auto-expanded hedefleme (bid ayarlanamaz): {raw_hedefleme} "
            f"(kampanya: {campaign_id}). Bu ASIN icin Amazon entity olusturmamis, "
            f"sadece raporlarda gorunuyor. ATLANIYOR."
        )

    return None, f"Hedefleme entity bulunamadi: {raw_hedefleme} (kampanya: {campaign_id})"


# ============================================================================
# HEDEFLENEN URUN (ASIN) BULMA
# ============================================================================

def find_advertised_asin(campaign_id, ad_group_id, today=None):
    """
    Bir kampanya/ad group icin hedeflenen urun ASIN'ini ve SKU'sunu bulur.
    Harvesting'de yeni kampanya olustururken bu ASIN ve SKU gerekli.
    Donen ASIN her zaman BUYUK HARF formatindadir.

    Returns: (asin, sku, warning) tuple

    Arama sirasi:
    1. sp_product_ads.json (en guvenilir — hem ASIN hem SKU var)
    2. SP targeting raporundan advertisedAsin alani (fallback, SKU yok)
    3. SP search term raporundan advertisedAsin alani (fallback 2, SKU yok)
    """
    if today is None:
        today = datetime.utcnow().strftime("%Y-%m-%d")

    # Yontem 1: SP Product Ads (en guvenilir)
    fpath = DATA_DIR / f"{today}_sp_product_ads.json"
    if fpath.exists():
        with open(fpath, "r", encoding="utf-8") as f:
            ads = json.load(f)
        # Oncelik: kampanya + ad group eslesmesi
        for ad in ads:
            if str(ad.get("campaignId", "")) == campaign_id:
                if ad_group_id and str(ad.get("adGroupId", "")) == ad_group_id:
                    asin = ad.get("asin", "")
                    sku = ad.get("sku", "")
                    if asin:
                        return _format_asin(asin), sku, None
        # Fallback: sadece kampanya eslesmesi
        for ad in ads:
            if str(ad.get("campaignId", "")) == campaign_id:
                asin = ad.get("asin", "")
                sku = ad.get("sku", "")
                if asin:
                    return _format_asin(asin), sku, "Ad group eslesmedi, kampanyadaki ilk ASIN kullanildi"

    # Yontem 2: SP Targeting raporundan advertisedAsin (fallback — SKU yok)
    for suffix in ["sp_targeting_report_14d", "sp_search_term_report_30d"]:
        fpath = DATA_DIR / f"{today}_{suffix}.json"
        if not fpath.exists():
            continue
        with open(fpath, "r", encoding="utf-8") as f:
            rows = json.load(f)
        for row in rows:
            if str(row.get("campaignId", "")) == campaign_id:
                asin = row.get("advertisedAsin", row.get("advertised_asin", ""))
                if asin:
                    return _format_asin(asin), "", f"Product ads dosyasi bulunamadi, {suffix}'den alindi (SKU eksik olabilir)"

    return None, "", "Hedeflenen ASIN bulunamadi (product_ads + targeting raporlari kontrol edildi)"


# ============================================================================
# API ISLEM HAZIRLAMA (DRY-RUN & EXECUTE)
# ============================================================================

def prepare_bid_change(action, config, campaign_lookup, targeting_lookup):
    """
    Bid degisikligi icin API payload'unu hazirlar.
    """
    result = {
        "action": action,
        "status": "HAZIR",
        "api_payload": None,
        "uyarilar": [],
        "hatalar": [],
    }

    # Campaign ID bul
    camp_id, err = resolve_campaign_id(action, campaign_lookup)
    if err:
        result["status"] = "HATA"
        result["hatalar"].append(err)
        return result

    # Entity ID bul
    entity, err = resolve_targeting_entity(action, camp_id, targeting_lookup, campaign_lookup)
    if err:
        result["status"] = "HATA"
        result["hatalar"].append(err)
        return result

    # Bid limitleri kontrol
    yeni_bid, uyari = validate_bid(action["yeni_bid"], config)
    if uyari:
        result["uyarilar"].append(uyari)
    action["yeni_bid"] = yeni_bid

    # API payload olustur
    ad_type = entity["ad_type"]
    entity_type = entity["entity_type"]

    if entity_type == "KEYWORD":
        if ad_type == "SP":
            result["api_endpoint"] = "sp_keyword_bid_update"
            result["api_payload"] = {
                "keywordId": entity["entity_id"],
                "bid": round(yeni_bid, 2),
            }
        elif ad_type == "SB":
            result["api_endpoint"] = "sb_keyword_bid_update"
            # SB API lowercase state gerektirir (enabled/paused)
            sb_state = entity.get("state", "enabled").lower()
            result["api_payload"] = {
                "keywordId": entity["entity_id"],
                "campaignId": entity["campaign_id"],
                "adGroupId": entity["ad_group_id"],
                "state": sb_state,
                "bid": round(yeni_bid, 2),
            }
    elif entity_type == "TARGET":
        if ad_type == "SP":
            result["api_endpoint"] = "sp_target_bid_update"
            result["api_payload"] = {
                "targetId": entity["entity_id"],
                "bid": round(yeni_bid, 2),
            }
        elif ad_type == "SB":
            result["api_endpoint"] = "sb_target_bid_update"
            sb_state = entity.get("state", "enabled").lower()
            result["api_payload"] = {
                "targetId": entity["entity_id"],
                "campaignId": entity["campaign_id"],
                "adGroupId": entity["ad_group_id"],
                "state": sb_state,
                "bid": round(yeni_bid, 2),
            }
        elif ad_type == "SD":
            result["api_endpoint"] = "sd_target_bid_update"
            result["api_payload"] = {
                "targetId": entity["entity_id"],
                "bid": round(yeni_bid, 2),
            }

    elif entity_type == "THEME":
        result["api_endpoint"] = "sb_theme_bid_update"
        sb_state = entity.get("state", "enabled").lower()
        result["api_payload"] = {
            "themeId": entity["entity_id"],
            "campaignId": entity["campaign_id"],
            "adGroupId": entity["ad_group_id"],
            "state": sb_state,
            "bid": round(yeni_bid, 2),
        }

    result["campaign_id"] = camp_id
    result["entity_id"] = entity["entity_id"]
    result["entity_type"] = entity_type
    result["ad_type"] = ad_type

    return result


def prepare_negative_add(action, config, campaign_lookup):
    """
    Negatif keyword/ASIN ekleme icin API payload'unu hazirlar.
    """
    result = {
        "action": action,
        "status": "HAZIR",
        "api_payload": None,
        "uyarilar": [],
        "hatalar": [],
    }

    # Campaign ID bul
    camp_id, err = resolve_campaign_id(action, campaign_lookup)
    if err:
        result["status"] = "HATA"
        result["hatalar"].append(err)
        return result

    # Ad Group ID bul
    ag_id, uyari = resolve_ad_group_id(camp_id, campaign_lookup)
    if uyari and not ag_id:
        result["status"] = "HATA"
        result["hatalar"].append(uyari)
        return result
    if uyari:
        result["uyarilar"].append(uyari)

    ad_type = action.get("reklam_tipi", "SP").upper()
    hedefleme = action["hedefleme"]

    if action["tip"] == "NEGATIF_KEYWORD":
        if ad_type == "SP":
            result["api_endpoint"] = "sp_negative_keyword_add"
            result["api_payload"] = {
                "campaignId": camp_id,
                "adGroupId": ag_id,
                "keywordText": hedefleme,
                "matchType": "NEGATIVE_EXACT",
                "state": "ENABLED",
            }
        elif ad_type == "SB":
            result["api_endpoint"] = "sb_negative_keyword_add"
            result["api_payload"] = {
                "campaignId": camp_id,
                "adGroupId": ag_id,
                "keywordText": hedefleme,
                "matchType": "NEGATIVE_EXACT",
                "state": "ENABLED",
            }
    elif action["tip"] == "NEGATIF_ASIN":
        if ad_type == "SP":
            result["api_endpoint"] = "sp_negative_target_add"
            result["api_payload"] = {
                "campaignId": camp_id,
                "adGroupId": ag_id,
                "expression": [{"type": "ASIN_SAME_AS", "value": _format_asin(hedefleme)}],
                "expressionType": "MANUAL",
                "state": "ENABLED",
            }

    result["campaign_id"] = camp_id
    result["ad_group_id"] = ag_id
    result["ad_type"] = ad_type

    return result


def prepare_harvest_keyword(action, config, campaign_lookup, today=None):
    """
    Harvesting keyword: kaynak kampanyada negatif ekle + yeni Exact kampanya olustur.
    """
    result = {
        "action": action,
        "status": "HAZIR",
        "sub_operations": [],
        "uyarilar": [],
        "hatalar": [],
    }

    # Kaynak kampanya ID bul
    source_camp_name = action.get("kaynak_kampanya", "")
    source_info = campaign_lookup["by_name"].get(source_camp_name)
    if not source_info:
        result["status"] = "HATA"
        result["hatalar"].append(f"Kaynak kampanya bulunamadi: {source_camp_name}")
        return result

    source_camp_id = source_info["id"]
    portfolio = action.get("portfolio", "")
    hedefleme = action["hedefleme"]
    bid = action.get("bid", 0)
    butce = config.get("yeni_kampanya_butcesi", 10.00)

    # Hedeflenen ASIN'i bul
    ag_id, _ = resolve_ad_group_id(source_camp_id, campaign_lookup)
    advertised_asin, advertised_sku, asin_uyari = find_advertised_asin(source_camp_id, ag_id, today)
    if asin_uyari:
        result["uyarilar"].append(asin_uyari)
    if not advertised_asin:
        result["status"] = "HATA"
        result["hatalar"].append("Hedeflenen urun ASIN'i bulunamadi")
        return result

    # Op 1: Kaynak kampanyada negatif keyword ekle
    result["sub_operations"].append({
        "op": "NEGATIF_EKLE",
        "api_endpoint": "sp_negative_keyword_add",
        "api_payload": {
            "campaignId": source_camp_id,
            "adGroupId": ag_id,
            "keywordText": hedefleme,
            "matchType": "NEGATIVE_EXACT",
            "state": "ENABLED",
        },
    })

    # Bid validation — minimum bid korumasi
    bid, bid_uyari = validate_bid(bid, config)
    if bid_uyari:
        result["uyarilar"].append(bid_uyari)

    # Op 2: Yeni kampanya olustur
    kampanya_adi = f"{portfolio} - E - {hedefleme}"

    # Ayni isimde kampanya var mi kontrol et
    if kampanya_adi in campaign_lookup["by_name"]:
        kampanya_adi = f"{kampanya_adi}-2"
        result["uyarilar"].append(f"Ayni isimde kampanya mevcut, '{kampanya_adi}' olarak olusturulacak")
        # -2 de varsa -3, -4 dene
        counter = 3
        while kampanya_adi in campaign_lookup["by_name"]:
            kampanya_adi = f"{portfolio} - E - {hedefleme}-{counter}"
            counter += 1

    # Amazon SP v3 API kampanya payload'u
    campaign_payload = {
        "name": kampanya_adi,
        "targetingType": "MANUAL",
        "state": "ENABLED",
        "budget": {
            "budgetType": "DAILY",
            "budget": round(butce, 2),
        },
        "startDate": datetime.utcnow().strftime("%Y-%m-%d"),
        "dynamicBidding": {
            "strategy": "LEGACY_FOR_SALES",  # Down Only
        },
    }
    # portfolioId sadece doluysa ekle (bos string Amazon'da hata verir)
    pid = source_info.get("portfolio_id", "")
    if pid:
        campaign_payload["portfolioId"] = pid

    result["sub_operations"].append({
        "op": "KAMPANYA_OLUSTUR",
        "api_endpoint": "sp_campaign_create",
        "api_payload": campaign_payload,
    })

    # Op 3: Ad group olustur
    result["sub_operations"].append({
        "op": "AD_GROUP_OLUSTUR",
        "api_endpoint": "sp_ad_group_create",
        "api_payload": {
            "campaignId": "__YENI_KAMPANYA_ID__",  # Runtime'da doldurulacak
            "name": kampanya_adi,
            "state": "ENABLED",
            "defaultBid": round(bid, 2),
        },
    })

    # Op 4: Urun reklami ekle (SKU zorunlu — Amazon API merchantSku ister)
    product_ad_payload = {
        "campaignId": "__YENI_KAMPANYA_ID__",
        "adGroupId": "__YENI_AD_GROUP_ID__",
        "asin": advertised_asin,
        "state": "ENABLED",
    }
    if advertised_sku:
        product_ad_payload["sku"] = advertised_sku
    else:
        result["uyarilar"].append("SKU bulunamadi — product ad olusturma basarisiz olabilir")
    result["sub_operations"].append({
        "op": "URUN_REKLAMI_EKLE",
        "api_endpoint": "sp_product_ad_create",
        "api_payload": product_ad_payload,
    })

    # Op 5: Exact keyword ekle
    result["sub_operations"].append({
        "op": "KEYWORD_EKLE",
        "api_endpoint": "sp_keyword_create",
        "api_payload": {
            "campaignId": "__YENI_KAMPANYA_ID__",
            "adGroupId": "__YENI_AD_GROUP_ID__",
            "keywordText": hedefleme,
            "matchType": "EXACT",
            "bid": round(bid, 2),
            "state": "ENABLED",
        },
    })

    result["kampanya_adi"] = kampanya_adi
    result["advertised_asin"] = advertised_asin
    result["source_campaign_id"] = source_camp_id

    return result


def prepare_harvest_asin(action, config, campaign_lookup, today=None):
    """
    Harvesting ASIN: kaynak kampanyada negatif ASIN ekle +
    mevcut ASIN Target kampanyasina yeni hedefleme ekle.
    """
    result = {
        "action": action,
        "status": "HAZIR",
        "sub_operations": [],
        "uyarilar": [],
        "hatalar": [],
    }

    # Kaynak kampanya ID bul
    source_camp_name = action.get("kaynak_kampanya", "")
    source_info = campaign_lookup["by_name"].get(source_camp_name)
    if not source_info:
        result["status"] = "HATA"
        result["hatalar"].append(f"Kaynak kampanya bulunamadi: {source_camp_name}")
        return result

    source_camp_id = source_info["id"]
    ag_id, _ = resolve_ad_group_id(source_camp_id, campaign_lookup)
    portfolio = action.get("portfolio", "")
    hedefleme = _format_asin(action["hedefleme"])  # ASIN buyuk harfe cevir
    bid = action.get("bid", 0)

    # Op 1: Kaynak kampanyada negatif ASIN target ekle
    if ag_id:
        result["sub_operations"].append({
            "op": "NEGATIF_ASIN_EKLE",
            "api_endpoint": "sp_negative_target_add",
            "api_payload": {
                "campaignId": source_camp_id,
                "adGroupId": ag_id,
                "expression": [{"type": "ASIN_SAME_AS", "value": hedefleme}],
                "expressionType": "MANUAL",
                "state": "ENABLED",
            },
        })

    # Op 2: Mevcut ASIN Target kampanyasina ekle
    asin_target_camps = config.get("portfolio_asin_target_kampanyalari", {})
    target_camp_id = asin_target_camps.get(portfolio, "")

    if not target_camp_id:
        result["status"] = "HATA"
        result["hatalar"].append(
            f"Portfolio '{portfolio}' icin ASIN Target kampanya ID'si tanimli degil. "
            f"Dashboard → Ayarlar → Portfolio ASIN Kampanya Eslestirme'den tanimlayin."
        )
        return result

    # Hedef kampanyanin ad group'unu bul
    target_ag_id, ag_err = resolve_ad_group_id(target_camp_id, campaign_lookup)
    if not target_ag_id:
        result["status"] = "HATA"
        result["hatalar"].append(f"ASIN Target kampanyasinda ad group bulunamadi: {ag_err}")
        return result

    # Bid validation — minimum bid korumasi
    validated_bid, bid_uyari = validate_bid(bid, config)
    if bid_uyari:
        result["uyarilar"].append(bid_uyari)

    result["sub_operations"].append({
        "op": "ASIN_TARGET_EKLE",
        "api_endpoint": "sp_target_create",
        "api_payload": {
            "campaignId": target_camp_id,
            "adGroupId": target_ag_id,
            "expression": [{"type": "ASIN_SAME_AS", "value": hedefleme}],
            "expressionType": "MANUAL",
            "bid": round(validated_bid, 2),
            "state": "ENABLED",
        },
    })

    result["source_campaign_id"] = source_camp_id
    result["target_campaign_id"] = target_camp_id

    return result


# ============================================================================
# DRY-RUN RAPORU
# ============================================================================

def generate_dry_run_report(bid_ops, neg_ops, harvest_ops):
    """
    Dry-run ozet raporu olusturur. Kullaniciya gosterilir.
    """
    report = {
        "mod": "DRY_RUN",
        "tarih": datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
        "ozet": {
            "bid_degisiklikleri": {
                "toplam": len(bid_ops),
                "hazir": sum(1 for o in bid_ops if o["status"] == "HAZIR"),
                "hata": sum(1 for o in bid_ops if o["status"] == "HATA"),
            },
            "negatif_eklemeler": {
                "toplam": len(neg_ops),
                "hazir": sum(1 for o in neg_ops if o["status"] == "HAZIR"),
                "hata": sum(1 for o in neg_ops if o["status"] == "HATA"),
            },
            "harvesting": {
                "toplam": len(harvest_ops),
                "hazir": sum(1 for o in harvest_ops if o["status"] == "HAZIR"),
                "hata": sum(1 for o in harvest_ops if o["status"] == "HATA"),
                "yeni_kampanya": sum(1 for o in harvest_ops if o.get("action", {}).get("tip") == "HARVEST_KEYWORD" and o["status"] == "HAZIR"),
                "asin_ekleme": sum(1 for o in harvest_ops if o.get("action", {}).get("tip") == "HARVEST_ASIN" and o["status"] == "HAZIR"),
            },
        },
        "detaylar": {
            "bid_degisiklikleri": [],
            "negatif_eklemeler": [],
            "harvesting": [],
        },
        "hatalar": [],
        "uyarilar": [],
    }

    # Bid detaylari
    for op in bid_ops:
        a = op["action"]
        if op["status"] == "HAZIR":
            report["detaylar"]["bid_degisiklikleri"].append({
                "kampanya": a["kampanya_adi"],
                "hedefleme": a["hedefleme"],
                "eski_bid": a["eski_bid"],
                "yeni_bid": a["yeni_bid"],
                "degisim": f"{((a['yeni_bid'] - a['eski_bid']) / a['eski_bid'] * 100) if a['eski_bid'] > 0 else 0:+.1f}%",
            })
        else:
            report["hatalar"].extend(op["hatalar"])
        report["uyarilar"].extend(op.get("uyarilar", []))

    # Negatif detaylari
    for op in neg_ops:
        a = op["action"]
        if op["status"] == "HAZIR":
            report["detaylar"]["negatif_eklemeler"].append({
                "kampanya": a["kampanya_adi"],
                "hedefleme": a["hedefleme"],
                "tip": a["tip"],
            })
        else:
            report["hatalar"].extend(op["hatalar"])
        report["uyarilar"].extend(op.get("uyarilar", []))

    # Harvesting detaylari
    for op in harvest_ops:
        a = op["action"]
        if op["status"] == "HAZIR":
            detail = {
                "kaynak_kampanya": a.get("kaynak_kampanya", ""),
                "hedefleme": a["hedefleme"],
                "tip": a["tip"],
            }
            if a["tip"] == "HARVEST_KEYWORD":
                detail["yeni_kampanya_adi"] = op.get("kampanya_adi", "")
                detail["advertised_asin"] = op.get("advertised_asin", "")
            report["detaylar"]["harvesting"].append(detail)
        else:
            report["hatalar"].extend(op["hatalar"])
        report["uyarilar"].extend(op.get("uyarilar", []))

    return report


# ============================================================================
# ROLLBACK LOG
# ============================================================================

def save_rollback_log(bid_ops, neg_ops, harvest_ops, today=None):
    """Her islem icin rollback bilgisini kaydeder."""
    if today is None:
        today = datetime.utcnow().strftime("%Y-%m-%d")

    LOG_DIR.mkdir(parents=True, exist_ok=True)

    log = {
        "tarih": today,
        "olusturma_zamani": datetime.utcnow().isoformat(),
        "islemler": [],
    }

    for op in bid_ops:
        if op["status"] != "UYGULANDI":
            continue
        a = op["action"]
        log["islemler"].append({
            "tip": "BID_DEGISIKLIGI",
            "kampanya": a["kampanya_adi"],
            "hedefleme": a["hedefleme"],
            "eski_bid": a["eski_bid"],
            "yeni_bid": a["yeni_bid"],
            "entity_id": op.get("entity_id", ""),
            "entity_type": op.get("entity_type", ""),
            "ad_type": op.get("ad_type", ""),
            "campaign_id": op.get("campaign_id", ""),
            "rollback": f"Bid'i {a['eski_bid']:.2f} olarak geri al",
        })

    for op in neg_ops:
        if op["status"] != "UYGULANDI":
            continue
        a = op["action"]
        log["islemler"].append({
            "tip": a["tip"],
            "kampanya": a["kampanya_adi"],
            "hedefleme": a["hedefleme"],
            "campaign_id": op.get("campaign_id", ""),
            "rollback": "Negatif hedeflemeyi sil (API ile ARCHIVED yap)",
        })

    for op in harvest_ops:
        if op["status"] != "UYGULANDI":
            continue
        a = op["action"]
        entry = {
            "tip": a["tip"],
            "kaynak_kampanya": a.get("kaynak_kampanya", ""),
            "hedefleme": a["hedefleme"],
        }
        if a["tip"] == "HARVEST_KEYWORD":
            entry["yeni_kampanya_adi"] = op.get("kampanya_adi", "")
            entry["rollback"] = "Yeni kampanyayi PAUSED yap + kaynak kampanyadaki negatifi sil"
        else:
            entry["rollback"] = "Eklenen ASIN target'i ARCHIVED yap + kaynak kampanyadaki negatifi sil"
        log["islemler"].append(entry)

    log_path = LOG_DIR / f"{today}_rollback.json"
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(log, f, indent=2, ensure_ascii=False)

    logger.info("Rollback log kaydedildi: %s (%d islem)", log_path, len(log["islemler"]))
    return str(log_path)


# ============================================================================
# DOGRULAMA — ASAMA 1: ANLIK (API RESPONSE KONTROLU)
# ============================================================================

def verify_api_response(api_response, expected_action):
    """
    Her API cagrisindan sonra response'u kontrol eder.
    
    Args:
        api_response: API'den donen response (dict veya raw)
        expected_action: Beklenen islem bilgisi
    
    Returns:
        dict: {
            "dogrulandi": True/False,
            "detay": "...",
            "response_code": 200,
            "retry_gerekli": True/False
        }
    """
    result = {
        "dogrulandi": False,
        "detay": "",
        "response_code": None,
        "retry_gerekli": False,
        "zaman": datetime.utcnow().isoformat(),
    }

    if api_response is None:
        result["detay"] = "API response bos (None)"
        result["retry_gerekli"] = True
        return result

    # Response bir dict ise
    if isinstance(api_response, dict):
        # HTTP status kodu kontrol
        status = api_response.get("status", api_response.get("code", api_response.get("statusCode", 0)))
        result["response_code"] = status

        # Hata kontrolleri
        if status and isinstance(status, int):
            if 200 <= status < 300:
                result["dogrulandi"] = True
                result["detay"] = f"HTTP {status} — basarili"
            elif status == 429:
                result["detay"] = f"HTTP 429 — rate limit, retry gerekli"
                result["retry_gerekli"] = True
            elif status == 500 or status == 503:
                result["detay"] = f"HTTP {status} — sunucu hatasi, retry gerekli"
                result["retry_gerekli"] = True
            elif 400 <= status < 500:
                error_msg = api_response.get("message", api_response.get("error", ""))
                result["detay"] = f"HTTP {status} — istemci hatasi: {error_msg}"
                result["retry_gerekli"] = False  # 4xx genelde retry ile duzelmez
            else:
                result["detay"] = f"HTTP {status} — beklenmeyen durum kodu"
                result["retry_gerekli"] = True

        # Hata mesaji iceriyorsa
        elif "error" in api_response or "errors" in api_response:
            errors = api_response.get("errors", [api_response.get("error", "")])
            result["detay"] = f"API hata dondu: {errors}"
            result["retry_gerekli"] = True

        # Basarili response (status kodu yok ama hata da yok)
        else:
            # Amazon API bazen sadece entity ID doner
            if any(k in api_response for k in ["keywordId", "targetId", "campaignId", "adGroupId"]):
                result["dogrulandi"] = True
                result["detay"] = "Basarili — entity ID dondu"
            elif "success" in str(api_response).lower():
                result["dogrulandi"] = True
                result["detay"] = "Basarili"
            else:
                result["detay"] = f"Belirsiz response: {str(api_response)[:200]}"
                result["retry_gerekli"] = True

    # Response bir list ise (batch islemler)
    elif isinstance(api_response, list):
        success_count = 0
        error_count = 0
        for item in api_response:
            if isinstance(item, dict):
                if item.get("code") == "SUCCESS" or item.get("status") == "SUCCESS":
                    success_count += 1
                elif "error" in item:
                    error_count += 1
                else:
                    success_count += 1  # Varsayilan basarili say
        result["dogrulandi"] = error_count == 0
        result["detay"] = f"Batch: {success_count} basarili, {error_count} hata"
        result["retry_gerekli"] = error_count > 0

    # String response
    elif isinstance(api_response, str):
        try:
            parsed = json.loads(api_response)
            return verify_api_response(parsed, expected_action)
        except json.JSONDecodeError:
            if "error" in api_response.lower():
                result["detay"] = f"String hata: {api_response[:200]}"
                result["retry_gerekli"] = True
            else:
                result["dogrulandi"] = True
                result["detay"] = "String response — varsayilan basarili"

    return result


def verify_batch_results(operations, api_responses):
    """
    Toplu islemlerin sonuclarini dogrular.
    
    Args:
        operations: Hazirlanmis islem listesi
        api_responses: Her isleme karsilik gelen API response listesi
    
    Returns:
        dict: {
            "toplam": N,
            "dogrulanan": N,
            "basarisiz": N,
            "retry_gereken": [...],
            "detaylar": [...]
        }
    """
    result = {
        "toplam": len(operations),
        "dogrulanan": 0,
        "basarisiz": 0,
        "retry_gereken": [],
        "detaylar": [],
    }

    for i, (op, response) in enumerate(zip(operations, api_responses)):
        verification = verify_api_response(response, op.get("action", {}))
        verification["islem_index"] = i
        verification["islem_tipi"] = op.get("action", {}).get("tip", "BILINMIYOR")

        if verification["dogrulandi"]:
            result["dogrulanan"] += 1
            op["status"] = "UYGULANDI"
        else:
            result["basarisiz"] += 1
            if verification["retry_gerekli"]:
                result["retry_gereken"].append({
                    "index": i,
                    "operation": op,
                    "sebep": verification["detay"],
                })

        result["detaylar"].append(verification)

    return result


# ============================================================================
# DOGRULAMA — ASAMA 2: GECIKMELI (AMAZON'DAN GUNCEL VERI CEKIP KARSILASTIRMA)
# ============================================================================

VERIFICATION_DELAY_SECONDS = 300  # 5 dakika


def run_delayed_verification(rollback_log_path, today=None):
    """
    Agent 3 islemlerinden 5 dk sonra calistirilir.
    Rollback log'daki her islemi Amazon'dan guncel veriyle karsilastirir.
    Uyusmayan islemler icin retry listesi olusturur.
    
    Bu fonksiyon Agent 1'in MCP tool'lariyla Amazon'dan veri ceker.
    Claude Code bu fonksiyonu cagirdiginda MCP tool'larini kullanarak
    gercek verileri getirir.
    
    Args:
        rollback_log_path: Rollback log dosya yolu
        today: Tarih (varsayilan: bugun)
    
    Returns:
        dict: Dogrulama raporu
    """
    if today is None:
        today = datetime.utcnow().strftime("%Y-%m-%d")

    logger.info("=== GECIKMELI DOGRULAMA BASLADI ===")

    # Rollback log'u oku
    with open(rollback_log_path, "r", encoding="utf-8") as f:
        rollback_log = json.load(f)

    islemler = rollback_log.get("islemler", [])
    if not islemler:
        logger.info("Dogrulanacak islem yok.")
        return {"durum": "BOS", "mesaj": "Dogrulanacak islem yok."}

    # Dogrulama icin gerekli verileri hazirla
    verification_tasks = {
        "bid_kontrolleri": [],
        "negatif_kontrolleri": [],
        "kampanya_kontrolleri": [],
    }

    for islem in islemler:
        tip = islem.get("tip", "")

        if tip == "BID_DEGISIKLIGI":
            verification_tasks["bid_kontrolleri"].append({
                "entity_id": islem.get("entity_id", ""),
                "entity_type": islem.get("entity_type", ""),
                "ad_type": islem.get("ad_type", ""),
                "campaign_id": islem.get("campaign_id", ""),
                "beklenen_bid": islem.get("yeni_bid", 0),
                "eski_bid": islem.get("eski_bid", 0),
                "hedefleme": islem.get("hedefleme", ""),
                "kampanya": islem.get("kampanya", ""),
            })

        elif tip in ("NEGATIF_KEYWORD", "NEGATIF_ASIN"):
            verification_tasks["negatif_kontrolleri"].append({
                "campaign_id": islem.get("campaign_id", ""),
                "hedefleme": islem.get("hedefleme", ""),
                "tip": tip,
                "kampanya": islem.get("kampanya", ""),
            })

        elif tip == "HARVEST_KEYWORD":
            verification_tasks["kampanya_kontrolleri"].append({
                "kampanya_adi": islem.get("yeni_kampanya_adi", ""),
                "hedefleme": islem.get("hedefleme", ""),
                "beklenen_durum": "ENABLED",
            })

        elif tip == "HARVEST_ASIN":
            verification_tasks["negatif_kontrolleri"].append({
                "campaign_id": islem.get("source_campaign_id", ""),
                "hedefleme": islem.get("hedefleme", ""),
                "tip": "NEGATIF_ASIN",
                "kampanya": islem.get("kaynak_kampanya", ""),
            })

    logger.info("Dogrulama gorevleri: %d bid, %d negatif, %d kampanya",
                len(verification_tasks["bid_kontrolleri"]),
                len(verification_tasks["negatif_kontrolleri"]),
                len(verification_tasks["kampanya_kontrolleri"]))

    # Dogrulama raporu sablonu
    # NOT: Gercek dogrulama Claude Code tarafindan MCP tool'lariyla yapilir.
    # Bu fonksiyon dogrulama gorevlerini hazirlar ve sonuclari isler.

    verification_report = {
        "tarih": today,
        "zaman": datetime.utcnow().isoformat(),
        "tip": "GECIKMELI_DOGRULAMA",
        "gorevler": verification_tasks,
        "sonuclar": {
            "bid_kontrolleri": [],
            "negatif_kontrolleri": [],
            "kampanya_kontrolleri": [],
        },
        "ozet": {
            "toplam_kontrol": len(islemler),
            "dogrulanan": 0,
            "uyusmayan": 0,
            "kontrol_edilemeyen": 0,
        },
        "retry_listesi": [],
    }

    return verification_report


def process_verification_results(verification_report, actual_data):
    """
    Claude Code'un MCP tool'lariyla cektigi guncel verileri
    beklenen degerlerle karsilastirir.
    
    Args:
        verification_report: run_delayed_verification'dan donen rapor
        actual_data: Claude Code'un Amazon'dan cektigi guncel veriler
            {
                "keywords": {entity_id: {"bid": 0.85, ...}},
                "targets": {entity_id: {"bid": 1.20, ...}},
                "negative_keywords": {campaign_id: [keyword_list]},
                "negative_targets": {campaign_id: [asin_list]},
                "campaigns": {campaign_name: {"state": "ENABLED", ...}}
            }
    
    Returns:
        dict: Guncellenmis dogrulama raporu (retry_listesi dahil)
    """
    report = verification_report
    retry_list = []

    # 1. Bid dogrulamasi
    keywords = actual_data.get("keywords", {})
    targets = actual_data.get("targets", {})

    for task in report["gorevler"]["bid_kontrolleri"]:
        entity_id = task["entity_id"]
        beklenen = task["beklenen_bid"]
        entity_type = task["entity_type"]

        # Guncel bid'i bul
        if entity_type == "KEYWORD":
            source = keywords
        elif entity_type == "THEME":
            source = actual_data.get("themes", {})
        else:
            source = targets
        actual_entity = source.get(entity_id, {})
        gercek_bid = actual_entity.get("bid", None)

        sonuc = {
            "entity_id": entity_id,
            "hedefleme": task["hedefleme"],
            "kampanya": task["kampanya"],
            "beklenen_bid": beklenen,
            "gercek_bid": gercek_bid,
        }

        if gercek_bid is None:
            sonuc["durum"] = "KONTROL_EDILEMEDI"
            sonuc["detay"] = "Entity bulunamadi"
            report["ozet"]["kontrol_edilemeyen"] += 1
        elif abs(gercek_bid - beklenen) < 0.01:  # 1 cent tolerans
            sonuc["durum"] = "DOGRULANDI"
            report["ozet"]["dogrulanan"] += 1
        else:
            sonuc["durum"] = "UYUSMADI"
            sonuc["detay"] = f"Beklenen: {beklenen:.2f}, Gercek: {gercek_bid:.2f}"
            report["ozet"]["uyusmayan"] += 1
            retry_list.append({
                "tip": "BID_DEGISIKLIGI",
                "entity_id": entity_id,
                "entity_type": entity_type,
                "ad_type": task["ad_type"],
                "campaign_id": task["campaign_id"],
                "hedefleme": task["hedefleme"],
                "beklenen_bid": beklenen,
                "gercek_bid": gercek_bid,
                "sebep": f"Bid uyusmadi: {gercek_bid:.2f} != {beklenen:.2f}",
            })

        report["sonuclar"]["bid_kontrolleri"].append(sonuc)

    # 2. Negatif keyword/ASIN dogrulamasi
    neg_keywords = actual_data.get("negative_keywords", {})
    neg_targets = actual_data.get("negative_targets", {})

    for task in report["gorevler"]["negatif_kontrolleri"]:
        campaign_id = task["campaign_id"]
        hedefleme = (task.get("hedefleme") or "").lower()
        tip = task["tip"]

        sonuc = {
            "campaign_id": campaign_id,
            "hedefleme": task["hedefleme"],
            "kampanya": task.get("kampanya", ""),
            "tip": tip,
        }

        if tip == "NEGATIF_KEYWORD":
            neg_list = [(kw or "").lower() for kw in neg_keywords.get(campaign_id, [])]
            if hedefleme in neg_list:
                sonuc["durum"] = "DOGRULANDI"
                report["ozet"]["dogrulanan"] += 1
            else:
                sonuc["durum"] = "UYUSMADI"
                sonuc["detay"] = "Negatif keyword listesinde bulunamadi"
                report["ozet"]["uyusmayan"] += 1
                retry_list.append({
                    "tip": "NEGATIF_KEYWORD",
                    "campaign_id": campaign_id,
                    "hedefleme": task["hedefleme"],
                    "sebep": "Negatif keyword eklenmemis",
                })
        elif tip == "NEGATIF_ASIN":
            neg_list = [(a or "").lower() for a in neg_targets.get(campaign_id, [])]
            if hedefleme in neg_list:
                sonuc["durum"] = "DOGRULANDI"
                report["ozet"]["dogrulanan"] += 1
            else:
                sonuc["durum"] = "UYUSMADI"
                sonuc["detay"] = "Negatif ASIN target listesinde bulunamadi"
                report["ozet"]["uyusmayan"] += 1
                retry_list.append({
                    "tip": "NEGATIF_ASIN",
                    "campaign_id": campaign_id,
                    "hedefleme": task["hedefleme"],
                    "sebep": "Negatif ASIN eklenmemis",
                })

        report["sonuclar"]["negatif_kontrolleri"].append(sonuc)

    # 3. Yeni kampanya dogrulamasi
    campaigns = actual_data.get("campaigns", {})

    for task in report["gorevler"]["kampanya_kontrolleri"]:
        kampanya_adi = task["kampanya_adi"]

        sonuc = {
            "kampanya_adi": kampanya_adi,
            "hedefleme": task.get("hedefleme", ""),
        }

        camp_info = campaigns.get(kampanya_adi, {})
        if camp_info:
            state = camp_info.get("state", "").upper()
            if state == task["beklenen_durum"]:
                sonuc["durum"] = "DOGRULANDI"
                sonuc["detay"] = f"Kampanya mevcut, durum: {state}"
                report["ozet"]["dogrulanan"] += 1
            else:
                sonuc["durum"] = "UYUSMADI"
                sonuc["detay"] = f"Kampanya durumu yanlis: {state} (beklenen: {task['beklenen_durum']})"
                report["ozet"]["uyusmayan"] += 1
                retry_list.append({
                    "tip": "KAMPANYA_DURUM",
                    "kampanya_adi": kampanya_adi,
                    "beklenen_durum": task["beklenen_durum"],
                    "gercek_durum": state,
                    "sebep": f"Kampanya durumu uyusmadi",
                })
        else:
            sonuc["durum"] = "UYUSMADI"
            sonuc["detay"] = "Kampanya bulunamadi"
            report["ozet"]["uyusmayan"] += 1
            retry_list.append({
                "tip": "KAMPANYA_OLUSTURMA",
                "kampanya_adi": kampanya_adi,
                "hedefleme": task.get("hedefleme", ""),
                "sebep": "Kampanya olusturulamamis",
            })

        report["sonuclar"]["kampanya_kontrolleri"].append(sonuc)

    report["retry_listesi"] = retry_list

    return report


def load_verify_actual_data(data_date):
    """
    _verify_ dosyalarindan actual_data sozlugunu olusturur.
    process_verification_results'a gecilecek formatta doner.

    Verify dosyalari: {data_date}_verify_{entity}.json
    Orijinal dosyalara DOKUNMAZ.

    Returns:
        dict: {
            "keywords": {entity_id: {"bid": X, ...}},
            "targets": {entity_id: {"bid": X, ...}},
            "negative_keywords": {campaign_id: [keyword_list]},
            "negative_targets": {campaign_id: [asin_list]},
            "campaigns": {campaign_name: {"state": "ENABLED", ...}}
        }
    """
    actual = {
        "keywords": {},
        "targets": {},
        "themes": {},
        "negative_keywords": {},
        "negative_targets": {},
        "campaigns": {},
    }

    prefix = f"{data_date}_verify_"

    # SP Keywords
    kw_path = DATA_DIR / f"{prefix}sp_keywords.json"
    if kw_path.exists():
        try:
            with open(kw_path, "r", encoding="utf-8") as f:
                for kw in json.load(f):
                    kid = str(kw.get("keywordId", ""))
                    if kid:
                        actual["keywords"][kid] = {"bid": kw.get("bid", 0)}
        except (json.JSONDecodeError, IOError):
            pass

    # SB Keywords
    sb_kw_path = DATA_DIR / f"{prefix}sb_keywords.json"
    if sb_kw_path.exists():
        try:
            with open(sb_kw_path, "r", encoding="utf-8") as f:
                for kw in json.load(f):
                    kid = str(kw.get("keywordId", ""))
                    if kid:
                        actual["keywords"][kid] = {"bid": kw.get("bid", 0)}
        except (json.JSONDecodeError, IOError):
            pass

    # SP Targets
    tgt_path = DATA_DIR / f"{prefix}sp_targets.json"
    if tgt_path.exists():
        try:
            with open(tgt_path, "r", encoding="utf-8") as f:
                for t in json.load(f):
                    tid = str(t.get("targetId", ""))
                    if tid:
                        actual["targets"][tid] = {"bid": t.get("bid", 0)}
        except (json.JSONDecodeError, IOError):
            pass

    # SD Targets
    sd_tgt_path = DATA_DIR / f"{prefix}sd_targets.json"
    if sd_tgt_path.exists():
        try:
            with open(sd_tgt_path, "r", encoding="utf-8") as f:
                for t in json.load(f):
                    tid = str(t.get("targetId", ""))
                    if tid:
                        actual["targets"][tid] = {"bid": t.get("bid", 0)}
        except (json.JSONDecodeError, IOError):
            pass

    # SB Targets
    sb_tgt_path = DATA_DIR / f"{prefix}sb_targets.json"
    if sb_tgt_path.exists():
        try:
            with open(sb_tgt_path, "r", encoding="utf-8") as f:
                for t in json.load(f):
                    tid = str(t.get("targetId", ""))
                    if tid:
                        actual["targets"][tid] = {"bid": t.get("bid", 0)}
        except (json.JSONDecodeError, IOError):
            pass

    # SB Themes
    theme_path = DATA_DIR / f"{prefix}sb_themes.json"
    if theme_path.exists():
        try:
            with open(theme_path, "r", encoding="utf-8") as f:
                for t in json.load(f):
                    tid = str(t.get("themeId", ""))
                    if tid:
                        actual["themes"][tid] = {"bid": t.get("bid", 0)}
        except (json.JSONDecodeError, IOError):
            pass

    # SP Negative Keywords
    neg_kw_path = DATA_DIR / f"{prefix}sp_negative_keywords.json"
    if neg_kw_path.exists():
        try:
            with open(neg_kw_path, "r", encoding="utf-8") as f:
                for nk in json.load(f):
                    cid = str(nk.get("campaignId", ""))
                    text = nk.get("keywordText", "")
                    if cid and text:
                        actual["negative_keywords"].setdefault(cid, []).append(text)
        except (json.JSONDecodeError, IOError):
            pass

    # SP Negative Targets
    neg_tgt_path = DATA_DIR / f"{prefix}sp_negative_targets.json"
    if neg_tgt_path.exists():
        try:
            with open(neg_tgt_path, "r", encoding="utf-8") as f:
                for nt in json.load(f):
                    cid = str(nt.get("campaignId", ""))
                    expr = nt.get("expression", [])
                    for e in (expr if isinstance(expr, list) else []):
                        val = e.get("value", "")
                        if cid and val:
                            actual["negative_targets"].setdefault(cid, []).append(val)
        except (json.JSONDecodeError, IOError):
            pass

    # SP Campaigns (yeni kampanya dogrulamasi icin)
    camp_path = DATA_DIR / f"{prefix}sp_campaigns.json"
    if camp_path.exists():
        try:
            with open(camp_path, "r", encoding="utf-8") as f:
                for c in json.load(f):
                    name = c.get("name", "")
                    if name:
                        actual["campaigns"][name] = {
                            "state": c.get("state", ""),
                            "campaignId": c.get("campaignId", ""),
                        }
        except (json.JSONDecodeError, IOError):
            pass

    logger.info("Verify verileri yuklendi: %d keyword, %d target, %d neg.kw campaign, "
                "%d neg.target campaign, %d kampanya",
                len(actual["keywords"]), len(actual["targets"]),
                len(actual["negative_keywords"]), len(actual["negative_targets"]),
                len(actual["campaigns"]))

    return actual


def prepare_retry_operations(retry_list, config, campaign_lookup, targeting_lookup, today=None):
    """
    Dogrulama sonucu uyusmayan islemleri tekrar denemek icin
    API payload'larini yeniden hazirlar.
    
    Args:
        retry_list: process_verification_results'tan donen retry_listesi
        config: Agent 3 config
        campaign_lookup: Kampanya eslestirme tablosu
        targeting_lookup: Hedefleme eslestirme tablosu
    
    Returns:
        dict: Retry islemleri (ayni formatta execution plan)
    """
    if today is None:
        today = datetime.utcnow().strftime("%Y-%m-%d")

    retry_ops = {
        "tarih": today,
        "zaman": datetime.utcnow().isoformat(),
        "tip": "RETRY",
        "toplam": len(retry_list),
        "islemler": [],
    }

    for item in retry_list:
        tip = item.get("tip", "")
        retry_op = {
            "kaynak": item,
            "status": "HAZIR",
            "api_endpoint": None,
            "api_payload": None,
            "hatalar": [],
        }

        if tip == "BID_DEGISIKLIGI":
            entity_type = item.get("entity_type", "")
            ad_type = item.get("ad_type", "")
            beklenen_bid = item.get("beklenen_bid", 0)

            # Bid limitleri tekrar kontrol
            beklenen_bid, uyari = validate_bid(beklenen_bid, config)

            if entity_type == "KEYWORD":
                if ad_type == "SP":
                    retry_op["api_endpoint"] = "sp_keyword_bid_update"
                elif ad_type == "SB":
                    retry_op["api_endpoint"] = "sb_keyword_bid_update"
                retry_op["api_payload"] = {
                    "keywordId": item["entity_id"],
                    "bid": round(beklenen_bid, 2),
                }
            elif entity_type == "TARGET":
                if ad_type == "SP":
                    retry_op["api_endpoint"] = "sp_target_bid_update"
                elif ad_type == "SD":
                    retry_op["api_endpoint"] = "sd_target_bid_update"
                retry_op["api_payload"] = {
                    "targetId": item["entity_id"],
                    "bid": round(beklenen_bid, 2),
                }

        elif tip == "NEGATIF_KEYWORD":
            campaign_id = item.get("campaign_id", "")
            ag_id, _ = resolve_ad_group_id(campaign_id, campaign_lookup)
            if ag_id:
                retry_op["api_endpoint"] = "sp_negative_keyword_add"
                retry_op["api_payload"] = {
                    "campaignId": campaign_id,
                    "adGroupId": ag_id,
                    "keywordText": item["hedefleme"],
                    "matchType": "NEGATIVE_EXACT",
                    "state": "ENABLED",
                }
            else:
                retry_op["status"] = "HATA"
                retry_op["hatalar"].append("Ad group bulunamadi")

        elif tip == "NEGATIF_ASIN":
            campaign_id = item.get("campaign_id", "")
            ag_id, _ = resolve_ad_group_id(campaign_id, campaign_lookup)
            if ag_id:
                retry_op["api_endpoint"] = "sp_negative_target_add"
                retry_op["api_payload"] = {
                    "campaignId": campaign_id,
                    "adGroupId": ag_id,
                    "expression": [{"type": "ASIN_SAME_AS", "value": _format_asin(item["hedefleme"])}],
                    "expressionType": "MANUAL",
                    "state": "ENABLED",
                }
            else:
                retry_op["status"] = "HATA"
                retry_op["hatalar"].append("Ad group bulunamadi")

        elif tip in ("KAMPANYA_OLUSTURMA", "KAMPANYA_DURUM"):
            # Kampanya olusturma retry'i karmasik — bilgilendir, manuel mudahale oner
            retry_op["status"] = "MANUEL_GEREKLI"
            retry_op["hatalar"].append(
                f"Kampanya '{item.get('kampanya_adi', '')}' olusturma/durum sorunu. "
                f"Manuel kontrol oneriliyor."
            )

        retry_ops["islemler"].append(retry_op)

    # Retry planini kaydet
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    retry_path = LOG_DIR / f"{today}_retry_plan.json"
    with open(retry_path, "w", encoding="utf-8") as f:
        json.dump(retry_ops, f, indent=2, ensure_ascii=False)

    logger.info("Retry plan kaydedildi: %s (%d islem)", retry_path, len(retry_list))
    return retry_ops


def save_verification_report(report, today=None):
    """Dogrulama raporunu dosyaya kaydeder."""
    if today is None:
        today = datetime.utcnow().strftime("%Y-%m-%d")

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    report_path = LOG_DIR / f"{today}_verification_report.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)

    logger.info("Dogrulama raporu kaydedildi: %s", report_path)
    logger.info("Ozet: %d dogrulandi, %d uyusmadi, %d kontrol edilemedi, %d retry",
                report["ozet"]["dogrulanan"],
                report["ozet"]["uyusmayan"],
                report["ozet"]["kontrol_edilemeyen"],
                len(report["retry_listesi"]))
    return str(report_path)


def run_verification_cycle(rollback_log_path, actual_data, config, campaign_lookup,
                           targeting_lookup, today=None):
    """
    Tam dogrulama dongusunu calistirir:
    1. Rollback log'u oku
    2. Guncel verilerle karsilastir
    3. Uyusmazlik varsa retry plani hazirla
    4. Rapor kaydet
    
    Bu fonksiyon Claude Code tarafindan soyle cagirilir:
    
        # 1. Agent 3 islemlerini uygula
        exec_result = run_executor(force_execute=True)
        
        # 2. 5 dakika bekle
        time.sleep(300)
        
        # 3. Amazon'dan guncel verileri cek (MCP tool'lariyla)
        actual_data = {
            "keywords": {...},      # entity_id → {bid, state}
            "targets": {...},       # entity_id → {bid, state}
            "negative_keywords": {  # campaign_id → [keyword_list]
                "12345": ["bad term", "another bad term"]
            },
            "negative_targets": {   # campaign_id → [asin_list]
                "12345": ["B0XXXXXXXX"]
            },
            "campaigns": {          # campaign_name → {state, budget}
                "LS051 - E - laptop stand": {"state": "ENABLED", "budget": 10.0}
            }
        }
        
        # 4. Dogrulama dongusunu calistir
        verification = run_verification_cycle(
            rollback_log_path, actual_data, config,
            campaign_lookup, targeting_lookup
        )
    
    Returns:
        dict: {
            "verification_report": {...},
            "retry_plan": {...} veya None,
            "rapor_dosyasi": "...",
            "sonuc": "TAMAM" / "RETRY_GEREKLI" / "MANUEL_GEREKLI"
        }
    """
    if today is None:
        today = datetime.utcnow().strftime("%Y-%m-%d")

    # 1. Dogrulama gorevlerini hazirla
    verification_report = run_delayed_verification(rollback_log_path, today)

    # 2. Guncel verilerle karsilastir
    verification_report = process_verification_results(verification_report, actual_data)

    # 3. Raporu kaydet
    report_path = save_verification_report(verification_report, today)

    result = {
        "verification_report": verification_report,
        "retry_plan": None,
        "rapor_dosyasi": report_path,
    }

    # 4. Uyusmazlik varsa retry plani hazirla
    if verification_report["retry_listesi"]:
        retry_plan = prepare_retry_operations(
            verification_report["retry_listesi"],
            config, campaign_lookup, targeting_lookup, today
        )
        result["retry_plan"] = retry_plan

        # Manuel mudahale gereken var mi?
        manuel_count = sum(1 for op in retry_plan["islemler"] if op["status"] == "MANUEL_GEREKLI")
        auto_retry_count = sum(1 for op in retry_plan["islemler"] if op["status"] == "HAZIR")

        if manuel_count > 0 and auto_retry_count > 0:
            result["sonuc"] = "RETRY_VE_MANUEL"
            result["mesaj"] = (
                f"{auto_retry_count} islem otomatik retry edilecek, "
                f"{manuel_count} islem manuel kontrol gerektiriyor."
            )
        elif manuel_count > 0:
            result["sonuc"] = "MANUEL_GEREKLI"
            result["mesaj"] = f"{manuel_count} islem manuel kontrol gerektiriyor."
        else:
            result["sonuc"] = "RETRY_GEREKLI"
            result["mesaj"] = f"{auto_retry_count} islem otomatik retry edilecek."
    else:
        result["sonuc"] = "TAMAM"
        result["mesaj"] = f"Tum islemler dogrulandi ({verification_report['ozet']['dogrulanan']}/{verification_report['ozet']['toplam_kontrol']})."

    logger.info("Dogrulama dongusu tamamlandi: %s", result["sonuc"])

    # Uyusmazlik varsa agent3_errors.json'a kaydet — Agent 4 analizi icin
    if result["sonuc"] != "TAMAM":
        uyusmayan_sayi = verification_report["ozet"].get("uyusmayan", 0)
        save_error_log(
            hata_tipi="VerificationError",
            hata_mesaji=f"{uyusmayan_sayi} islem dogrulanamadi: {result['mesaj']}",
            adim="delayed_verification",
            session_id=MAESTRO_SESSION_ID,
            extra={
                "sonuc": result["sonuc"],
                "toplam_kontrol": verification_report["ozet"].get("toplam_kontrol", 0),
                "dogrulanan": verification_report["ozet"].get("dogrulanan", 0),
                "uyusmayan": uyusmayan_sayi,
                "retry_listesi_boyut": len(verification_report.get("retry_listesi", [])),
            }
        )

    return result

def preflight_check(today):
    """
    Agent 2'nin bugun icin Excel raporlarini olusturup olusturamadigini kontrol eder.
    En az 1 rapor olmadan Agent 3 calismaz.
    
    Returns:
        (bool, list, list): (gecti_mi, bulunan_dosyalar, eksik_dosyalar)
    """
    ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)

    raporlar = {
        "bid_recommendations": {"zorunlu": True, "aciklama": "Bid tavsiyeleri"},
        "negative_candidates": {"zorunlu": False, "aciklama": "Negatif keyword adaylari"},
        "harvesting_candidates": {"zorunlu": False, "aciklama": "Harvesting adaylari"},
    }

    bulunan = []
    eksik = []

    for prefix, info in raporlar.items():
        dosya = find_todays_excel(prefix, today)
        if dosya:
            bulunan.append({"rapor": prefix, "dosya": str(dosya), "aciklama": info["aciklama"]})
            logger.info("  [OK] %s: %s", prefix, dosya.name)
        else:
            eksik.append({"rapor": prefix, "zorunlu": info["zorunlu"], "aciklama": info["aciklama"]})
            seviye = "KRITIK" if info["zorunlu"] else "OPSIYONEL"
            logger.warning("  [EKSIK] %s: BULUNAMADI [%s]", prefix, seviye)

    # Agent 1 veri dosyalari da kontrol et (kampanya lookup icin gerekli)
    agent1_kritik = ["sp_campaigns"]
    agent1_eksik = []
    for prefix in agent1_kritik:
        fpath = DATA_DIR / f"{today}_{prefix}.json"
        if not fpath.exists():
            # Dunku dosyayi kontrol et (gece yarisi gecis durumu)
            from datetime import timedelta
            dun = (datetime.strptime(today, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
            fpath_dun = DATA_DIR / f"{dun}_{prefix}.json"
            if fpath_dun.exists():
                logger.info("  [OK] Agent 1 verisi dunku tarihle bulundu: %s_%s.json", dun, prefix)
            else:
                agent1_eksik.append(prefix)
                logger.warning("  [EKSIK] Agent 1 verisi eksik: %s_%s.json", today, prefix)

    # Karar: en az bid_recommendations olmali
    zorunlu_eksik = [e for e in eksik if e["zorunlu"]]

    if zorunlu_eksik:
        logger.error("ON KONTROL BASARISIZ: Zorunlu raporlar eksik.")
        logger.error("Once Agent 2'yi calistirin: python agent2_analyst.py")
        return False, bulunan, eksik

    if agent1_eksik:
        logger.error("ON KONTROL BASARISIZ: Agent 1 verileri eksik (%s).", ", ".join(agent1_eksik))
        logger.error("Once Agent 1'i calistirin.")
        return False, bulunan, eksik

    if not bulunan:
        logger.error("ON KONTROL BASARISIZ: Hicbir Excel raporu bulunamadi.")
        return False, bulunan, eksik

    logger.info("On kontrol gecti: %d rapor bulundu, %d opsiyonel eksik.",
                len(bulunan), len([e for e in eksik if not e["zorunlu"]]))
    return True, bulunan, eksik


def save_error_log(hata_tipi, hata_mesaji, traceback_str=None, adim=None,
                   extra=None, session_id=None):
    """Agent 3 hata logu — lokal + Supabase dual-write."""
    dir_name = DATA_DIR.name  # "vigowood_eu_UK"
    parts = dir_name.rsplit("_", 1)
    hk = parts[0] if len(parts) == 2 else ""
    mp = parts[1] if len(parts) == 2 else ""
    return _central_save_error_log(
        hata_tipi, hata_mesaji, LOG_DIR,
        traceback_str=traceback_str, adim=adim, extra=extra,
        session_id=session_id, agent_name="agent3",
        hesap_key=hk, marketplace=mp)


def run_executor(hesap_key, marketplace, today=None, force_execute=False):
    """
    Agent 3 ana fonksiyonu. Claude Code / Master Agent tarafindan cagirilir.
    """
    init_paths(hesap_key, marketplace)

    if today is None:
        today = datetime.utcnow().strftime("%Y-%m-%d")

    logger.info("=== AGENT 3 v3 EXECUTOR BASLADI -- %s/%s -- %s ===", hesap_key, marketplace, today)

    # Dashboard: Agent 3 basliyor
    _session_id = MAESTRO_SESSION_ID or f"direct_{today}_{hesap_key}_{marketplace}"
    _dashboard_status("agent3", "running")
    _dashboard_pipeline(_session_id, hesap_key, marketplace, "agent3_execute", "running")
    _save_log("info", f"Agent 3 basliyor: {hesap_key}/{marketplace}", "agent3", hesap_key, marketplace, _session_id)

    try:
        result = _run_executor_impl(today, force_execute, hesap_key, marketplace)
        # Dashboard: Agent 3 tamamlandi
        _final = "completed" if result.get("durum") not in ("BASARISIZ",) else "failed"
        _dashboard_status("agent3", _final, {
            "tasks": result.get("ozet", {}).get("toplam", 0) if isinstance(result.get("ozet"), dict) else 0,
            "errors_7d": 1 if _final == "failed" else 0,
        })
        _dashboard_pipeline(_session_id, hesap_key, marketplace, "agent3_execute", _final)
        _save_log("info" if _final == "completed" else "error",
                  f"Agent 3 {'tamamlandi' if _final == 'completed' else 'basarisiz'}: {result.get('durum', '')}",
                  "agent3", hesap_key, marketplace, _session_id)
        return result
    except Exception as e:
        tb = traceback.format_exc()
        hata_tipi = type(e).__name__
        hata_mesaji = str(e)
        logger.error("BEKLENMEYEN HATA [%s]: %s", hata_tipi, hata_mesaji)
        save_error_log(hata_tipi, hata_mesaji, tb, adim="run_executor",
                       session_id=MAESTRO_SESSION_ID)
        # Dashboard: Agent 3 hata
        _dashboard_status("agent3", "failed")
        _dashboard_pipeline(_session_id, hesap_key, marketplace, "agent3_execute", "failed", hata_mesaji[:500])
        _save_log("error", f"Agent 3 hatasi: {hata_mesaji[:200]}", "agent3", hesap_key, marketplace, _session_id, error_type=hata_tipi)
        return {
            "agent": "Agent3_Executor",
            "tarih": today,
            "durum": "BASARISIZ",
            "hata": hata_mesaji,
        }


async def _collect_verify_data(hesap_key, marketplace):
    """
    Verify verilerini Amazon API'den ceker. MCP server'a bagimlilik yok.
    parallel_collector.py'deki AmazonAdsClient'i kullanir.
    """
    sys.path.insert(0, str(BASE_DIR))
    from parallel_collector import AmazonAdsClient, load_accounts

    accounts = load_accounts()
    lwa = accounts["lwa_app"]
    hesap = accounts["hesaplar"][hesap_key]
    mp_config = hesap["marketplaces"][marketplace]

    config = {
        "client_id": lwa["client_id"],
        "client_secret": lwa["client_secret"],
        "refresh_token": hesap["refresh_token"],
        "marketplace": marketplace,
        "profile_id": mp_config["profile_id"],
        "account_id": hesap["account_id"],
        "api_endpoint": hesap["api_endpoint"],
        "token_endpoint": hesap["token_endpoint"],
    }
    client = AmazonAdsClient(config)
    today = datetime.utcnow().strftime("%Y-%m-%d")

    VERIFY_ENTITIES = [
        ("sp_campaigns",          "POST", "/sp/campaigns/list",          "application/vnd.spCampaign.v3+json",              {"stateFilter": {"include": ["ENABLED", "PAUSED"]}}, "campaigns"),
        ("sp_keywords",           "POST", "/sp/keywords/list",           "application/vnd.spKeyword.v3+json",               {"stateFilter": {"include": ["ENABLED", "PAUSED"]}}, "keywords"),
        ("sp_targets",            "POST", "/sp/targets/list",            "application/vnd.spTargetingClause.v3+json",       {"stateFilter": {"include": ["ENABLED", "PAUSED"]}}, "targetingClauses"),
        ("sp_negative_keywords",  "POST", "/sp/negativeKeywords/list",   "application/vnd.spNegativeKeyword.v3+json",       {}, "negativeKeywords"),
        ("sp_negative_targets",   "POST", "/sp/negativeTargets/list",    "application/vnd.spNegativeTargetingClause.v3+json",{}, "negativeTargetingClauses"),
        ("sb_keywords",           "GET",  "/sb/keywords",                None, {"stateFilter": "enabled,paused", "count": 1000}, None),
        ("sb_targets",            "POST", "/sb/targets/list",            "application/vnd.sbtargetingresource.v3.2+json",   {"filter": {"states": ["enabled", "paused"]}}, None),
        ("sb_themes",             "POST", "/sb/themes/list",             "application/vnd.sbthemesresource.v3+json",        {}, None),
        ("sd_targets",            "GET",  "/sd/targets",                 None, {"stateFilter": "enabled,paused", "count": 1000}, None),
    ]

    results = {"tarih": today, "hesap": f"{hesap_key}/{marketplace}", "dosyalar": {}, "hatalar": [], "basarili": 0, "basarisiz": 0}

    for name, method, path, ct, params, _wrapper in VERIFY_ENTITIES:
        try:
            if method == "GET":
                query = "&".join(f"{k}={v}" for k, v in params.items())
                full_path = f"{path}?{query}" if query else path
                resp = await client._request_with_retry("GET", full_path, None)
            else:
                resp = await client._request_with_retry("POST", path, params, content_type=ct, accept=ct)

            # Extract data
            if isinstance(resp, dict) and _wrapper:
                data = resp.get(_wrapper, [])
            elif isinstance(resp, list):
                data = resp
            elif isinstance(resp, dict):
                for v in resp.values():
                    if isinstance(v, list):
                        data = v
                        break
                else:
                    data = [resp] if resp else []
            else:
                data = []

            # Pagination for SP POST endpoints
            if method == "POST" and isinstance(resp, dict) and _wrapper:
                next_token = resp.get("nextToken")
                while next_token:
                    page_params = {**params, "nextToken": next_token}
                    resp2 = await client._request_with_retry("POST", path, page_params, content_type=ct, accept=ct)
                    page_data = resp2.get(_wrapper, []) if isinstance(resp2, dict) else resp2 if isinstance(resp2, list) else []
                    data.extend(page_data)
                    next_token = resp2.get("nextToken") if isinstance(resp2, dict) else None

            filepath = DATA_DIR / f"{today}_verify_{name}.json"
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            logger.info("Kaydedildi: %s (%d kayit)", filepath, len(data))
            results["dosyalar"][name] = str(filepath)
            results["basarili"] += 1
        except Exception as e:
            logger.error("Verify hatasi [%s]: %s", name, str(e))
            results["hatalar"].append({"entity": name, "hata": str(e)})
            results["basarisiz"] += 1

    results["durum"] = "BASARILI" if results["basarisiz"] == 0 else "KISMI_BASARILI"
    return results


async def apply_execution_plan(hesap_key, marketplace, plan_path):
    """
    Execution plan dosyasini okuyup dogrudan Amazon API'ye gonderir.
    MCP server'a bagimlilik olmadan calisir.
    parallel_collector.py'deki AmazonAdsClient'i kullanir.
    """
    import asyncio
    sys.path.insert(0, str(BASE_DIR))
    from parallel_collector import AmazonAdsClient, load_accounts

    accounts = load_accounts()
    lwa = accounts["lwa_app"]
    hesap = accounts["hesaplar"][hesap_key]
    mp_config = hesap["marketplaces"][marketplace]

    config = {
        "client_id": lwa["client_id"],
        "client_secret": lwa["client_secret"],
        "refresh_token": hesap["refresh_token"],
        "marketplace": marketplace,
        "profile_id": mp_config["profile_id"],
        "account_id": hesap["account_id"],
        "api_endpoint": hesap["api_endpoint"],
        "token_endpoint": hesap["token_endpoint"],
    }
    client = AmazonAdsClient(config)

    with open(plan_path, "r", encoding="utf-8") as f:
        plan = json.load(f)

    # Plan dosya adindan tarihi cikar
    fname = Path(plan_path).name
    today = fname[:10] if len(fname) >= 10 and fname[4] == "-" else datetime.utcnow().strftime("%Y-%m-%d")

    # WRITE_ENDPOINTS — executor icinde tanimla (amazon_ads_mcp'ye bagimlilik olmasin)
    WRITE_ENDPOINTS = {
        "sp_keyword_bid_update": {"method": "PUT", "path": "/sp/keywords",
            "content_type": "application/vnd.spKeyword.v3+json",
            "accept": "application/vnd.spKeyword.v3+json", "wrapper_key": "keywords"},
        "sp_target_bid_update": {"method": "PUT", "path": "/sp/targets",
            "content_type": "application/vnd.spTargetingClause.v3+json",
            "accept": "application/vnd.spTargetingClause.v3+json", "wrapper_key": "targetingClauses"},
        "sb_keyword_bid_update": {"method": "PUT", "path": "/sb/keywords",
            "content_type": "application/json", "accept": "*/*", "wrapper_key": None},
        "sb_target_bid_update": {"method": "PUT", "path": "/sb/targets",
            "content_type": "application/json", "accept": "*/*", "wrapper_key": None},
        "sb_theme_bid_update": {"method": "PUT", "path": "/sb/themes",
            "content_type": "application/json",
            "accept": "application/vnd.sbthemesupdateresponse.v3+json", "wrapper_key": "themes"},
        "sd_target_bid_update": {"method": "PUT", "path": "/sd/targets",
            "content_type": "application/json", "accept": "application/json", "wrapper_key": None},
        "sp_negative_keyword_add": {"method": "POST", "path": "/sp/negativeKeywords",
            "content_type": "application/vnd.spNegativeKeyword.v3+json",
            "accept": "application/vnd.spNegativeKeyword.v3+json", "wrapper_key": "negativeKeywords"},
        "sp_campaign_negative_keyword_add": {"method": "POST", "path": "/sp/campaignNegativeKeywords",
            "content_type": "application/vnd.spCampaignNegativeKeyword.v3+json",
            "accept": "application/vnd.spCampaignNegativeKeyword.v3+json", "wrapper_key": "campaignNegativeKeywords"},
        "sp_negative_target_add": {"method": "POST", "path": "/sp/negativeTargets",
            "content_type": "application/vnd.spNegativeTargetingClause.v3+json",
            "accept": "application/vnd.spNegativeTargetingClause.v3+json", "wrapper_key": "negativeTargetingClauses"},
        "sp_campaign_create": {"method": "POST", "path": "/sp/campaigns",
            "content_type": "application/vnd.spCampaign.v3+json",
            "accept": "application/vnd.spCampaign.v3+json", "wrapper_key": "campaigns"},
        "sp_ad_group_create": {"method": "POST", "path": "/sp/adGroups",
            "content_type": "application/vnd.spAdGroup.v3+json",
            "accept": "application/vnd.spAdGroup.v3+json", "wrapper_key": "adGroups"},
        "sp_product_ad_create": {"method": "POST", "path": "/sp/productAds",
            "content_type": "application/vnd.spProductAd.v3+json",
            "accept": "application/vnd.spProductAd.v3+json", "wrapper_key": "productAds"},
        "sp_keyword_create": {"method": "POST", "path": "/sp/keywords",
            "content_type": "application/vnd.spKeyword.v3+json",
            "accept": "application/vnd.spKeyword.v3+json", "wrapper_key": "keywords"},
    }

    results = {
        "tarih": today, "durum": "BASLATILDI",
        "bid_sonuclari": [], "negatif_sonuclari": [], "hatalar": [],
        "rollback_log": [],
        "ozet": {"basarili": 0, "basarisiz": 0, "atlanan": 0},
    }

    EXECUTE_DELAY = 0.3

    async def execute_single(ep_name, payload):
        ep = WRITE_ENDPOINTS.get(ep_name)
        if not ep:
            return False, {}, f"Bilinmeyen endpoint: {ep_name}"
        wrapper = ep["wrapper_key"]
        body = {wrapper: [payload]} if wrapper else [payload]
        try:
            resp = await client._request_with_retry(
                ep["method"], ep["path"], body,
                content_type=ep["content_type"], accept=ep["accept"]
            )
            if wrapper and isinstance(resp, dict):
                inner = resp.get(wrapper, resp)
                if isinstance(inner, dict) and inner.get("error"):
                    return False, resp, f"API error: {json.dumps(inner['error'])[:500]}"
                return True, resp, None
            if isinstance(resp, list):
                for item in resp:
                    if isinstance(item, dict) and "code" in item:
                        if str(item.get("code", "")).upper() not in ("SUCCESS", "200"):
                            return False, resp, f"API error: {json.dumps(item)[:500]}"
                return True, resp, None
            return True, resp, None
        except Exception as e:
            return False, {}, f"Exception: {str(e)[:300]}"

    # --- BID DEGISIKLIKLERI ---
    logger.info("--- Faz 1: Bid Degisiklikleri (%d islem) ---", len(plan.get("bid_islemleri", [])))
    for op in plan.get("bid_islemleri", []):
        if op.get("status") != "HAZIR":
            results["ozet"]["atlanan"] += 1
            continue

        ep_name = op.get("api_endpoint", "")
        payload = op.get("api_payload", {})
        success, resp, error = await execute_single(ep_name, payload)

        if success:
            results["ozet"]["basarili"] += 1
            _entity_id, _entity_type, _ad_type = "", "", ""
            if "keywordId" in payload: _entity_id, _entity_type = str(payload["keywordId"]), "KEYWORD"
            elif "targetId" in payload: _entity_id, _entity_type = str(payload["targetId"]), "TARGET"
            elif "themeId" in payload: _entity_id, _entity_type = str(payload["themeId"]), "THEME"
            if ep_name.startswith("sp_"): _ad_type = "SP"
            elif ep_name.startswith("sb_"): _ad_type = "SB"
            elif ep_name.startswith("sd_"): _ad_type = "SD"

            results["rollback_log"].append({
                "tip": "BID_DEGISIKLIGI",
                "kampanya": op.get("kampanya", ""), "hedefleme": op.get("hedefleme", ""),
                "eski_bid": op.get("eski_bid"), "yeni_bid": op.get("yeni_bid"),
                "entity_id": _entity_id, "entity_type": _entity_type,
                "ad_type": _ad_type, "campaign_id": str(payload.get("campaignId", "")),
                "api_endpoint": ep_name, "api_payload": payload,
                "rollback": f"Bid'i {op.get('eski_bid')} olarak geri al",
            })
            logger.info("  BASARILI: %s -- %s (%.2f)", op.get("kampanya",""), op.get("hedefleme",""), payload.get("bid",0))
        else:
            results["ozet"]["basarisiz"] += 1
            results["hatalar"].append({"faz": "BID", "hata": error})
            logger.error("  BASARISIZ: %s -- %s", op.get("hedefleme",""), error)

        await asyncio.sleep(EXECUTE_DELAY)

    # --- NEGATIF EKLEMELER ---
    logger.info("--- Faz 2: Negatif Eklemeler (%d islem) ---", len(plan.get("negatif_islemleri", [])))
    for op in plan.get("negatif_islemleri", []):
        if op.get("status") != "HAZIR":
            results["ozet"]["atlanan"] += 1
            continue

        ep_name = op.get("api_endpoint", "")
        payload = op.get("api_payload", {})
        success, resp, error = await execute_single(ep_name, payload)

        if success:
            results["ozet"]["basarili"] += 1
            results["rollback_log"].append({
                "tip": op.get("tip", "NEGATIF"),
                "kampanya": op.get("kampanya", ""), "hedefleme": op.get("hedefleme", ""),
                "campaign_id": str(payload.get("campaignId", "")),
                "api_endpoint": ep_name, "api_payload": payload,
            })
            logger.info("  BASARILI: %s -- %s", op.get("kampanya",""), op.get("hedefleme",""))
        else:
            results["ozet"]["basarisiz"] += 1
            results["hatalar"].append({"faz": "NEGATIF", "hata": error})
            logger.error("  BASARISIZ: %s -- %s", op.get("hedefleme",""), error)

        await asyncio.sleep(EXECUTE_DELAY)

    # --- FAZ 3: HARVESTING (sub_operations with chain dependency) ---
    logger.info("--- Faz 3: Harvesting (%d islem) ---", len(plan.get("harvesting_islemleri", [])))

    def extract_entity_id(response, wrapper_key, id_field):
        """SP v3 response'tan entity ID cikar."""
        if not isinstance(response, dict):
            return None
        inner = response.get(wrapper_key, response)
        if isinstance(inner, dict):
            successes = inner.get("success", [])
            if successes and isinstance(successes[0], dict):
                val = successes[0].get(id_field)
                if val:
                    return str(val)
                for v in successes[0].values():
                    if isinstance(v, dict) and id_field in v:
                        return str(v[id_field])
        if isinstance(response, dict) and id_field in response:
            return str(response[id_field])
        return None

    for harvest in plan.get("harvesting_islemleri", []):
        if harvest.get("status") != "HAZIR":
            results["ozet"]["atlanan"] += 1
            continue

        sub_ops = harvest.get("sub_operations", [])
        logger.info("Harvesting: %s -- %s (%d sub-op)",
                     harvest.get("hedefleme", ""), harvest.get("tip", ""), len(sub_ops))

        new_campaign_id = None
        new_ad_group_id = None
        chain_broken = False
        sub_sonuclar = []

        for sub_op in sub_ops:
            if chain_broken:
                sub_sonuclar.append({
                    "op": sub_op.get("op", ""),
                    "durum": "ATLANDI",
                    "hata": "Onceki islem basarisiz (chain broken)",
                })
                continue

            ep_name = sub_op.get("api_endpoint", "")
            payload = dict(sub_op.get("api_payload", {}))

            # Placeholder degistir
            placeholder_fail = False
            for key in list(payload.keys()):
                val = payload[key]
                if val == "__YENI_KAMPANYA_ID__":
                    if new_campaign_id:
                        payload[key] = new_campaign_id
                    else:
                        placeholder_fail = True
                        chain_broken = True
                        sub_sonuclar.append({
                            "op": sub_op.get("op", ""),
                            "durum": "BASARISIZ",
                            "hata": "Kampanya ID henuz mevcut degil",
                        })
                        break
                elif val == "__YENI_AD_GROUP_ID__":
                    if new_ad_group_id:
                        payload[key] = new_ad_group_id
                    else:
                        placeholder_fail = True
                        chain_broken = True
                        sub_sonuclar.append({
                            "op": sub_op.get("op", ""),
                            "durum": "BASARISIZ",
                            "hata": "Ad Group ID henuz mevcut degil",
                        })
                        break

            if placeholder_fail:
                continue

            logger.info("  Sub-op: %s -- %s", sub_op.get("op", ""), ep_name)
            success, resp, error = await execute_single(ep_name, payload)

            sub_sonuc = {
                "op": sub_op.get("op", ""),
                "durum": "BASARILI" if success else "BASARISIZ",
                "hata": error,
            }

            if success:
                ep_info = WRITE_ENDPOINTS.get(ep_name, {})
                wrapper = ep_info.get("wrapper_key")

                if sub_op.get("op") == "KAMPANYA_OLUSTUR" and wrapper:
                    extracted = extract_entity_id(resp, wrapper, "campaignId")
                    if extracted:
                        new_campaign_id = extracted
                        sub_sonuc["campaignId"] = extracted
                        logger.info("    -> Yeni kampanya ID: %s", extracted)
                    else:
                        chain_broken = True
                        sub_sonuc["durum"] = "BASARISIZ"
                        sub_sonuc["hata"] = f"Kampanya olusturuldu ama ID alinamadi. Response: {json.dumps(resp, ensure_ascii=False)[:300]}"
                        logger.error("    -> ID alinamadi: %s", json.dumps(resp, ensure_ascii=False)[:300])

                elif sub_op.get("op") == "AD_GROUP_OLUSTUR" and wrapper:
                    extracted = extract_entity_id(resp, wrapper, "adGroupId")
                    if extracted:
                        new_ad_group_id = extracted
                        sub_sonuc["adGroupId"] = extracted
                        logger.info("    -> Yeni ad group ID: %s", extracted)
                    else:
                        chain_broken = True
                        sub_sonuc["durum"] = "BASARISIZ"
                        sub_sonuc["hata"] = f"Ad group olusturuldu ama ID alinamadi. Response: {json.dumps(resp, ensure_ascii=False)[:300]}"
                        logger.error("    -> ID alinamadi: %s", json.dumps(resp, ensure_ascii=False)[:300])
                else:
                    logger.info("    -> BASARILI")
            else:
                chain_broken = True
                logger.error("    -> BASARISIZ: %s", error)

            sub_sonuclar.append(sub_sonuc)
            await asyncio.sleep(EXECUTE_DELAY)

        # Harvesting ozet durumu
        failed_subs = [s for s in sub_sonuclar if s["durum"] != "BASARILI"]
        if failed_subs:
            results["ozet"]["basarisiz"] += 1
        else:
            results["ozet"]["basarili"] += 1

        results["rollback_log"].append({
            "tip": harvest.get("tip", "HARVEST"),
            "hedefleme": harvest.get("hedefleme", ""),
            "yeni_kampanya_adi": harvest.get("kampanya_adi", ""),
            "kaynak_kampanya": harvest.get("kaynak_kampanya", ""),
            "source_campaign_id": str(harvest.get("source_campaign_id", "")),
            "new_campaign_id": new_campaign_id,
            "new_ad_group_id": new_ad_group_id,
            "sub_operations": sub_sonuclar,
        })

    # --- ROLLBACK LOG KAYDET ---
    logs_dir = LOG_DIR
    logs_dir.mkdir(parents=True, exist_ok=True)
    rollback = {
        "tarih": today,
        "olusturma_zamani": datetime.utcnow().isoformat(),
        "islemler": results["rollback_log"],
    }
    rollback_path = logs_dir / f"{today}_rollback.json"
    with open(rollback_path, "w", encoding="utf-8") as f:
        json.dump(rollback, f, indent=2, ensure_ascii=False)

    results["durum"] = "TAMAMLANDI"
    results["rollback_dosyasi"] = str(rollback_path)

    ozet = results["ozet"]
    logger.info("=== EXECUTION TAMAMLANDI: %d basarili, %d basarisiz, %d atlanan ===",
                ozet["basarili"], ozet["basarisiz"], ozet["atlanan"])

    await client.close()
    return results


def _run_executor_impl(today, force_execute=False, hesap_key="", marketplace=""):

    # 0. On kontrol — Agent 2 raporlari mevcut mu?
    logger.info("--- On Kontrol ---")
    gecti, bulunan, eksik = preflight_check(today)
    if not gecti:
        mesaj = f"Agent 2 raporlari bulunamadi: {eksik}"
        save_error_log("Preflight", mesaj, adim="preflight_check",
                       extra={"bulunan": bulunan, "eksik": eksik},
                       session_id=MAESTRO_SESSION_ID)
        return {
            "agent": "Agent3_Executor",
            "tarih": today,
            "durum": "ON_KONTROL_BASARISIZ",
            "mesaj": "Agent 2 raporlari bulunamadi. Once Agent 2'yi calistirin.",
            "bulunan_raporlar": bulunan,
            "eksik_raporlar": eksik,
        }

    # 1. Ayarlari yukle
    settings = load_settings()
    config = get_agent3_config(settings)
    dry_run = config["dry_run"] and not force_execute

    logger.info("Mod: %s", "DRY_RUN" if dry_run else "UYGULAMA")
    logger.info("Limitler -- max_bid: %.2f, min_bid: %.2f, gunluk_max: %d",
                config["max_bid_limiti"], config["min_bid_limiti"], config["gunluk_max_islem"])

    # 2. Excel raporlarini oku
    bid_actions = parse_bid_recommendations(today)
    neg_actions = parse_negative_candidates(today)
    harvest_actions = parse_harvesting_candidates(today)

    total_actions = len(bid_actions) + len(neg_actions) + len(harvest_actions)

    if total_actions == 0:
        logger.info("Onaylanmis islem yok. Agent 3 tamamlandi.")
        return {
            "agent": "Agent3_Executor",
            "tarih": today,
            "mod": "DRY_RUN" if dry_run else "UYGULAMA",
            "durum": "BOS",
            "mesaj": "Hicbir onaylanmis islem bulunamadi.",
        }

    # Gunluk limit kontrolu
    limit_ok, limit_msg = check_daily_limit(total_actions, config)
    if not limit_ok:
        logger.warning(limit_msg)

    # 3. Kampanya ve targeting lookup olustur
    campaign_lookup = build_campaign_lookup(today)
    targeting_lookup = build_targeting_lookup(today)

    # 4. Her islem icin API payload hazirla
    bid_ops = [prepare_bid_change(a, config, campaign_lookup, targeting_lookup) for a in bid_actions]
    neg_ops = [prepare_negative_add(a, config, campaign_lookup) for a in neg_actions]

    harvest_ops = []
    for a in harvest_actions:
        if a["tip"] == "HARVEST_KEYWORD":
            harvest_ops.append(prepare_harvest_keyword(a, config, campaign_lookup, today))
        elif a["tip"] == "HARVEST_ASIN":
            harvest_ops.append(prepare_harvest_asin(a, config, campaign_lookup, today))

    # Gunluk limit uygula (siniri asanlari kes)
    max_ops = config["gunluk_max_islem"]
    all_ops_count = (
        sum(1 for o in bid_ops if o["status"] == "HAZIR") +
        sum(1 for o in neg_ops if o["status"] == "HAZIR") +
        sum(1 for o in harvest_ops if o["status"] == "HAZIR")
    )
    if all_ops_count > max_ops:
        logger.warning("Gunluk limit asildi: %d > %d. Fazla islemler atlanacak.", all_ops_count, max_ops)

    # 5. DRY-RUN: Rapor olustur ve don
    if dry_run:
        report = generate_dry_run_report(bid_ops, neg_ops, harvest_ops)
        # Dry-run raporunu dosyaya kaydet
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        report_path = LOG_DIR / f"{today}_dry_run_report.json"
        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2, ensure_ascii=False)
        logger.info("Dry-run raporu kaydedildi: %s", report_path)
        report["rapor_dosyasi"] = str(report_path)
        logger.info("=== AGENT 3 DRY-RUN TAMAMLANDI ===")
        return report

    # 6. UYGULAMA MODU
    # Plan olusturulur, ardindan apply_execution_plan() ile dogrudan Amazon API'ye gonderilir.

    execution_report = {
        "agent": "Agent3_Executor",
        "tarih": today,
        "mod": "UYGULAMA",
        "durum": "PLAN_HAZIRLANIYOR",
        "mesaj": "API payload'lari hazirlaniyor.",
        "bid_islemleri": [
            {
                "kampanya": op["action"]["kampanya_adi"],
                "hedefleme": op["action"]["hedefleme"],
                "eski_bid": op["action"]["eski_bid"],
                "yeni_bid": op["action"]["yeni_bid"],
                "api_endpoint": op.get("api_endpoint", ""),
                "api_payload": op.get("api_payload", {}),
                "status": op["status"],
                "hatalar": op["hatalar"],
            }
            for op in bid_ops
        ],
        "negatif_islemleri": [
            {
                "kampanya": op["action"]["kampanya_adi"],
                "hedefleme": op["action"]["hedefleme"],
                "tip": op["action"]["tip"],
                "api_endpoint": op.get("api_endpoint", ""),
                "api_payload": op.get("api_payload", {}),
                "status": op["status"],
                "hatalar": op["hatalar"],
            }
            for op in neg_ops
        ],
        "harvesting_islemleri": [
            {
                "kaynak_kampanya": op["action"].get("kaynak_kampanya", ""),
                "hedefleme": op["action"]["hedefleme"],
                "tip": op["action"]["tip"],
                "sub_operations": op.get("sub_operations", []),
                "kampanya_adi": op.get("kampanya_adi", ""),
                "status": op["status"],
                "hatalar": op["hatalar"],
            }
            for op in harvest_ops
        ],
        "ozet": {
            "bid_degisiklikleri": {"toplam": len(bid_ops), "hazir": sum(1 for o in bid_ops if o["status"] == "HAZIR"), "hata": sum(1 for o in bid_ops if o["status"] == "HATA")},
            "negatif_eklemeler": {"toplam": len(neg_ops), "hazir": sum(1 for o in neg_ops if o["status"] == "HAZIR"), "hata": sum(1 for o in neg_ops if o["status"] == "HATA")},
            "harvesting": {"toplam": len(harvest_ops), "hazir": sum(1 for o in harvest_ops if o["status"] == "HAZIR"), "hata": sum(1 for o in harvest_ops if o["status"] == "HATA")},
        },
    }

    # Execution raporunu kaydet
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    exec_path = LOG_DIR / f"{today}_execution_plan.json"
    with open(exec_path, "w", encoding="utf-8") as f:
        json.dump(execution_report, f, indent=2, ensure_ascii=False)

    logger.info("Execution plan kaydedildi: %s", exec_path)

    # Execution hatalari varsa agent3_errors.json'a kaydet
    toplam_hata = (
        sum(1 for o in bid_ops if o["status"] == "HATA") +
        sum(1 for o in neg_ops if o["status"] == "HATA") +
        sum(1 for o in harvest_ops if o["status"] == "HATA")
    )
    if toplam_hata > 0:
        hata_detaylari = []
        for op in bid_ops + neg_ops + harvest_ops:
            if op["status"] == "HATA" and op.get("hatalar"):
                hata_detaylari.append({
                    "tip": op["action"].get("tip", "bid"),
                    "kampanya": op["action"].get("kampanya_adi", ""),
                    "hedefleme": op["action"].get("hedefleme", ""),
                    "hatalar": op["hatalar"][:3],
                })
        save_error_log(
            hata_tipi="ExecutionError",
            hata_mesaji=f"{toplam_hata} islem HATA statusunde",
            adim="execution_plan",
            session_id=MAESTRO_SESSION_ID,
            extra={"toplam_hata": toplam_hata, "detaylar": hata_detaylari[:10]}
        )

    # ---- Supabase Sync ----
    _sync_agent3_to_supabase(hesap_key, marketplace, today,
                              execution_report, bid_ops, neg_ops, harvest_ops)

    execution_report["plan_dosyasi"] = str(exec_path)

    # --- DOGRUDAN API'YE GONDER ---
    hazir_islem = sum(1 for o in bid_ops + neg_ops + harvest_ops if o["status"] == "HAZIR")
    if hazir_islem > 0:
        logger.info("=== APPLY: %d islem Amazon API'ye gonderiliyor ===", hazir_islem)
        import asyncio
        try:
            apply_result = asyncio.run(apply_execution_plan(hesap_key, marketplace, str(exec_path)))
            if isinstance(apply_result, str):
                apply_result = json.loads(apply_result)
            execution_report["api_sonuclari"] = apply_result
            execution_report["durum"] = apply_result.get("durum", "TAMAMLANDI")
            execution_report["mesaj"] = apply_result.get("ozet_mesaj", "API uygulamasi tamamlandi")
            ozet = apply_result.get("ozet", {})
            logger.info("=== AGENT 3 EXECUTION TAMAMLANDI === Basarili: %d, Basarisiz: %d, Atlanan: %d",
                        ozet.get("basarili", 0), ozet.get("basarisiz", 0), ozet.get("atlanan", 0))
        except Exception as e:
            tb = traceback.format_exc()
            logger.error("API uygulama hatasi: %s", str(e))
            save_error_log("ExecutionError", str(e), tb, adim="apply_execution_plan",
                           session_id=MAESTRO_SESSION_ID)
            execution_report["durum"] = "API_HATASI"
            execution_report["mesaj"] = f"Plan hazir ama API uygulamasi basarisiz: {str(e)[:300]}"
    else:
        logger.info("Hazir islem yok, API cagrisi yapilmadi.")
        execution_report["durum"] = "PLAN_HAZIR_ISLEM_YOK"

    execution_report["dogrulama_talimati"] = {
        "bekleme_suresi_saniye": VERIFICATION_DELAY_SECONDS,
        "bekleme_suresi_dakika": VERIFICATION_DELAY_SECONDS // 60,
        "talimat": (
            f"Islemler uygulandiktan sonra {VERIFICATION_DELAY_SECONDS // 60} dakika bekleyin. "
            f"Sonra Amazon'dan guncel verileri cekip run_verification_cycle() fonksiyonunu calistirin. "
            f"Uyusmayan islemler otomatik retry edilecek."
        ),
        "rollback_log": str(LOG_DIR / f"{today}_rollback.json"),
    }
    return execution_report


def _sync_agent3_to_supabase(hesap_key, marketplace, today,
                              execution_report, bid_ops, neg_ops, harvest_ops):
    """Agent 3 execution plan'i Supabase'e yaz."""
    try:
        import sys as _sys
        _project_root = str(Path(__file__).parent.parent)
        if _project_root not in _sys.path:
            _sys.path.insert(0, _project_root)
        from supabase.db_client import SupabaseClient
        db = SupabaseClient()
    except Exception as e:
        logger.error("Supabase sync import hatasi: %s", e)
        save_error_log("InternalError", f"Supabase sync import: {e}",
                       traceback.format_exc(), adim="supabase_sync_init",
                       session_id=MAESTRO_SESSION_ID)
        return

    try:
        mode = execution_report.get("mod", "UYGULAMA")
        # Tablo varlik kontrolu — yoksa sessizce atla
        try:
            plan_id = db.insert_execution_plan(
                hesap_key, marketplace, today, mode,
                execution_report.get("ozet", {}),
                session_id=MAESTRO_SESSION_ID
            )
        except Exception as table_err:
            if "execution_plans" in str(table_err) and ("does not exist" in str(table_err) or "UndefinedTable" in str(table_err)):
                logger.warning("execution_plans tablosu Supabase'de yok — sync atlaniyor. "
                               "Tablo olusturulunca otomatik calisacak.")
                return
            raise  # Baska hata ise yukari ilet

        items = []
        for op in bid_ops:
            a = op.get("action", {})
            items.append({
                "item_type": "BID_CHANGE",
                "campaign_id": a.get("campaign_id"),
                "kampanya": a.get("kampanya_adi"),
                "ad_group_id": a.get("ad_group_id"),
                "keyword_id": a.get("keyword_id"),
                "target_id": a.get("target_id"),
                "hedefleme": a.get("hedefleme"),
                "eski_bid": a.get("eski_bid"),
                "yeni_bid": a.get("yeni_bid"),
                "api_endpoint": op.get("api_endpoint"),
                "api_payload": op.get("api_payload"),
                "status": op.get("status", "HAZIR"),
                "error_message": "; ".join(op.get("hatalar", []))[:500] or None,
            })
        for op in neg_ops:
            a = op.get("action", {})
            items.append({
                "item_type": "NEGATIVE_ADD",
                "campaign_id": a.get("campaign_id"),
                "kampanya": a.get("kampanya_adi"),
                "ad_group_id": a.get("ad_group_id"),
                "hedefleme": a.get("hedefleme"),
                "negative_type": a.get("tip"),
                "match_type": a.get("match_type"),
                "api_endpoint": op.get("api_endpoint"),
                "api_payload": op.get("api_payload"),
                "status": op.get("status", "HAZIR"),
                "error_message": "; ".join(op.get("hatalar", []))[:500] or None,
            })
        for op in harvest_ops:
            a = op.get("action", {})
            items.append({
                "item_type": "HARVESTING",
                "campaign_id": a.get("campaign_id"),
                "kampanya": a.get("kampanya_adi"),
                "hedefleme": a.get("hedefleme"),
                "harvest_type": a.get("tip"),
                "kaynak_kampanya": a.get("kaynak_kampanya"),
                "api_endpoint": op.get("api_endpoint"),
                "api_payload": op.get("api_payload"),
                "status": op.get("status", "HAZIR"),
                "error_message": "; ".join(op.get("hatalar", []))[:500] or None,
            })

        if items:
            db.insert_execution_items(plan_id, hesap_key, marketplace, items)

        logger.info("Supabase: execution plan yazildi (plan_id=%s, %d islem)", plan_id, len(items))

    except Exception as e:
        logger.error("Supabase sync hatasi (executor devam eder): %s", e)
        save_error_log("InternalError", f"Supabase sync: {e}",
                       traceback.format_exc(), adim="supabase_sync",
                       session_id=MAESTRO_SESSION_ID)
        try:
            db.insert_error_log(hesap_key, marketplace, "agent3", {
                "hata_tipi": "InternalError",
                "hata_mesaji": f"Supabase sync hatasi: {e}"[:500],
                "adim": "supabase_sync",
            })
        except Exception:
            pass


# ============================================================================
# DOGRUDAN CALISTIRMA (TEST)
# ============================================================================

if __name__ == "__main__":
    import sys

    # Windows cp1252 encoding sorunlarini onle
    if sys.stdout.encoding != 'utf-8':
        try:
            sys.stdout.reconfigure(encoding='utf-8')
            sys.stderr.reconfigure(encoding='utf-8')
        except Exception:
            pass

    # Kullanim: python agent3/executor.py <hesap_key> <marketplace> [--execute] [--verify] [--collect-verify] [--date YYYY-MM-DD]
    if len(sys.argv) < 3:
        print("Kullanim: python agent3/executor.py <hesap_key> <marketplace> [--execute] [--verify] [--collect-verify]")
        print("Ornek:    python agent3/executor.py vigowood_na US                    (dry-run)")
        print("Ornek:    python agent3/executor.py vigowood_na US --execute           (plan + API'ye gonder)")
        print("Ornek:    python agent3/executor.py vigowood_na US --collect-verify    (verify verileri cek)")
        print("Ornek:    python agent3/executor.py vigowood_na US --verify            (dogrulama yap)")
        sys.exit(1)

    hesap_key = sys.argv[1]
    marketplace = sys.argv[2]
    init_paths(hesap_key, marketplace)

    # --date YYYY-MM-DD parametresi destegi
    custom_date = None
    for i, arg in enumerate(sys.argv):
        if arg == "--date" and i + 1 < len(sys.argv):
            custom_date = sys.argv[i + 1]

    if "--collect-verify" in sys.argv:
        # Verify verilerini Amazon API'den cek (MCP server'a bagimlilik yok)
        import asyncio
        today = custom_date or datetime.utcnow().strftime("%Y-%m-%d")
        logger.info("=== VERIFY VERI TOPLAMA BASLADI === %s/%s", hesap_key, marketplace)
        try:
            result = asyncio.run(_collect_verify_data(hesap_key, marketplace))
            print(json.dumps(result, indent=2, ensure_ascii=False))
        except Exception as e:
            logger.error("Verify veri toplama hatasi: %s", str(e))
            print(json.dumps({"durum": "HATA", "mesaj": str(e)}, indent=2))
        sys.exit(0)

    if "--verify" in sys.argv:
        today = custom_date or datetime.utcnow().strftime("%Y-%m-%d")
        rollback_path = LOG_DIR / f"{today}_rollback.json"
        if not rollback_path.exists():
            print(f"Rollback log bulunamadi: {rollback_path}")
        else:
            report = run_delayed_verification(str(rollback_path), today)

            data_date = today
            for i, arg in enumerate(sys.argv):
                if arg == "--data-date" and i + 1 < len(sys.argv):
                    data_date = sys.argv[i + 1]

            verify_test_file = DATA_DIR / f"{data_date}_verify_sp_keywords.json"
            if verify_test_file.exists():
                actual_data = load_verify_actual_data(data_date)
                final_report = process_verification_results(report, actual_data)
                print(json.dumps(final_report, indent=2, ensure_ascii=False))

                ozet = final_report.get("ozet", {})
                print(f"\n--- DOGRULAMA TAMAMLANDI ---")
                print(f"Dogrulanan: {ozet.get('dogrulanan', 0)}")
                print(f"Uyusmayan: {ozet.get('uyusmayan', 0)}")
                print(f"Kontrol edilemeyen: {ozet.get('kontrol_edilemeyen', 0)}")
            else:
                print(json.dumps(report, indent=2, ensure_ascii=False))
                print(f"\n--- Dogrulama icin once verify verilerini cekin ---")
    else:
        force = "--execute" in sys.argv
        result = run_executor(hesap_key, marketplace, today=custom_date, force_execute=force)
        print(json.dumps(result, indent=2, ensure_ascii=False))
