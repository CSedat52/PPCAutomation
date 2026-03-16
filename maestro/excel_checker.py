"""
Maestro Agent — Excel Checker
================================
Agent 2'nin urettigi tavsiye Excel dosyalarindaki
Onay sutunlarinin doldurulup doldurulmadigini kontrol eder.
"""

import os
import logging
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("openpyxl gerekli: pip install openpyxl --break-system-packages")

from . import config

logger = logging.getLogger("maestro.excel")


# ============================================================================
# EXCEL ONAY KONTROL
# ============================================================================

# Agent 2'nin urettigi 3 Excel dosyasinin prefixleri
EXCEL_PREFIXES = [
    "bid_recommendations",
    "negative_candidates",
    "harvesting_candidates",
]

# Onay sutununun olasi isimleri
APPROVAL_COLUMN_NAMES = ["onay", "Onay", "ONAY", "approval", "Approval"]


def find_todays_excels(today=None):
    """
    Bugunun tarihiyle eslesen Agent 2 Excel dosyalarini bulur.
    
    Returns:
        dict: {prefix: filepath} — bulunan dosyalar
    """
    if today is None:
        today = datetime.utcnow().strftime(config.DATE_FORMAT)

    analysis_dir = Path(config.RECOMMENDATION_EXCEL_DIR)
    if not analysis_dir.exists():
        logger.warning("Analysis klasoru bulunamadi: %s", analysis_dir)
        return {}

    found = {}
    for prefix in EXCEL_PREFIXES:
        pattern = f"{today}_{prefix}"
        for f in analysis_dir.iterdir():
            if f.name.startswith(pattern) and f.suffix == ".xlsx":
                found[prefix] = str(f)
                break

    return found


def check_approval_status(today=None):
    """
    Tum tavsiye Excel'lerindeki Onay sutunlarini kontrol eder.
    
    Returns:
        dict: {
            "excels_found": int,
            "excels_checked": int,
            "total_rows": int,
            "approved_rows": int,
            "has_any_approval": bool,     — En az bir onay var mi
            "all_empty": bool,            — Tum onaylar bos mu
            "details": {prefix: {...}}    — Dosya bazli detay
        }
    """
    if today is None:
        today = datetime.utcnow().strftime(config.DATE_FORMAT)

    excels = find_todays_excels(today)

    result = {
        "excels_found": len(excels),
        "excels_checked": 0,
        "total_rows": 0,
        "approved_rows": 0,
        "has_any_approval": False,
        "all_empty": True,
        "details": {},
    }

    if not excels:
        logger.warning("Hicbir tavsiye Excel dosyasi bulunamadi (tarih: %s)", today)
        return result

    for prefix, filepath in excels.items():
        try:
            detail = _check_single_excel(filepath)
            result["details"][prefix] = detail
            result["excels_checked"] += 1
            result["total_rows"] += detail["total_rows"]
            result["approved_rows"] += detail["approved_rows"]

            if detail["approved_rows"] > 0:
                result["has_any_approval"] = True
                result["all_empty"] = False

        except Exception as e:
            logger.error("Excel kontrol hatasi (%s): %s", prefix, e)
            result["details"][prefix] = {
                "error": str(e),
                "total_rows": 0,
                "approved_rows": 0,
            }

    logger.info(
        "Excel kontrol sonucu: %d dosya, %d/%d satir onaylanmis, herhangi_onay=%s",
        result["excels_checked"],
        result["approved_rows"],
        result["total_rows"],
        result["has_any_approval"],
    )
    return result


def _check_single_excel(filepath):
    """
    Tek bir Excel dosyasindaki Onay sutununu kontrol eder.
    
    Returns:
        dict: {total_rows, approved_rows, approval_column_index, sample_approvals}
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return {"total_rows": 0, "approved_rows": 0, "approval_column_index": None}

    headers = [str(h).strip() if h else "" for h in rows[0]]

    # Onay sutununu bul
    approval_col = None
    for i, h in enumerate(headers):
        if h in APPROVAL_COLUMN_NAMES:
            approval_col = i
            break

    if approval_col is None:
        # Son sutun genellikle Onay
        approval_col = len(headers) - 1
        logger.info("Onay sutunu bulunamadi, son sutun kullaniliyor (index %d: '%s')",
                     approval_col, headers[approval_col])

    total_rows = 0
    approved_rows = 0
    sample_approvals = []

    for row in rows[1:]:
        total_rows += 1
        if approval_col < len(row):
            val = row[approval_col]
            if val is not None and str(val).strip() != "":
                approved_rows += 1
                if len(sample_approvals) < 5:
                    sample_approvals.append(str(val).strip())

    return {
        "total_rows": total_rows,
        "approved_rows": approved_rows,
        "approval_column_index": approval_col,
        "approval_column_name": headers[approval_col] if approval_col < len(headers) else "?",
        "sample_approvals": sample_approvals,
    }


def get_approval_summary(today=None):
    """Kullanici icin okunabilir onay ozeti doner."""
    status = check_approval_status(today)

    lines = []
    lines.append(f"Excel Dosyalari: {status['excels_found']} bulundu, {status['excels_checked']} kontrol edildi")
    lines.append(f"Toplam Satir: {status['total_rows']}")
    lines.append(f"Onaylanmis Satir: {status['approved_rows']}")

    for prefix, detail in status["details"].items():
        if "error" in detail:
            lines.append(f"  {prefix}: HATA — {detail['error']}")
        else:
            lines.append(f"  {prefix}: {detail['approved_rows']}/{detail['total_rows']} onaylanmis")

    return "\n".join(lines), status
